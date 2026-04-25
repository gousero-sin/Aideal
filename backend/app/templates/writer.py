"""Writer de templates — escrita não destrutiva em workbooks Excel.

Estratégia de preservação (RF-05, RNF-02):
- Abrir template com openpyxl preservando tudo (data_only=False)
- Escrever SOMENTE nas áreas de dados previstas
- Nunca recriar sheets, fórmulas, filtros ou layout
- Preservar partes OOXML visuais do template sem recriar layout
- Validar integridade estrutural após escrita
"""

import logging
import os
import re
import tempfile
import zipfile
from copy import deepcopy
from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook import Workbook

from ..config import settings
from .slicer_builder import (
    ResolvedSlicerSpec,
    build_drawing_slicer_anchor,
    build_slicer_cache_xml,
    build_slicers_xml,
    sanitize_identifier,
)

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class TableSpec:
    """Especificação de table OOXML a ser materializada no salvar final."""

    sheet_name: str
    table_name: str
    ref: str
    colunas: tuple[str, ...]
    style: str = "TableStyleMedium2"


@dataclass(frozen=True)
class SlicerSpec:
    """Especificação declarativa de slicer a ser injetado no pacote final."""

    table_name: str
    column_name: str
    caption: str
    sheet_destino: str = "DRE"
    cache_name: str | None = None
    slicer_name: str | None = None


class TemplateWriter:
    """Escrita controlada em templates Excel com preservação total de estrutura."""

    NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    NS_REL_PKG = "http://schemas.openxmlformats.org/package/2006/relationships"
    NS_MC = "http://schemas.openxmlformats.org/markup-compatibility/2006"
    NS_X14 = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
    NS_X14AC = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"
    NS_XR = "http://schemas.microsoft.com/office/spreadsheetml/2014/revision"
    NS_XR2 = "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2"
    NS_XR3 = "http://schemas.microsoft.com/office/spreadsheetml/2016/revision3"
    REL_TABLE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/table"
    REL_PRINTER = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/printerSettings"
    REL_SLICER = "http://schemas.microsoft.com/office/2007/relationships/slicer"
    REL_SLICER_CACHE = "http://schemas.microsoft.com/office/2007/relationships/slicerCache"
    EXT_URI_SLICER_LIST_X15 = "{3A4CF648-6AED-40f4-86FF-DC5316D8AED3}"
    EXT_URI_SLICER_CACHES_X15 = "{46BE6895-7355-4a93-B00E-2C351335B9C9}"
    NS_X15 = "http://schemas.microsoft.com/office/spreadsheetml/2010/11/main"
    SHEETS_MERGE_VISUAL = {"DRE", "BD_FLUXO"}

    def __init__(self, template_path: Path):
        self.template_path = Path(template_path)
        self._wb: Workbook | None = None
        self._original_sheets: list[str] = []
        self._original_defined_names: list[str] = []
        self._original_tables: list[str] = []
        self._modified_sheets: set[str] = set()
        # name -> novo ref (ex.: "BD_FLUXO1" -> "A1:R1851")
        self._table_refs_override: dict[str, str] = {}
        # tables com colunas calculadas do template que devem ser removidas
        # do OOXML final porque os valores foram materializados pelo backend.
        self._table_strip_calculated_formulas: set[str] = set()
        # Tabelas novas registradas em runtime para materialização no pacote final.
        # table_name -> TableSpec
        self._table_specs: dict[str, TableSpec] = {}
        # Slicers declarados para reconstrução sobre table OOXML.
        self._slicer_specs: list[SlicerSpec] = []
        self._slicers_novos_ativos = False
        # Modo de tratamento de pivot cache records no pacote final:
        # - None: preserva template
        # - remove: remove records + referências
        # - empty: mantém records com count=0 e sem itens
        self._pivot_cache_records_mode: str | None = None
        # Se True, limpa <items> dos slicerCaches para forçar rebuild no Excel.
        self._clear_slicer_cache_items = False
        # Se True, remove slicers/slicerCaches do pacote final e limpa referências.
        self._strip_slicers = False
        # Se True, aplica modo "excel safe": remove charts, pivots, drawings de abas
        # hidden (Painel/APOIO), printerSettings, themeOverride e customXml, limpando
        # todas as referências (Content_Types, workbook, sheet rels). Preserva apenas
        # a logo da DRE (drawing2) e as tables fundamentais.
        self._excel_safe_mode = False
        # Sheets que devem manter o XML do template e trocar apenas <sheetData>.
        self._sheet_data_overrides: set[str] = set()
        # sheet_name -> {col_idx_1based: hidden_bool} — patch direto no XML do template
        # (usado para DRE; evita reserialização do openpyxl que quebra slicers).
        self._cols_patch_only: dict[str, dict[int, bool]] = {}
        # sheet_name -> XML string do <sheetData>…</sheetData> construído
        # externamente em XML puro (bypassa openpyxl inteiramente para a sheet).
        self._sheetdata_xml_puro: dict[str, str] = {}

    def abrir(self) -> None:
        """Abre o template para edição preservando toda a estrutura."""
        if not self.template_path.exists():
            raise FileNotFoundError(f"Template não encontrado: {self.template_path}")

        self._wb = load_workbook(
            self.template_path,
            keep_vba=False,
            keep_links=True,
            data_only=False,
        )
        self._original_sheets = list(self._wb.sheetnames)
        self._original_defined_names = list(self._wb.defined_names.keys())
        self._original_tables = self._listar_tabelas()

        logger.info(
            f"Template aberto: {self.template_path.name} "
            f"({len(self._original_sheets)} sheet(s))"
        )

    def listar_sheets(self) -> list[str]:
        """Retorna lista de sheets do template."""
        if not self._wb:
            raise RuntimeError("Template não aberto. Chame abrir() primeiro.")
        return list(self._wb.sheetnames)

    def obter_info_sheet(self, sheet_name: str) -> dict:
        """Retorna informações sobre uma sheet específica.

        Returns:
            dict com: min_row, max_row, min_col, max_col, merged_cells, auto_filter
        """
        if not self._wb:
            raise RuntimeError("Template não aberto.")

        ws = self._wb[sheet_name]
        return {
            "nome": sheet_name,
            "min_row": ws.min_row,
            "max_row": ws.max_row,
            "min_col": ws.min_column,
            "max_col": ws.max_column,
            "merged_cells": [str(mc) for mc in ws.merged_cells.ranges],
            "auto_filter": str(ws.auto_filter.ref) if ws.auto_filter.ref else None,
            "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
        }

    def escrever_celula(self, sheet_name: str, row: int, col: int, valor) -> None:
        """Escreve valor em uma célula específica."""
        if not self._wb:
            raise RuntimeError("Template não aberto.")
        ws = self._wb[sheet_name]
        ws.cell(row=row, column=col, value=valor)
        self._modified_sheets.add(sheet_name)

    def escrever_area(
        self, sheet_name: str, dados: list[list], linha_inicio: int, coluna_inicio: int
    ) -> int:
        """Escreve uma área retangular de dados no template.

        Args:
            sheet_name: nome da sheet
            dados: lista de listas (linhas x colunas)
            linha_inicio: linha inicial (1-based)
            coluna_inicio: coluna inicial (1-based)

        Returns:
            Número de linhas escritas
        """
        if not self._wb:
            raise RuntimeError("Template não aberto.")

        ws = self._wb[sheet_name]

        for i, linha in enumerate(dados):
            for j, valor in enumerate(linha):
                ws.cell(
                    row=linha_inicio + i,
                    column=coluna_inicio + j,
                    value=valor,
                )
        self._modified_sheets.add(sheet_name)

        logger.debug(
            f"Área escrita em '{sheet_name}': {len(dados)} linha(s) "
            f"a partir de ({linha_inicio}, {coluna_inicio})"
        )
        return len(dados)

    def limpar_area(
        self,
        sheet_name: str,
        linha_inicio: int,
        linha_fim: int,
        coluna_inicio: int,
        coluna_fim: int,
    ) -> None:
        """Limpa os valores de uma área retangular sem mexer na formatação."""
        if not self._wb:
            raise RuntimeError("Template não aberto.")

        ws = self._wb[sheet_name]
        for row in range(linha_inicio, linha_fim + 1):
            for col in range(coluna_inicio, coluna_fim + 1):
                ws.cell(row=row, column=col).value = None
        self._modified_sheets.add(sheet_name)

    def ajustar_tabela_range(
        self,
        sheet_name: str,
        table_name: str,
        linha_fim: int,
        coluna_fim: int | None = None,
    ) -> None:
        """Ajusta o ref da tabela nomeada para refletir o volume real de dados.

        Necessário quando a tabela é fonte de pivotCache: ref divergente faz o
        pivot/slicer ler linhas vazias após o refresh.
        """
        if not self._wb:
            raise RuntimeError("Template não aberto.")

        ws = self._wb[sheet_name]
        tables = getattr(ws, "tables", {})
        tbl = tables.get(table_name)
        if tbl is None:
            logger.warning("Tabela '%s' não encontrada em '%s'", table_name, sheet_name)
            return

        from openpyxl.utils import range_boundaries, get_column_letter

        min_col, min_row, max_col, _ = range_boundaries(tbl.ref)
        if coluna_fim is not None:
            max_col = coluna_fim
        novo_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{linha_fim}"
        if novo_ref == tbl.ref:
            return

        tbl.ref = novo_ref
        if getattr(tbl, "autoFilter", None) is not None:
            tbl.autoFilter.ref = novo_ref
        self._modified_sheets.add(sheet_name)
        # Registra override para patch direto no XML do template (pois o pacote
        # final usa tables/*.xml do template, não da saída do openpyxl).
        self._table_refs_override[table_name] = novo_ref
        logger.info("Tabela '%s' ajustada: ref=%s", table_name, novo_ref)

    def remover_formulas_calculadas_tabela(self, table_name: str) -> None:
        """Remove metadados de calculatedColumnFormula da tabela no OOXML final.

        Use quando o backend materializa os valores das colunas derivadas e a
        definição calculada herdada do template passa a ficar inconsistente.
        """
        self._table_strip_calculated_formulas.add(table_name)

    def registrar_table_ooxml(
        self,
        sheet_name: str,
        table_name: str,
        ref: str,
        colunas: list[str],
        style: str = "TableStyleMedium2",
    ) -> None:
        """Registra criação de table OOXML para aplicar no salvar.

        Usado para promover áreas de dados (ex.: DETALHE_MENSAL_DB granular)
        para `tableParts` reais no pacote final.
        """
        if not self._wb:
            raise RuntimeError("Template não aberto.")
        if sheet_name not in self._wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' não encontrada para registrar table.")
        if not table_name:
            raise ValueError("table_name é obrigatório.")
        if not ref:
            raise ValueError("ref é obrigatório (ex.: A1:Q200).")
        if not colunas:
            raise ValueError("Lista de colunas da table não pode ser vazia.")

        from openpyxl.utils import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(ref)
        if max_row <= min_row:
            raise ValueError("Table precisa conter cabeçalho e ao menos uma linha de dados.")
        total_cols = max_col - min_col + 1
        if total_cols != len(colunas):
            raise ValueError(
                "Quantidade de colunas não bate com o range da table: "
                f"range={total_cols}, colunas={len(colunas)}"
            )

        # Garante que não colidimos com tabela existente em outra aba.
        for ws in self._wb.worksheets:
            if table_name in ws.tables and ws.title != sheet_name:
                raise ValueError(
                    f"Table '{table_name}' já existe na aba '{ws.title}'."
                )

        self._table_specs[table_name] = TableSpec(
            sheet_name=sheet_name,
            table_name=table_name,
            ref=ref,
            colunas=tuple(colunas),
            style=style,
        )
        self._modified_sheets.add(sheet_name)
        logger.info(
            "Table registrada para criação: %s!%s (%s)",
            sheet_name,
            table_name,
            ref,
        )

    def registrar_slicer(
        self,
        *,
        table_name: str,
        column_name: str,
        caption: str | None = None,
        sheet_destino: str = "DRE",
        cache_name: str | None = None,
        slicer_name: str | None = None,
    ) -> None:
        """Registra slicer para reconstrução sobre table OOXML.

        A injeção OOXML é feita após o merge final do pacote para evitar que
        o openpyxl descarte `extLst` de slicers.
        """
        if not self._wb:
            raise RuntimeError("Template não aberto.")
        if table_name not in self._table_specs and not any(
            table_name in ws.tables for ws in self._wb.worksheets
        ):
            raise ValueError(f"Table '{table_name}' não encontrada para slicer.")
        if sheet_destino not in self._wb.sheetnames:
            raise ValueError(f"Sheet destino '{sheet_destino}' não existe.")
        if not column_name:
            raise ValueError("column_name é obrigatório.")

        label = caption or column_name
        self._slicer_specs.append(
            SlicerSpec(
                table_name=table_name,
                column_name=column_name,
                caption=label,
                sheet_destino=sheet_destino,
                cache_name=cache_name,
                slicer_name=slicer_name,
            )
        )
        self._slicers_novos_ativos = True

    def remover_pivot_cache_records(self) -> None:
        """Remove snapshots pivotCacheRecords para forçar rebuild no Excel."""
        self._pivot_cache_records_mode = "remove"

    def esvaziar_pivot_cache_records(self) -> None:
        """Mantém pivotCacheRecords no pacote, mas com snapshot vazio."""
        self._pivot_cache_records_mode = "empty"

    def limpar_slicer_cache_items(self) -> None:
        """Limpa itens de slicer caches para reconstrução no Excel."""
        self._clear_slicer_cache_items = True

    def remover_slicers(self) -> None:
        """Remove artefatos de slicer/slicerCache no pacote final."""
        self._strip_slicers = True

    def ativar_modo_excel_safe(self) -> None:
        """Ativa modo Excel-safe: remove charts, pivots, drawings de abas hidden,
        printerSettings, themeOverride e customXml para garantir abertura sem reparo
        no Excel Desktop. Preserva DRE, BD_FLUXO, DETALHE_MENSAL_DB e a logo.
        """
        self._excel_safe_mode = True
        self._strip_slicers = True

    def registrar_sheet_data_override(
        self,
        sheet_name: str,
        dados: list[list],
        linha_inicio: int,
        coluna_inicio: int,
        limpar_area: tuple[int, int, int, int] | None = None,
    ) -> int:
        """Registra escrita em sheet com override final de apenas <sheetData>.

        Usa openpyxl para popular os valores em memória e, no salvar final, aplica
        patch no XML da sheet do template trocando somente o bloco <sheetData>.
        """
        if limpar_area:
            linha_i, linha_f, col_i, col_f = limpar_area
            self.limpar_area(sheet_name, linha_i, linha_f, col_i, col_f)

        linhas = self.escrever_area(
            sheet_name=sheet_name,
            dados=dados,
            linha_inicio=linha_inicio,
            coluna_inicio=coluna_inicio,
        )
        self._sheet_data_overrides.add(sheet_name)
        return linhas

    def definir_colunas_ocultas(
        self,
        sheet_name: str,
        colunas: list[int | str],
        ocultar: bool,
    ) -> None:
        """Define visibilidade de colunas sem alterar conteúdo.

        Args:
            sheet_name: nome da sheet
            colunas: lista de índices (1-based int) ou letras de coluna ('A', 'B', ...)
            ocultar: True para ocultar, False para exibir
        """
        if not self._wb:
            raise RuntimeError("Template não aberto.")
        if not colunas:
            return

        from openpyxl.utils import get_column_letter

        ws = self._wb[sheet_name]
        for col in colunas:
            letra = get_column_letter(col) if isinstance(col, int) else str(col).upper()
            ws.column_dimensions[letra].hidden = ocultar
        self._modified_sheets.add(sheet_name)

    def substituir_sheetdata_xml_puro(self, sheet_name: str, sheetdata_xml: str) -> None:
        """Registra substituição de <sheetData> por XML construído externamente.

        Substitui apenas o bloco <sheetData>…</sheetData> no XML do template,
        sem passar pelo openpyxl. Style IDs, row attrs e rels do template
        são preservados integralmente. A sheet NÃO é adicionada a _modified_sheets.
        """
        self._sheetdata_xml_puro[sheet_name] = sheetdata_xml

    def ocultar_colunas_xml_patch(
        self,
        sheet_name: str,
        colunas: list[int | str],
        ocultar: bool,
    ) -> None:
        """Registra patch de ocultação de colunas aplicado direto no XML do template.

        Alternativa a `definir_colunas_ocultas` para abas onde a reserialização
        do openpyxl quebra artefatos visuais (ex.: slicers na aba DRE).
        O patch é aplicado diretamente nos bytes do template durante o salvar.
        """
        if not colunas:
            return
        from openpyxl.utils import column_index_from_string, get_column_letter

        mapa = self._cols_patch_only.setdefault(sheet_name, {})
        ws = self._wb[sheet_name] if self._wb else None
        for col in colunas:
            if isinstance(col, int):
                idx = col
            else:
                idx = column_index_from_string(str(col).upper())
            mapa[idx] = ocultar
            if ws is not None:
                # Mantém o estado em memória sincronizado para validações e
                # chamadas subsequentes, sem forçar a aba para o fluxo de merge
                # do openpyxl que hoje é evitado para preservar slicers.
                ws.column_dimensions[get_column_letter(idx)].hidden = ocultar

    def definir_linhas_ocultas(
        self,
        sheet_name: str,
        linha_inicio: int,
        linha_fim: int,
        ocultar: bool,
    ) -> None:
        """Define visibilidade de um intervalo de linhas sem alterar conteúdo."""
        if not self._wb:
            raise RuntimeError("Template não aberto.")
        if linha_inicio > linha_fim:
            return

        ws = self._wb[sheet_name]
        for row in range(linha_inicio, linha_fim + 1):
            ws.row_dimensions[row].hidden = ocultar
        self._modified_sheets.add(sheet_name)

    @staticmethod
    def _normalizar_target_workbook_rel(target: str) -> str:
        target = target.strip()
        if target.startswith("/"):
            target = target[1:]
        if target.startswith("xl/"):
            return target

        # workbook rels costumam ser relativos a xl/workbook.xml
        normalizado = os.path.normpath(f"xl/{target}").replace("\\", "/")
        return normalizado

    @classmethod
    def _mapear_sheets_para_partes(cls, xlsx_path: Path) -> dict[str, str]:
        with zipfile.ZipFile(xlsx_path, "r") as zf:
            wb_xml = ET.fromstring(zf.read("xl/workbook.xml"))
            rels_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))

        rel_map: dict[str, str] = {}
        for rel in rels_xml.findall(f"{{{cls.NS_REL_PKG}}}Relationship"):
            rel_id = rel.attrib.get("Id")
            target = rel.attrib.get("Target")
            if rel_id and target:
                rel_map[rel_id] = cls._normalizar_target_workbook_rel(target)

        sheet_parts: dict[str, str] = {}
        for sheet in wb_xml.findall(f".//{{{cls.NS_MAIN}}}sheet"):
            nome = sheet.attrib.get("name")
            rid = sheet.attrib.get(f"{{{cls.NS_REL}}}id")
            if not nome or not rid:
                continue
            target = rel_map.get(rid)
            if target:
                sheet_parts[nome] = target

        return sheet_parts

    @staticmethod
    def _path_rel_sheet(part_sheet: str) -> str:
        p = Path(part_sheet)
        return str(p.parent / "_rels" / f"{p.name}.rels").replace("\\", "/")

    @staticmethod
    def _serializar_xml(root: ET.Element) -> bytes:
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    @classmethod
    def _registrar_namespaces_xml(cls) -> None:
        """Registra namespaces usados pelo template para serialização estável."""
        ET.register_namespace("", cls.NS_MAIN)
        ET.register_namespace("r", cls.NS_REL)
        ET.register_namespace("mc", cls.NS_MC)
        ET.register_namespace("x14", cls.NS_X14)
        ET.register_namespace("x14ac", cls.NS_X14AC)
        ET.register_namespace("xr", cls.NS_XR)
        ET.register_namespace("xr2", cls.NS_XR2)
        ET.register_namespace("xr3", cls.NS_XR3)

    @staticmethod
    def _remover_ref_calc_chain_workbook_rels(xml_bytes: bytes) -> bytes:
        """Remove referência ao calcChain do workbook.xml.rels (ou do workbook.xml).

        calcChain.xml é omitido do pacote final (dados novos invalidam a chain);
        manter a referência com o arquivo ausente gera prompt de reparo no Excel.
        """
        try:
            text = xml_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return xml_bytes
        updated = re.sub(
            r'<[A-Za-z0-9]*:?Relationship\b[^/>]*Type="[^"]*/calcChain"[^/>]*/>\s*',
            "",
            text,
        )
        if updated == text:
            return xml_bytes
        return updated.encode("utf-8")

    def _remover_rel_pivot_cache_records(self, rels_bytes: bytes) -> bytes:
        """Remove relationship de pivotCacheRecords de um arquivo .rels.

        Usa manipulação por regex para preservar os prefixos/namespaces
        originais do arquivo (ET reescreveria como `ns0:` e o Excel trata
        divergência de prefix como inconsistência).
        """
        try:
            rels_xml = rels_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return rels_bytes

        pattern = r'<[A-Za-z0-9]*:?Relationship\b[^/>]*Type="[^"]*/pivotCacheRecords"[^/>]*/>'
        updated = re.sub(pattern, "", rels_xml)
        if updated == rels_xml:
            return rels_bytes
        return updated.encode("utf-8")

    @staticmethod
    def _remover_content_type_pivot_cache_records(content_types_bytes: bytes) -> bytes:
        """Remove overrides de pivotCacheRecords do [Content_Types].xml."""
        try:
            content_types_xml = content_types_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return content_types_bytes

        pattern = (
            r'<Override\b[^>]*PartName=(?:"|\')/xl/pivotCache/pivotCacheRecords[^"\']+\.xml'
            r'(?:"|\')[^>]*/>'
        )
        updated = re.sub(pattern, "", content_types_xml)

        if updated == content_types_xml:
            return content_types_bytes

        return updated.encode("utf-8")

    @staticmethod
    def _is_slicer_part(nome_parte: str) -> bool:
        return nome_parte.startswith("xl/slicerCaches/") or nome_parte.startswith("xl/slicers/")

    @staticmethod
    def _remover_rels_por_tipo_fragmento(rels_bytes: bytes, tipo_fragmento: str) -> bytes:
        """Remove Relationship(s) cujo Type contenha o fragmento informado."""
        try:
            rels_xml = rels_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return rels_bytes

        pattern = (
            r'<[A-Za-z0-9]*:?Relationship\b[^/>]*Type=(?:"|\')[^"\']*'
            + re.escape(tipo_fragmento)
            + r'[^"\']*(?:"|\')[^>]*/>\s*'
        )
        updated = re.sub(pattern, "", rels_xml)
        if updated == rels_xml:
            return rels_bytes
        return updated.encode("utf-8")

    @staticmethod
    def _remover_content_type_slicer_parts(content_types_bytes: bytes) -> bytes:
        """Remove overrides de slicers/slicerCaches do [Content_Types].xml."""
        try:
            content_types_xml = content_types_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return content_types_bytes

        pattern = (
            r'<Override\b[^>]*PartName=(?:"|\')/xl/(?:slicerCaches/slicerCache|slicers/slicer)'
            r'[^"\']+\.xml(?:"|\')[^>]*/>\s*'
        )
        updated = re.sub(pattern, "", content_types_xml)
        if updated == content_types_xml:
            return content_types_bytes
        return updated.encode("utf-8")

    @staticmethod
    def _remover_slicer_ext_workbook_xml(workbook_xml_bytes: bytes) -> bytes:
        """Remove bloco de slicerCaches do workbook.xml (extLst/x14:slicerCaches)."""
        try:
            workbook_xml = workbook_xml_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return workbook_xml_bytes

        updated = re.sub(
            r"<ext\b[^>]*>\s*<x14:slicerCaches>.*?</x14:slicerCaches>\s*</ext>\s*",
            "",
            workbook_xml,
            flags=re.DOTALL,
        )
        if updated == workbook_xml:
            return workbook_xml_bytes
        return updated.encode("utf-8")

    # --- Excel Safe Mode ------------------------------------------------
    @staticmethod
    def _is_excel_safe_drop(nome_parte: str) -> bool:
        """Retorna True se a parte deve ser removida no modo Excel-safe."""
        prefixos_drop = (
            "xl/charts/",
            "xl/pivotCache/",
            "xl/pivotTables/",
            "xl/printerSettings/",
            "customXml/",
        )
        if nome_parte.startswith(prefixos_drop):
            return True
        if nome_parte.startswith("xl/theme/themeOverride"):
            return True
        # drawing1 (Painel) e drawing3 (APOIO) — drawing2 (logo DRE) preserva
        if nome_parte in (
            "xl/drawings/drawing1.xml",
            "xl/drawings/drawing3.xml",
            "xl/drawings/_rels/drawing1.xml.rels",
            "xl/drawings/_rels/drawing3.xml.rels",
        ):
            return True
        # imagens usadas só por Painel/APOIO
        if nome_parte in (
            "xl/media/image1.png",
            "xl/media/image1.jpeg",
            "xl/media/image3.png",
            "xl/media/image4.svg",
        ):
            return True
        return False

    @staticmethod
    def _remover_pivot_caches_workbook_xml(workbook_xml_bytes: bytes) -> bytes:
        """Remove bloco <pivotCaches>...</pivotCaches> do workbook.xml."""
        try:
            xml = workbook_xml_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return workbook_xml_bytes
        updated = re.sub(r"<pivotCaches\b[^>]*>.*?</pivotCaches>\s*", "", xml, flags=re.DOTALL)
        updated = re.sub(r"<pivotCaches\b[^/]*/>\s*", "", updated)
        if updated == xml:
            return workbook_xml_bytes
        return updated.encode("utf-8")

    @staticmethod
    def _excel_safe_limpar_workbook_rels(rels_bytes: bytes) -> bytes:
        """Remove relationships de pivotCacheDefinition no workbook.xml.rels."""
        try:
            xml = rels_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return rels_bytes
        updated = re.sub(
            r"<Relationship\b[^>]*Type=\"[^\"]*/pivotCacheDefinition\"[^>]*/>\s*",
            "",
            xml,
        )
        if updated == xml:
            return rels_bytes
        return updated.encode("utf-8")

    @staticmethod
    def _excel_safe_limpar_content_types(content_types_bytes: bytes) -> bytes:
        """Remove Overrides de chart, pivot, drawing1, drawing3, printerSettings, themeOverride."""
        try:
            xml = content_types_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return content_types_bytes
        padroes = [
            r'<Override\b[^>]*PartName="/xl/charts/[^"]+"[^>]*/>\s*',
            r'<Override\b[^>]*PartName="/xl/pivotTables/[^"]+"[^>]*/>\s*',
            r'<Override\b[^>]*PartName="/xl/pivotCache/[^"]+"[^>]*/>\s*',
            r'<Override\b[^>]*PartName="/xl/printerSettings/[^"]+"[^>]*/>\s*',
            r'<Override\b[^>]*PartName="/xl/theme/themeOverride[^"]*"[^>]*/>\s*',
            r'<Override\b[^>]*PartName="/xl/drawings/drawing1\.xml"[^>]*/>\s*',
            r'<Override\b[^>]*PartName="/xl/drawings/drawing3\.xml"[^>]*/>\s*',
            r'<Override\b[^>]*PartName="/customXml/[^"]+"[^>]*/>\s*',
        ]
        updated = xml
        for p in padroes:
            updated = re.sub(p, "", updated)
        if updated == xml:
            return content_types_bytes
        return updated.encode("utf-8")

    @staticmethod
    def _excel_safe_limpar_sheet_rels(rels_bytes: bytes, sheet_parte: str) -> bytes:
        """Remove rels de pivotTable, printerSettings e drawings descartados."""
        try:
            xml = rels_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return rels_bytes
        padroes = [
            r"<Relationship\b[^>]*Type=\"[^\"]*/pivotTable\"[^>]*/>\s*",
            r"<Relationship\b[^>]*Type=\"[^\"]*/printerSettings\"[^>]*/>\s*",
        ]
        # sheets que usam drawing1/drawing3 (Painel/APOIO) perdem drawing e imagens
        if sheet_parte in (
            "xl/worksheets/_rels/sheet1.xml.rels",
            "xl/worksheets/_rels/sheet5.xml.rels",
        ):
            padroes.append(r"<Relationship\b[^>]*Type=\"[^\"]*/drawing\"[^>]*/>\s*")
            padroes.append(r"<Relationship\b[^>]*Type=\"[^\"]*/image\"[^>]*/>\s*")
        updated = xml
        for p in padroes:
            updated = re.sub(p, "", updated)
        if updated == xml:
            return rels_bytes
        return updated.encode("utf-8")

    @staticmethod
    def _excel_safe_limpar_sheet_xml(sheet_xml_bytes: bytes, sheet_parte: str) -> bytes:
        """Remove tags <drawing>, <pageSetup r:id>, <picture>, <legacyDrawing> que
        referenciam rels removidos. Para sheet1 e sheet5 remove também <drawing>."""
        try:
            xml = sheet_xml_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return sheet_xml_bytes
        # Remove pageSetup com r:id (aponta para printerSettings removido)
        xml_novo = re.sub(
            r'<pageSetup\b([^/>]*?)\br:id="[^"]*"([^/>]*)/>',
            r"<pageSetup\1\2/>",
            xml,
        )
        # Remove drawing de sheets cujo drawing foi descartado
        if sheet_parte in ("xl/worksheets/sheet1.xml", "xl/worksheets/sheet5.xml"):
            # Usa [^>]* para aceitar atributos com URIs (contêm '/'),
            # cenário em que [^/]* deixava <drawing r:id="..."/> órfão.
            xml_novo = re.sub(r"<drawing\b[^>]*?/>\s*", "", xml_novo)
            xml_novo = re.sub(r"<drawing\b[^>]*?>.*?</drawing>\s*", "", xml_novo, flags=re.DOTALL)
        # picture e legacyDrawing podem apontar para recursos removidos
        xml_novo = re.sub(r"<picture\b[^>]*?/>\s*", "", xml_novo)
        xml_novo = re.sub(r"<picture\b[^>]*?>.*?</picture>\s*", "", xml_novo, flags=re.DOTALL)
        if xml_novo == xml:
            return sheet_xml_bytes
        return xml_novo.encode("utf-8")

    @staticmethod
    def _remover_defined_names_orfaos_workbook_xml(workbook_xml_bytes: bytes) -> bytes:
        """Remove definedNames órfãos (#N/A ou de slicers) que causam reparo no Excel."""
        try:
            workbook_xml = workbook_xml_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return workbook_xml_bytes

        def _filtrar_defined_names(match: "re.Match[str]") -> str:
            bloco = match.group(0)
            inner = match.group(1)

            def _manter(dn_match: "re.Match[str]") -> str:
                nome_attr = re.search(r'name="([^"]*)"', dn_match.group(0))
                valor = dn_match.group(2).strip()
                nome = nome_attr.group(1) if nome_attr else ""
                if valor in ("#N/A", "#REF!") or nome.startswith("SegmentaçãodeDados") or nome.startswith("SegmentacaodeDados"):
                    return ""
                return dn_match.group(0)

            novo_inner = re.sub(
                r"<definedName\b([^>]*)>(.*?)</definedName>",
                _manter,
                inner,
                flags=re.DOTALL,
            )
            novo_inner_strip = novo_inner.strip()
            if not novo_inner_strip:
                return ""
            return bloco.replace(inner, novo_inner)

        updated = re.sub(
            r"<definedNames>(.*?)</definedNames>",
            _filtrar_defined_names,
            workbook_xml,
            flags=re.DOTALL,
        )
        if updated == workbook_xml:
            return workbook_xml_bytes
        return updated.encode("utf-8")

    @staticmethod
    def _remover_slicer_ext_worksheet_xml(sheet_xml_bytes: bytes) -> bytes:
        """Remove bloco de slicerList da worksheet e limpa extLst vazio."""
        try:
            sheet_xml = sheet_xml_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return sheet_xml_bytes

        updated = re.sub(
            r"<ext\b[^>]*>\s*<x14:slicerList>.*?</x14:slicerList>\s*</ext>\s*",
            "",
            sheet_xml,
            flags=re.DOTALL,
        )
        updated = re.sub(r"<extLst>\s*</extLst>\s*", "", updated, flags=re.DOTALL)
        if updated == sheet_xml:
            return sheet_xml_bytes
        return updated.encode("utf-8")

    @staticmethod
    def _remover_slicer_shapes_drawing_xml(drawing_xml_bytes: bytes) -> bytes:
        """Remove anchors de slicer em drawing*.xml (objetos visuais de slicer)."""
        try:
            drawing_xml = drawing_xml_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return drawing_xml_bytes

        updated = re.sub(
            r"<xdr:(?:twoCellAnchor|oneCellAnchor|absoluteAnchor)\b[^>]*>"
            r".*?(?:drawing/2010/slicer|<[A-Za-z_][\\w.-]*:slicer\\b).*?"
            r"</xdr:(?:twoCellAnchor|oneCellAnchor|absoluteAnchor)>",
            "",
            drawing_xml,
            flags=re.DOTALL,
        )
        updated = re.sub(
            r"<(?:[A-Za-z_][\\w.-]*:)?AlternateContent\b[^>]*>"
            r".*?(?:drawing/2010/slicer|<[A-Za-z_][\\w.-]*:slicer\\b).*?"
            r"</(?:[A-Za-z_][\\w.-]*:)?AlternateContent>",
            "",
            updated,
            flags=re.DOTALL,
        )
        if updated == drawing_xml:
            return drawing_xml_bytes
        return updated.encode("utf-8")

    @staticmethod
    def _garantir_rel_shared_strings_workbook(rels_bytes: bytes) -> bytes:
        """Garante relationship de sharedStrings em xl/_rels/workbook.xml.rels."""
        try:
            rels_xml = rels_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return rels_bytes

        if "/relationships/sharedStrings" in rels_xml:
            return rels_bytes

        ids = re.findall(r'Id="rId(\d+)"', rels_xml)
        next_id = max((int(i) for i in ids), default=0) + 1
        rel_node = (
            f'<Relationship Id="rId{next_id}" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" '
            'Target="sharedStrings.xml"/>'
        )
        m_end = re.search(r"</(?:[A-Za-z0-9]+:)?Relationships>\s*$", rels_xml)
        if not m_end:
            return rels_bytes
        updated = rels_xml[: m_end.start()] + rel_node + rels_xml[m_end.start() :]
        return updated.encode("utf-8")

    @staticmethod
    def _garantir_content_type_shared_strings(content_types_bytes: bytes) -> bytes:
        """Garante override de sharedStrings no [Content_Types].xml."""
        try:
            content_types_xml = content_types_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return content_types_bytes

        if "/xl/sharedStrings.xml" in content_types_xml:
            return content_types_bytes

        override = (
            '<Override PartName="/xl/sharedStrings.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
        )
        m_end = re.search(r"</(?:[A-Za-z0-9]+:)?Types>\s*$", content_types_xml)
        if not m_end:
            return content_types_bytes
        updated = content_types_xml[: m_end.start()] + override + content_types_xml[m_end.start() :]
        return updated.encode("utf-8")

    @staticmethod
    def _normalizar_targets_workbook_rels(rels_bytes: bytes) -> bytes:
        """Normaliza targets absolutos '/xl/...' para relativos em workbook.xml.rels."""
        try:
            rels_xml = rels_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return rels_bytes

        updated = re.sub(
            r'(Target=(?:"|\'))/xl/',
            r"\1",
            rels_xml,
        )
        if updated == rels_xml:
            return rels_bytes
        return updated.encode("utf-8")

    @staticmethod
    def _normalizar_targets_sheet_rels(rels_bytes: bytes) -> bytes:
        """Normaliza targets de sheet .rels para caminho relativo (`../...`).

        openpyxl pode serializar targets como `/xl/tables/tableN.xml` em
        `xl/worksheets/_rels/sheetN.xml.rels`. Esse formato absoluto quebra a
        resolução de metadados de table slicers em parsers compatíveis (e pode
        impedir o vínculo slicer->table no Excel). Aqui convertemos para o
        formato canônico relativo, ex.: `../tables/tableN.xml`.
        """
        try:
            rels_xml = rels_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return rels_bytes

        updated = re.sub(
            r'(Target=(?:"|\'))(?:/)?xl/',
            r"\1../",
            rels_xml,
        )
        if updated == rels_xml:
            return rels_bytes
        return updated.encode("utf-8")

    def _forcar_recalculo_workbook_xml(self, workbook_xml_bytes: bytes) -> bytes:
        """Garante fullCalcOnLoad/forceFullCalc preservando namespaces originais.

        IMPORTANTE:
        - Não usar parser XML aqui (ElementTree) para evitar reescrever prefixes
          (`mc`, `x15`, `xr`, etc.) usados por MCE (markup compatibility) no
          workbook.xml. Mudança de prefix pode levar o Excel a "reparar" arquivo.
        """
        try:
            workbook_xml = workbook_xml_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return workbook_xml_bytes

        def _normalize_calc_attrs(attrs: str) -> str:
            cleaned = re.sub(
                r"""\s(?:fullCalcOnLoad|forceFullCalc|calcMode)\s*=\s*(?:"[^"]*"|'[^']*')""",
                "",
                attrs,
            )
            cleaned = cleaned.rstrip()
            if "calcMode=" not in cleaned:
                cleaned += ' calcMode="auto"'
            cleaned += ' fullCalcOnLoad="1" forceFullCalc="1"'
            return cleaned

        # 1) calcPr self-closing: <calcPr .../>
        m_self = re.search(r"<((?:\w+:)?calcPr)\b([^>]*)/>", workbook_xml)
        if m_self:
            tag = m_self.group(1)
            attrs = _normalize_calc_attrs(m_self.group(2))
            replaced = f"<{tag}{attrs}/>"
            workbook_xml = workbook_xml[:m_self.start()] + replaced + workbook_xml[m_self.end():]

        else:
            # 2) calcPr com abertura/fechamento: <calcPr ...>...</calcPr>
            m_open = re.search(r"<((?:\w+:)?calcPr)\b([^>]*)>", workbook_xml)
            if m_open:
                tag = m_open.group(1)
                attrs = _normalize_calc_attrs(m_open.group(2))
                replaced = f"<{tag}{attrs}>"
                workbook_xml = workbook_xml[:m_open.start()] + replaced + workbook_xml[m_open.end():]

            else:
                # 3) Sem calcPr: injeta antes do fechamento de workbook.
                m_end = re.search(r"</(?:\w+:)?workbook>\s*$", workbook_xml)
                if m_end:
                    calc_node = '<calcPr calcMode="auto" fullCalcOnLoad="1" forceFullCalc="1"/>'
                    workbook_xml = workbook_xml[:m_end.start()] + calc_node + workbook_xml[m_end.start():]
                    # segue para normalizar workbookPr quando houver

        # 4) Não altera workbookPr para minimizar risco de incompatibilidade.

        return workbook_xml.encode("utf-8")

    def _aplicar_override_table_ref(self, table_xml_bytes: bytes) -> bytes:
        """Atualiza ref/autoFilter ref e remove fórmulas calculadas quando pedido.

        Necessário porque o salvar final copia tables/*.xml do template; sem
        patch direto o range antigo prevalece e o pivot lê linhas vazias.
        """
        if not self._table_refs_override and not self._table_strip_calculated_formulas:
            return table_xml_bytes
        try:
            xml = table_xml_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return table_xml_bytes

        m = re.search(r'<table\b[^>]*\bname="([^"]+)"', xml)
        if not m:
            return table_xml_bytes
        name = m.group(1)
        novo_ref = self._table_refs_override.get(name)
        xml_novo = xml
        if novo_ref:
            xml_novo = re.sub(
                r'(<table\b[^>]*\bref=")[^"]*(")',
                rf'\g<1>{novo_ref}\g<2>',
                xml_novo,
                count=1,
            )
            xml_novo = re.sub(
                r'(<autoFilter\b[^>]*\bref=")[^"]*(")',
                rf'\g<1>{novo_ref}\g<2>',
                xml_novo,
                count=1,
            )
        if name in self._table_strip_calculated_formulas:
            xml_novo = re.sub(
                r"<calculatedColumnFormula>.*?</calculatedColumnFormula>",
                "",
                xml_novo,
                flags=re.DOTALL,
            )
        if xml_novo == xml:
            return table_xml_bytes
        return xml_novo.encode("utf-8")

    def _forcar_refresh_pivot_cache_xml(self, pivot_cache_xml_bytes: bytes) -> bytes:
        """Força refreshOnLoad nos pivot caches sem reescrever namespaces."""
        try:
            pivot_xml = pivot_cache_xml_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return pivot_cache_xml_bytes

        m_open = re.search(r"<((?:\w+:)?pivotCacheDefinition)\b([^>]*)>", pivot_xml)
        if not m_open:
            return pivot_cache_xml_bytes

        tag = m_open.group(1)
        attrs = m_open.group(2)
        attrs = re.sub(
            r"""\s(?:refreshOnLoad|saveData|recordCount)\s*=\s*(?:"[^"]*"|'[^']*')""",
            "",
            attrs,
        )
        attrs = attrs.rstrip() + ' refreshOnLoad="1"'
        if self._pivot_cache_records_mode in {"remove", "empty"}:
            attrs += ' saveData="0"'
        if self._pivot_cache_records_mode == "empty":
            attrs += ' recordCount="0"'
        replaced = f"<{tag}{attrs}>"
        pivot_xml = pivot_xml[:m_open.start()] + replaced + pivot_xml[m_open.end():]
        return pivot_xml.encode("utf-8")

    @staticmethod
    def _extrair_sheet_data(xml: str) -> str | None:
        m = re.search(r"<sheetData\b[^>]*>.*?</sheetData>", xml, re.DOTALL)
        if not m:
            return None
        return m.group(0)

    @staticmethod
    def _aplicar_sheetdata_puro(sheet_tpl_bytes: bytes, sheetdata_xml: str) -> bytes:
        """Substitui <sheetData> no template por XML construído externamente.

        Preserva o header (row 1) do template seguido das novas linhas de dados.
        """
        try:
            tpl = sheet_tpl_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return sheet_tpl_bytes

        m_sd = re.search(r"<sheetData\b[^>]*>(.*?)</sheetData>", tpl, re.DOTALL)
        if not m_sd:
            return sheet_tpl_bytes

        # Extrai header row 1 do template
        header_row = ""
        m_row1 = re.search(r"<row\b[^>]*\br=\"1\"[^>]*>.*?</row>", m_sd.group(1), re.DOTALL)
        if m_row1:
            header_row = m_row1.group(0)

        # Monta novo sheetData: tag de abertura + header + dados fornecidos
        m_open = re.search(r"<sheetData\b[^>]*>", tpl)
        novo_sd = m_open.group(0) + header_row + sheetdata_xml + "</sheetData>"
        novo = tpl[: m_sd.start()] + novo_sd + tpl[m_sd.end() :]

        # Atualiza <dimension>
        m_ultimo_r = re.findall(r'<row\b[^>]*\br="(\d+)"', sheetdata_xml)
        if m_ultimo_r:
            ultima = m_ultimo_r[-1]
            novo = re.sub(
                r'(<dimension\b[^>]*\bref=")[^"]*(")',
                lambda x: re.sub(r'\d+(?=")', ultima, x.group(0)),
                novo,
                count=1,
            )
        return novo.encode("utf-8")

    def _aplicar_sheet_data_override(
        self,
        sheet_tpl_bytes: bytes,
        sheet_edit_bytes: bytes,
    ) -> bytes:
        """Substitui apenas o bloco <sheetData> da sheet do template."""
        try:
            tpl = sheet_tpl_bytes.decode("utf-8")
            edt = sheet_edit_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return sheet_edit_bytes

        sheet_data_edit = self._extrair_sheet_data(edt)
        sheet_data_tpl = self._extrair_sheet_data(tpl)
        if not sheet_data_edit or not sheet_data_tpl:
            return sheet_edit_bytes

        tpl_novo = tpl.replace(sheet_data_tpl, sheet_data_edit, 1)
        return tpl_novo.encode("utf-8")

    def _limpar_items_slicer_cache_xml(self, slicer_cache_xml_bytes: bytes) -> bytes:
        """Limpa bloco <items> de um slicerCache mantendo estrutura e namespaces."""
        try:
            xml = slicer_cache_xml_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return slicer_cache_xml_bytes

        pattern = re.compile(
            r"<(?P<tag>(?:[A-Za-z_][\w.-]*:)?items)\b(?P<attrs>[^>]*)>"
            r"(?P<body>.*?)"
            r"</(?P=tag)>",
            re.DOTALL,
        )

        def _replace(m: re.Match) -> str:
            tag = m.group("tag")
            attrs = m.group("attrs")
            if re.search(r"""\bcount\s*=\s*(['"])[^'"]*\1""", attrs):
                attrs = re.sub(
                    r"""(\bcount\s*=\s*)(['"])[^'"]*(\2)""",
                    r"""\g<1>\g<2>0\g<3>""",
                    attrs,
                    count=1,
                )
            else:
                attrs = attrs.rstrip() + ' count="0"'
            return f"<{tag}{attrs}></{tag}>"

        updated = pattern.sub(_replace, xml)
        if updated == xml:
            return slicer_cache_xml_bytes
        return updated.encode("utf-8")

    def _esvaziar_pivot_cache_records_xml(self, records_xml_bytes: bytes) -> bytes:
        """Substitui pivotCacheRecords por raiz vazia count=0."""
        try:
            xml = records_xml_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return records_xml_bytes

        m_open = re.search(r"<((?:\w+:)?pivotCacheRecords)\b([^>]*)>", xml)
        if not m_open:
            return records_xml_bytes
        tag = m_open.group(1)
        attrs = m_open.group(2)
        attrs = re.sub(r"""\scount\s*=\s*(?:"[^"]*"|'[^']*')""", "", attrs)
        attrs = attrs.rstrip() + ' count="0"'
        replacement = f"<{tag}{attrs}/>"

        m_close = re.search(rf"</{re.escape(tag)}>", xml)
        if not m_close:
            # Já self-closing, apenas normaliza count.
            updated = re.sub(
                rf"<{re.escape(tag)}\b[^>]*/>",
                replacement,
                xml,
                count=1,
            )
            if updated == xml:
                return records_xml_bytes
            return updated.encode("utf-8")

        updated = xml[:m_open.start()] + replacement + xml[m_close.end():]
        return updated.encode("utf-8")

    def _aplicar_table_specs_workbook(self) -> None:
        """Materializa tables registradas via `registrar_table_ooxml` no workbook."""
        if not self._wb or not self._table_specs:
            return

        for spec in self._table_specs.values():
            ws = self._wb[spec.sheet_name]

            # Se a table já existir nessa aba, substitui para manter ref/estilo atualizados.
            if spec.table_name in ws.tables:
                del ws.tables[spec.table_name]

            table = Table(displayName=spec.table_name, ref=spec.ref)
            table.tableStyleInfo = TableStyleInfo(
                name=spec.style,
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)
            # AutoFilter passa a ser controlado pela própria table.
            ws.auto_filter.ref = None
            self._modified_sheets.add(spec.sheet_name)

    @staticmethod
    def _next_rid(existing_ids: set[str]) -> str:
        idx = 1
        while f"rId{idx}" in existing_ids:
            idx += 1
        return f"rId{idx}"

    def _mesclar_sheet_com_template(
        self,
        sheet_tpl_bytes: bytes,
        sheet_edit_bytes: bytes,
        rel_tpl_bytes: bytes | None,
        rel_edit_bytes: bytes | None,
    ) -> tuple[bytes, bytes | None]:
        """Mescla sheet modificada preservando relações visuais do template.

        Mantém relações de tabela da versão editada (consistência do tablePart)
        e recupera printerSettings do template quando o openpyxl remove.
        """
        if not sheet_tpl_bytes:
            return sheet_edit_bytes, rel_edit_bytes

        self._registrar_namespaces_xml()
        sheet_tpl = ET.fromstring(sheet_tpl_bytes)
        sheet_edit = ET.fromstring(sheet_edit_bytes)

        # Preserva atributos de compatibilidade da raiz da worksheet (ex.: mc:Ignorable, xr:uid).
        for attr_name, attr_val in sheet_tpl.attrib.items():
            if attr_name not in sheet_edit.attrib:
                sheet_edit.attrib[attr_name] = attr_val

        # Garante declaração de todos os namespaces listados em mc:Ignorable.
        # Excel trata como corrupção se Ignorable="x14ac xr xr2 xr3" sem xmlns correspondente.
        _NS_IGNORABLE_MAP = {
            "x14ac": self.NS_X14AC,
            "xr": self.NS_XR,
            "xr2": self.NS_XR2,
            "xr3": self.NS_XR3,
        }
        ignorable = sheet_edit.attrib.get(f"{{{self.NS_MC}}}Ignorable", "")
        for prefix in ignorable.split():
            uri = _NS_IGNORABLE_MAP.get(prefix)
            if uri:
                ET.register_namespace(prefix, uri)

        had_rel_edit = bool(rel_edit_bytes)
        had_rel_tpl = bool(rel_tpl_bytes)
        rel_edit = (
            ET.fromstring(rel_edit_bytes)
            if rel_edit_bytes
            else ET.Element(f"{{{self.NS_REL_PKG}}}Relationships")
        )
        rel_tpl = ET.fromstring(rel_tpl_bytes) if rel_tpl_bytes else None

        rels_por_id_tpl: dict[str, ET.Element] = {}
        if rel_tpl is not None:
            for rel in rel_tpl.findall(f"{{{self.NS_REL_PKG}}}Relationship"):
                rid = rel.attrib.get("Id")
                if rid:
                    rels_por_id_tpl[rid] = rel

        rels_por_id_edit: dict[str, ET.Element] = {}
        for rel in rel_edit.findall(f"{{{self.NS_REL_PKG}}}Relationship"):
            rid = rel.attrib.get("Id")
            if rid:
                rels_por_id_edit[rid] = rel

        existing_ids = set(rels_por_id_edit.keys())
        rid_printer_final = None
        rid_remap: dict[str, str] = {}

        # Procurar printerSettings do template.
        rid_printer_tpl = None
        for rid, rel in rels_por_id_tpl.items():
            if rel.attrib.get("Type") == self.REL_PRINTER:
                rid_printer_tpl = rid
                break

        if rid_printer_tpl:
            # Se já existir printerSettings no editado, reaproveita.
            for rid, rel in rels_por_id_edit.items():
                if rel.attrib.get("Type") == self.REL_PRINTER:
                    rid_printer_final = rid
                    break

            if not rid_printer_final:
                rel_tpl_printer = rels_por_id_tpl[rid_printer_tpl]
                rid_novo = rid_printer_tpl
                if rid_novo in existing_ids:
                    rid_novo = self._next_rid(existing_ids)
                existing_ids.add(rid_novo)

                novo = ET.Element(f"{{{self.NS_REL_PKG}}}Relationship")
                novo.attrib.update(
                    {
                        "Id": rid_novo,
                        "Type": rel_tpl_printer.attrib.get("Type", self.REL_PRINTER),
                        "Target": rel_tpl_printer.attrib.get("Target", ""),
                    }
                )
                rel_edit.append(novo)
                rid_printer_final = rid_novo

            page_setup_tpl = sheet_tpl.find(f"{{{self.NS_MAIN}}}pageSetup")
            page_setup_edit = sheet_edit.find(f"{{{self.NS_MAIN}}}pageSetup")
            if (
                rid_printer_final
                and page_setup_tpl is not None
                and page_setup_tpl.attrib.get(f"{{{self.NS_REL}}}id")
                and page_setup_edit is not None
            ):
                page_setup_edit.attrib[f"{{{self.NS_REL}}}id"] = rid_printer_final

        # Preservar relacionamento de slicer quando o openpyxl remove (modo legado).
        if not self._strip_slicers:
            rid_slicer_tpl = None
            for rid, rel in rels_por_id_tpl.items():
                if rel.attrib.get("Type") == self.REL_SLICER:
                    rid_slicer_tpl = rid
                    break

            if rid_slicer_tpl:
                existe_slicer_edit = any(
                    rel.attrib.get("Type") == self.REL_SLICER
                    for rel in rels_por_id_edit.values()
                )
                if not existe_slicer_edit:
                    rel_tpl_slicer = rels_por_id_tpl[rid_slicer_tpl]
                    rid_novo = rid_slicer_tpl
                    if rid_novo in existing_ids:
                        rid_novo = self._next_rid(existing_ids)
                        rid_remap[rid_slicer_tpl] = rid_novo
                    existing_ids.add(rid_novo)

                    novo = ET.Element(f"{{{self.NS_REL_PKG}}}Relationship")
                    novo.attrib.update(
                        {
                            "Id": rid_novo,
                            "Type": rel_tpl_slicer.attrib.get("Type", self.REL_SLICER),
                            "Target": rel_tpl_slicer.attrib.get("Target", ""),
                        }
                    )
                    rel_edit.append(novo)

        # Preserva extLst da sheet quando removido pelo openpyxl.
        ext_tpl = sheet_tpl.find(f"{{{self.NS_MAIN}}}extLst")
        ext_edit = sheet_edit.find(f"{{{self.NS_MAIN}}}extLst")
        if ext_tpl is not None and ext_edit is None and not self._strip_slicers:
            ext_clone = deepcopy(ext_tpl)
            if rid_remap:
                rid_attr = f"{{{self.NS_REL}}}id"
                for node in ext_clone.iter():
                    rid_val = node.attrib.get(rid_attr)
                    if rid_val and rid_val in rid_remap:
                        node.attrib[rid_attr] = rid_remap[rid_val]
            sheet_edit.append(ext_clone)

        for rel in rel_edit.findall(f"{{{self.NS_REL_PKG}}}Relationship"):
            rel_type = rel.attrib.get("Type", "")
            target = rel.attrib.get("Target", "")
            if target.startswith("/xl/"):
                rel.attrib["Target"] = f"../{target[len('/xl/'):]}"

        rel_result: bytes | None
        if had_rel_edit or had_rel_tpl:
            rel_result = self._serializar_rels(rel_edit)
        else:
            rel_result = None

        sheet_bytes = self._serializar_xml(sheet_edit)
        sheet_bytes = self._garantir_xmlns_ignorable(sheet_bytes, ignorable)
        if not self._strip_slicers:
            sheet_bytes = self._garantir_xmlns_ext_slicer(sheet_bytes)
        return sheet_bytes, rel_result

    @classmethod
    def _serializar_rels(cls, rel_root: ET.Element) -> bytes:
        """Serializa um .rels garantindo namespace default (sem prefix ns0:)."""
        raw = ET.tostring(rel_root, encoding="utf-8", xml_declaration=True)
        text = raw.decode("utf-8")
        # ElementTree pode reescrever como `ns0:Relationship ... xmlns:ns0="..."`.
        # Normaliza para o formato canônico OOXML: `<Relationships xmlns="...">`.
        text = re.sub(r"<ns0:Relationships\b", "<Relationships", text)
        text = re.sub(r"</ns0:Relationships>", "</Relationships>", text)
        text = re.sub(r"<ns0:Relationship\b", "<Relationship", text)
        text = re.sub(r"</ns0:Relationship>", "</Relationship>", text)
        text = text.replace('xmlns:ns0="', 'xmlns="')
        return text.encode("utf-8")

    def _garantir_xmlns_ignorable(self, sheet_bytes: bytes, ignorable: str) -> bytes:
        """Garante que todos os prefixos em mc:Ignorable tenham xmlns declarado no root.

        Excel trata como corrupção se `mc:Ignorable` referenciar prefixo sem xmlns.
        """
        if not ignorable:
            return sheet_bytes
        try:
            text = sheet_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return sheet_bytes

        ns_map = {
            "x14ac": self.NS_X14AC,
            "xr": self.NS_XR,
            "xr2": self.NS_XR2,
            "xr3": self.NS_XR3,
        }
        m = re.search(r"<worksheet\b[^>]*>", text)
        if not m:
            return sheet_bytes
        root_tag = m.group(0)
        novos: list[str] = []
        for prefix in ignorable.split():
            uri = ns_map.get(prefix)
            if not uri:
                continue
            if f'xmlns:{prefix}="' in root_tag:
                continue
            novos.append(f'xmlns:{prefix}="{uri}"')
        if not novos:
            return sheet_bytes

        # Insere antes do `>` (ou `/>`) no final da tag.
        if root_tag.endswith("/>"):
            novo_tag = root_tag[:-2] + " " + " ".join(novos) + "/>"
        else:
            novo_tag = root_tag[:-1] + " " + " ".join(novos) + ">"
        text = text[: m.start()] + novo_tag + text[m.end() :]
        return text.encode("utf-8")

    def _aplicar_patch_cols_hidden(
        self, sheet_xml_bytes: bytes, cols_map: dict[int, bool]
    ) -> bytes:
        """Atualiza atributos `hidden` de `<col>` no XML bruto da sheet.

        Divide ranges `min=a max=b` em sub-ranges quando necessário para aplicar
        o flag por coluna. Elementos não tocados permanecem intactos.
        """
        if not cols_map:
            return sheet_xml_bytes
        try:
            xml = sheet_xml_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return sheet_xml_bytes

        m_cols = re.search(r"<cols\b[^>]*>(.*?)</cols>", xml, re.DOTALL)
        if not m_cols:
            return sheet_xml_bytes

        cols_inner = m_cols.group(1)
        col_re = re.compile(r"<col\b([^/]*?)/>", re.DOTALL)
        novos_fragmentos: list[str] = []
        last = 0
        for m in col_re.finditer(cols_inner):
            novos_fragmentos.append(cols_inner[last:m.start()])
            last = m.end()
            attrs = m.group(1)
            min_m = re.search(r'\bmin="(\d+)"', attrs)
            max_m = re.search(r'\bmax="(\d+)"', attrs)
            if not min_m or not max_m:
                novos_fragmentos.append(m.group(0))
                continue
            cmin, cmax = int(min_m.group(1)), int(max_m.group(1))
            # Sem colunas afetadas neste range — mantém.
            afeta = {c: cols_map[c] for c in range(cmin, cmax + 1) if c in cols_map}
            if not afeta:
                novos_fragmentos.append(m.group(0))
                continue
            # Expande em sub-ranges contíguos por estado final de hidden.
            estado_por_col: dict[int, bool | None] = {}
            hidden_atual = bool(re.search(r'\bhidden="1"', attrs))
            for c in range(cmin, cmax + 1):
                estado_por_col[c] = afeta.get(c, hidden_atual)

            def _render_sub(sub_min: int, sub_max: int, hidden: bool) -> str:
                novo_attrs = re.sub(r'\s+min="\d+"', "", attrs)
                novo_attrs = re.sub(r'\s+max="\d+"', "", novo_attrs)
                novo_attrs = re.sub(r'\s+hidden="(?:0|1|true|false)"', "", novo_attrs)
                novo_attrs = novo_attrs.strip()
                hidden_attr = ' hidden="1"' if hidden else ""
                prefix = " " if novo_attrs else ""
                return f'<col min="{sub_min}" max="{sub_max}"{prefix}{novo_attrs}{hidden_attr}/>'

            sub_min = cmin
            sub_estado = estado_por_col[cmin]
            for c in range(cmin + 1, cmax + 1):
                if estado_por_col[c] != sub_estado:
                    novos_fragmentos.append(_render_sub(sub_min, c - 1, bool(sub_estado)))
                    sub_min = c
                    sub_estado = estado_por_col[c]
            novos_fragmentos.append(_render_sub(sub_min, cmax, bool(sub_estado)))

        novos_fragmentos.append(cols_inner[last:])
        novo_cols_inner = "".join(novos_fragmentos)
        novo_xml = xml[: m_cols.start(1)] + novo_cols_inner + xml[m_cols.end(1) :]
        return novo_xml.encode("utf-8")

    def _garantir_xmlns_ext_slicer(self, sheet_bytes: bytes) -> bytes:
        """Garante xmlns:x14 no `<ext>` do slicerList (URI A8765BA9...).

        Excel exige declaração do namespace no próprio `<ext>`; apenas no
        root `<worksheet>` não basta para o slicer renderizar.
        """
        try:
            text = sheet_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return sheet_bytes

        uri_slicer = "{A8765BA9-456A-4dab-B4F3-ACF838C121DE}"
        pattern = re.compile(
            r'<ext\b([^>]*\buri="' + re.escape(uri_slicer) + r'"[^>]*)>',
            re.IGNORECASE,
        )

        def _inject(m: re.Match) -> str:
            attrs = m.group(1)
            if 'xmlns:x14=' in attrs:
                return m.group(0)
            novo = attrs.rstrip() + f' xmlns:x14="{self.NS_X14}"'
            return f"<ext{novo}>"

        novo_text, n = pattern.subn(_inject, text)
        if n == 0:
            return sheet_bytes
        return novo_text.encode("utf-8")

    def _salvar_com_preservacao_visual(self, workbook_salvo: Path, output_path: Path) -> None:
        """Monta o arquivo final preservando partes visuais do template.

        Estratégia:
        - Base: pacote OOXML do template original (preserva layout e estilos)
        - Substitui apenas as parts XML das sheets modificadas com versões editadas pelo openpyxl.
        """
        # Sheets com sheetData puro são tratadas via patch no template — não
        # precisam passar pelo loop de openpyxl. Remove-as de _modified_sheets
        # para que rels, styles e XML do openpyxl não contaminem o output.
        modified_efetivo = self._modified_sheets - set(self._sheetdata_xml_puro)

        sem_patch_estrutural = (
            not modified_efetivo
            and not self._cols_patch_only
            and not self._sheetdata_xml_puro
            and self._pivot_cache_records_mode is None
            and not self._clear_slicer_cache_items
            and not self._strip_slicers
            and not self._table_refs_override
            and not self._table_strip_calculated_formulas
            and not self._table_specs
            and not self._slicer_specs
        )
        if sem_patch_estrutural:
            workbook_salvo.replace(output_path)
            return

        partes_editado = self._mapear_sheets_para_partes(workbook_salvo)
        partes_template = self._mapear_sheets_para_partes(self.template_path)
        # Mapeia nome_parte_template -> sheet_name para aplicar patch de colunas.
        parte_para_sheet_tpl: dict[str, str] = {v: k for k, v in partes_template.items()}

        partes_substituir: set[str] = set()
        for sheet in modified_efetivo:
            parte_editada = partes_editado.get(sheet)
            if not parte_editada:
                continue
            partes_substituir.add(parte_editada)

            rel_path = self._path_rel_sheet(parte_editada)
            with zipfile.ZipFile(workbook_salvo, "r") as z_edit:
                if rel_path in z_edit.namelist():
                    partes_substituir.add(rel_path)

        # Mantém consistência de style_id apenas para sheets com serialização
        # "raw" do openpyxl (não sheetData-override). Se todas as sheets
        # modificadas usam sheetData-override, o template preserva os IDs de
        # estilo intactos — substituir styles.xml nesse caso descarta estilos
        # que a DRE e outras abas preservadas do template ainda referenciam.
        sheets_raw = self._modified_sheets - self._sheet_data_overrides - set(self._sheetdata_xml_puro)
        if sheets_raw:
            partes_substituir.add("xl/styles.xml")

        with (
            zipfile.ZipFile(self.template_path, "r") as z_tpl,
            zipfile.ZipFile(workbook_salvo, "r") as z_edit,
            zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as z_out,
        ):
            edit_names = set(z_edit.namelist())
            tpl_names = set(z_tpl.namelist())
            payload_override: dict[str, bytes] = {}
            escrito: set[str] = set()

            # Quando há novas sheets, workbook.xml/workbook.rels/[Content_Types]
            # precisam vir do pacote editado para manter o cadastro das partes.
            tem_sheet_nova = any(
                parte_editada not in tpl_names
                for parte_editada in partes_editado.values()
            )
            core_from_edit: set[str] = set()
            if tem_sheet_nova:
                core_from_edit.update(
                    {
                        "xl/workbook.xml",
                        "xl/_rels/workbook.xml.rels",
                        "[Content_Types].xml",
                        "docProps/app.xml",
                    }
                )
            if "xl/sharedStrings.xml" in edit_names:
                core_from_edit.add("xl/sharedStrings.xml")
                if "xl/sharedStrings.xml" not in tpl_names:
                    core_from_edit.add("[Content_Types].xml")

            def _aplicar_patches(nome: str, payload: bytes) -> bytes:
                if nome == "xl/workbook.xml":
                    payload = self._forcar_recalculo_workbook_xml(payload)
                    payload = self._remover_ref_calc_chain_workbook_rels(payload)
                    if self._strip_slicers:
                        payload = self._remover_slicer_ext_workbook_xml(payload)
                        payload = self._remover_defined_names_orfaos_workbook_xml(payload)
                    if self._excel_safe_mode:
                        payload = self._remover_pivot_caches_workbook_xml(payload)
                elif nome == "xl/_rels/workbook.xml.rels":
                    payload = self._remover_ref_calc_chain_workbook_rels(payload)
                    payload = self._normalizar_targets_workbook_rels(payload)
                    if self._strip_slicers:
                        payload = self._remover_rels_por_tipo_fragmento(payload, "/slicerCache")
                    if self._excel_safe_mode:
                        payload = self._excel_safe_limpar_workbook_rels(payload)
                    if "xl/sharedStrings.xml" in edit_names or "xl/sharedStrings.xml" in tpl_names:
                        payload = self._garantir_rel_shared_strings_workbook(payload)
                elif (
                    self._pivot_cache_records_mode == "remove"
                    and nome.startswith("xl/pivotCache/_rels/pivotCacheDefinition")
                    and nome.endswith(".rels")
                ):
                    payload = self._remover_rel_pivot_cache_records(payload)
                elif nome == "[Content_Types].xml":
                    if self._pivot_cache_records_mode == "remove":
                        payload = self._remover_content_type_pivot_cache_records(payload)
                    if self._strip_slicers:
                        payload = self._remover_content_type_slicer_parts(payload)
                    if self._excel_safe_mode:
                        payload = self._excel_safe_limpar_content_types(payload)
                    if "xl/sharedStrings.xml" in edit_names or "xl/sharedStrings.xml" in tpl_names:
                        payload = self._garantir_content_type_shared_strings(payload)
                elif nome.startswith("xl/worksheets/_rels/") and nome.endswith(".rels"):
                    payload = self._normalizar_targets_sheet_rels(payload)
                elif nome.startswith("xl/tables/") and nome.endswith(".xml"):
                    payload = self._aplicar_override_table_ref(payload)
                elif nome.startswith("xl/pivotCache/pivotCacheDefinition") and nome.endswith(".xml"):
                    payload = self._forcar_refresh_pivot_cache_xml(payload)
                elif (
                    self._pivot_cache_records_mode == "empty"
                    and nome.startswith("xl/pivotCache/pivotCacheRecords")
                    and nome.endswith(".xml")
                ):
                    payload = self._esvaziar_pivot_cache_records_xml(payload)
                elif (
                    self._clear_slicer_cache_items
                    and nome.startswith("xl/slicerCaches/slicerCache")
                    and nome.endswith(".xml")
                ):
                    payload = self._limpar_items_slicer_cache_xml(payload)

                if self._strip_slicers and nome.startswith("xl/worksheets/") and nome.endswith(".xml"):
                    payload = self._remover_slicer_ext_worksheet_xml(payload)
                if self._strip_slicers and nome.startswith("xl/worksheets/_rels/") and nome.endswith(".rels"):
                    payload = self._remover_rels_por_tipo_fragmento(payload, "/slicer")
                if self._strip_slicers and nome.startswith("xl/drawings/drawing") and nome.endswith(".xml"):
                    payload = self._remover_slicer_shapes_drawing_xml(payload)
                if self._excel_safe_mode:
                    if nome.startswith("xl/worksheets/_rels/") and nome.endswith(".rels"):
                        payload = self._excel_safe_limpar_sheet_rels(payload, nome)
                    elif nome.startswith("xl/worksheets/") and nome.endswith(".xml"):
                        payload = self._excel_safe_limpar_sheet_xml(payload, nome)
                return payload

            for sheet in modified_efetivo:
                parte = partes_editado.get(sheet)
                if not parte or parte not in edit_names:
                    continue

                rel_path = self._path_rel_sheet(parte)
                sheet_edit = z_edit.read(parte)
                rel_edit = z_edit.read(rel_path) if rel_path in edit_names else None

                # Modo conservador: só aplica mesclagem OOXML (ElementTree) nas
                # abas que realmente precisam preservar artefatos visuais
                # removidos pelo openpyxl (ex.: slicers/extLst da aba DRE).
                # Para abas de dados (ex.: BD_FLUXO), usa bytes "raw" do
                # openpyxl para evitar reserialização XML incompatível.
                if sheet in self._sheet_data_overrides:
                    if parte in tpl_names:
                        sheet_tpl = z_tpl.read(parte)
                        payload_override[parte] = self._aplicar_sheet_data_override(
                            sheet_tpl,
                            sheet_edit,
                        )
                    else:
                        payload_override[parte] = sheet_edit
                    # Para sheet com override de sheetData, preserva rels do
                    # template para manter ids estáveis (ex.: tablePart rId2).
                    if rel_path in tpl_names:
                        payload_override[rel_path] = z_tpl.read(rel_path)
                    elif rel_edit is not None:
                        payload_override[rel_path] = rel_edit
                elif sheet in self.SHEETS_MERGE_VISUAL:
                    sheet_tpl = z_tpl.read(parte) if parte in tpl_names else b""
                    rel_tpl = z_tpl.read(rel_path) if rel_path in tpl_names else None
                    merged_sheet, merged_rel = self._mesclar_sheet_com_template(
                        sheet_tpl,
                        sheet_edit,
                        rel_tpl,
                        rel_edit,
                    )
                    payload_override[parte] = merged_sheet
                    if merged_rel is not None:
                        payload_override[rel_path] = merged_rel
                else:
                    payload_override[parte] = sheet_edit
                    if rel_edit is not None:
                        payload_override[rel_path] = rel_edit

            for info in z_tpl.infolist():
                nome = info.filename

                # calcChain.xml do template fica obsoleto quando dados são
                # reescritos — referências a células antigas causam prompt de
                # reparo no Excel. Omitindo, o Excel reconstrói no primeiro open.
                if nome == "xl/calcChain.xml":
                    continue

                if self._strip_slicers and self._is_slicer_part(nome):
                    continue

                if self._excel_safe_mode and self._is_excel_safe_drop(nome):
                    continue

                if self._pivot_cache_records_mode == "remove" and (
                    nome.startswith("xl/pivotCache/pivotCacheRecords")
                    and nome.endswith(".xml")
                ):
                    continue

                if nome in payload_override:
                    payload = payload_override[nome]
                elif nome in core_from_edit and nome in edit_names:
                    payload = z_edit.read(nome)
                elif nome in partes_substituir and nome in edit_names:
                    payload = z_edit.read(nome)
                else:
                    payload = z_tpl.read(nome)

                payload = _aplicar_patches(nome, payload)
                # Patch de colunas ocultas e/ou sheetData puro — aplicados
                # diretamente nos bytes do template, sem reserialização openpyxl.
                if nome in parte_para_sheet_tpl:
                    sheet_nome = parte_para_sheet_tpl[nome]
                    sheetdata_puro = self._sheetdata_xml_puro.get(sheet_nome)
                    if sheetdata_puro:
                        payload = self._aplicar_sheetdata_puro(payload, sheetdata_puro)
                    cols_map = self._cols_patch_only.get(sheet_nome)
                    if cols_map:
                        payload = self._aplicar_patch_cols_hidden(payload, cols_map)
                z_out.writestr(nome, payload)
                escrito.add(nome)

            # Inclui partes novas vindas do openpyxl (ex.: novas sheets).
            for nome in sorted(edit_names - tpl_names):
                if nome in escrito:
                    continue
                if nome == "xl/calcChain.xml":
                    continue
                if self._strip_slicers and self._is_slicer_part(nome):
                    continue
                if self._excel_safe_mode and self._is_excel_safe_drop(nome):
                    continue
                if self._pivot_cache_records_mode == "remove" and (
                    nome.startswith("xl/pivotCache/pivotCacheRecords")
                    and nome.endswith(".xml")
                ):
                    continue

                payload = z_edit.read(nome)
                payload = _aplicar_patches(nome, payload)
                z_out.writestr(nome, payload)

    @staticmethod
    def _resolver_rel_target(base_rels_path: str, target: str) -> str:
        """Resolve Target de .rels para caminho de part OOXML no zip."""
        target = target.strip()
        if target.startswith("/"):
            return target.lstrip("/")

        base = base_rels_path.rsplit("/_rels/", 1)[0] if "/_rels/" in base_rels_path else ""
        parts = (base.split("/") if base else []) + target.split("/")
        resolved: list[str] = []
        for part in parts:
            if part == "..":
                if resolved:
                    resolved.pop()
            elif part and part != ".":
                resolved.append(part)
        return "/".join(resolved)

    @classmethod
    def _mapear_sheets_partes_from_parts(cls, parts: dict[str, bytes]) -> dict[str, str]:
        """Mapeia nome da sheet para part `xl/worksheets/sheetN.xml`."""
        wb_raw = parts.get("xl/workbook.xml")
        rels_raw = parts.get("xl/_rels/workbook.xml.rels")
        if not wb_raw or not rels_raw:
            return {}

        wb_xml = ET.fromstring(wb_raw)
        rels_xml = ET.fromstring(rels_raw)

        rel_map: dict[str, str] = {}
        for rel in rels_xml.findall(f"{{{cls.NS_REL_PKG}}}Relationship"):
            rel_id = rel.attrib.get("Id")
            target = rel.attrib.get("Target")
            if rel_id and target:
                rel_map[rel_id] = cls._normalizar_target_workbook_rel(target)

        out: dict[str, str] = {}
        for sheet in wb_xml.findall(f".//{{{cls.NS_MAIN}}}sheet"):
            name = sheet.attrib.get("name")
            rid = sheet.attrib.get(f"{{{cls.NS_REL}}}id")
            if not name or not rid:
                continue
            part = rel_map.get(rid)
            if part:
                out[name] = part
        return out

    @staticmethod
    def _next_part_index(parts: set[str], prefix: str, suffix: str) -> int:
        maior = 0
        for name in parts:
            if not name.startswith(prefix) or not name.endswith(suffix):
                continue
            frag = name[len(prefix) : -len(suffix)]
            if frag.isdigit():
                maior = max(maior, int(frag))
        return maior + 1

    @staticmethod
    def _append_content_type_override(
        content_types_xml: str,
        *,
        part_name: str,
        content_type: str,
    ) -> str:
        if part_name in content_types_xml:
            return content_types_xml
        override = (
            f'<Override PartName="/{part_name}" '
            f'ContentType="{content_type}"/>'
        )
        m = re.search(r"</(?:[A-Za-z0-9]+:)?Types>\s*$", content_types_xml)
        if not m:
            return content_types_xml
        return content_types_xml[: m.start()] + override + content_types_xml[m.start() :]

    @staticmethod
    def _coletar_table_metadata(parts: dict[str, bytes]) -> dict[str, dict]:
        """Lê metadados de tables do pacote final por displayName."""
        meta: dict[str, dict] = {}
        for name, payload in parts.items():
            if not name.startswith("xl/tables/") or not name.endswith(".xml"):
                continue
            try:
                root = ET.fromstring(payload)
            except ET.ParseError:
                continue
            display_name = root.attrib.get("displayName") or root.attrib.get("name")
            table_id_raw = root.attrib.get("id")
            if not display_name or not table_id_raw or not table_id_raw.isdigit():
                continue
            cols = []
            for col in root.findall(f".//{{{TemplateWriter.NS_MAIN}}}tableColumn"):
                col_name = col.attrib.get("name")
                if col_name:
                    cols.append(col_name)
            meta[display_name] = {
                "id": int(table_id_raw),
                "columns": cols,
            }
        return meta

    def _resolver_slicer_specs(self, parts: dict[str, bytes]) -> list[ResolvedSlicerSpec]:
        table_meta = self._coletar_table_metadata(parts)
        resolved: list[ResolvedSlicerSpec] = []
        used_cache_names: set[str] = set()
        used_slicer_names: set[str] = set()

        for raw in self._slicer_specs:
            table = table_meta.get(raw.table_name)
            if not table:
                logger.warning(
                    "Table '%s' não encontrada; slicer '%s' ignorado.",
                    raw.table_name,
                    raw.column_name,
                )
                continue
            columns: list[str] = table["columns"]
            try:
                col_idx = columns.index(raw.column_name) + 1
                source_name = raw.column_name
            except ValueError:
                source_name = ""
                col_idx = 0
                for idx, col in enumerate(columns, start=1):
                    if col.casefold() == raw.column_name.casefold():
                        col_idx = idx
                        source_name = col
                        break
                if col_idx == 0:
                    logger.warning(
                        "Coluna '%s' não encontrada na table '%s'; slicer ignorado.",
                        raw.column_name,
                        raw.table_name,
                    )
                    continue

            base_cache = raw.cache_name or f"Slicer_{sanitize_identifier(source_name)}"
            cache_name = base_cache
            suffix = 1
            while cache_name in used_cache_names:
                suffix += 1
                cache_name = f"{base_cache}_{suffix}"
            used_cache_names.add(cache_name)

            base_slicer = raw.slicer_name or raw.caption
            slicer_name = base_slicer
            suffix = 1
            while slicer_name in used_slicer_names:
                suffix += 1
                slicer_name = f"{base_slicer} {suffix}"
            used_slicer_names.add(slicer_name)

            resolved.append(
                ResolvedSlicerSpec(
                    cache_name=cache_name,
                    slicer_name=slicer_name,
                    caption=raw.caption,
                    source_name=source_name,
                    table_id=table["id"],
                    column_index=col_idx,
                )
            )
        return resolved

    def _injetar_slicer_caches_workbook_xml(
        self,
        workbook_xml: str,
        *,
        cache_rids: list[str],
        cache_names: list[str],
    ) -> str:
        """Injeta definedNames + extLst de slicer caches no workbook.xml."""
        for cache_name in cache_names:
            if f'name="{cache_name}"' in workbook_xml:
                continue
            if "<definedNames>" in workbook_xml:
                workbook_xml = workbook_xml.replace(
                    "</definedNames>",
                    f'<definedName name="{cache_name}">#N/A</definedName></definedNames>',
                    1,
                )
            else:
                block = "<definedNames>" + f'<definedName name="{cache_name}">#N/A</definedName>' + "</definedNames>"
                insert_at = workbook_xml.find("<calcPr")
                if insert_at == -1:
                    insert_at = workbook_xml.rfind("</workbook>")
                if insert_at != -1:
                    workbook_xml = workbook_xml[:insert_at] + block + workbook_xml[insert_at:]

        caches_nodes = "".join([f'<x14:slicerCache r:id="{rid}"/>' for rid in cache_rids])
        ext_pattern = (
            r'(<ext\b[^>]*uri="' + re.escape(self.EXT_URI_SLICER_CACHES_X15) + r'"[^>]*>)(.*?)(</ext>)'
        )
        ext_match = re.search(ext_pattern, workbook_xml, flags=re.DOTALL)
        if ext_match:
            body = ext_match.group(2)
            if "<x15:slicerCaches" in body:
                body = body.replace("</x15:slicerCaches>", caches_nodes + "</x15:slicerCaches>", 1)
            else:
                body = (
                    '<x15:slicerCaches xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main">'
                    + caches_nodes
                    + "</x15:slicerCaches>"
                )
            workbook_xml = (
                workbook_xml[: ext_match.start()]
                + ext_match.group(1)
                + body
                + ext_match.group(3)
                + workbook_xml[ext_match.end() :]
            )
            return workbook_xml

        ext_block = (
            f'<ext uri="{self.EXT_URI_SLICER_CACHES_X15}" xmlns:x15="{self.NS_X15}">'
            '<x15:slicerCaches xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main">'
            + caches_nodes
            + "</x15:slicerCaches>"
            "</ext>"
        )
        if "<extLst>" in workbook_xml:
            workbook_xml = workbook_xml.replace("</extLst>", ext_block + "</extLst>", 1)
        else:
            workbook_xml = workbook_xml.replace("</workbook>", f"<extLst>{ext_block}</extLst></workbook>", 1)
        return workbook_xml

    def _injetar_slicer_list_sheet_xml(self, sheet_xml: str, rid_slicer: str) -> str:
        """Injeta extLst/x14:slicerList na worksheet hospedeira."""
        sheet_xml = self._remover_slicer_ext_worksheet_xml(sheet_xml.encode("utf-8")).decode(
            "utf-8",
            errors="ignore",
        )
        ext_block = (
            f'<ext uri="{self.EXT_URI_SLICER_LIST_X15}" '
            f'xmlns:x15="{self.NS_X15}">'
            '<x14:slicerList xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main">'
            f'<x14:slicer r:id="{rid_slicer}"/>'
            "</x14:slicerList>"
            "</ext>"
        )
        if "<extLst>" in sheet_xml:
            sheet_xml = sheet_xml.replace("</extLst>", ext_block + "</extLst>", 1)
            return sheet_xml
        return sheet_xml.replace("</worksheet>", f"<extLst>{ext_block}</extLst></worksheet>", 1)

    def _injetar_anchors_slicer_drawing(
        self,
        drawing_xml: str,
        resolved_specs: list[ResolvedSlicerSpec],
    ) -> str:
        """Anexa anchors visuais de slicer no drawing da sheet destino."""
        drawing_xml = self._remover_slicer_shapes_drawing_xml(drawing_xml.encode("utf-8")).decode(
            "utf-8",
            errors="ignore",
        )
        max_id = max([int(v) for v in re.findall(r'<xdr:cNvPr\b[^>]*\bid="(\d+)"', drawing_xml)] or [1])
        # Grid simples (2 colunas) no topo da aba.
        positions = [
            (0, 0),
            (0, 4),
            (0, 8),
            (0, 12),
            (0, 16),
            (0, 20),
        ]
        anchors: list[str] = []
        for idx, spec in enumerate(resolved_specs):
            pos = positions[idx] if idx < len(positions) else (0, 4 + (idx * 4))
            max_id += 1
            anchors.append(
                build_drawing_slicer_anchor(
                    c_nv_pr_id=max_id,
                    slicer_name=spec.slicer_name,
                    from_col=pos[0],
                    from_row=pos[1],
                )
            )
        if not anchors:
            return drawing_xml
        return drawing_xml.replace("</xdr:wsDr>", "".join(anchors) + "</xdr:wsDr>", 1)

    def _injetar_slicers_reconstruidos(self, output_path: Path) -> None:
        """Injeta slicers de table no pacote OOXML final (pós-serialização)."""
        if not self._slicer_specs:
            return

        with zipfile.ZipFile(output_path, "r") as zin:
            parts = {info.filename: zin.read(info.filename) for info in zin.infolist()}

        # Garante base limpa de slicers legados.
        for name in list(parts.keys()):
            if self._is_slicer_part(name):
                parts.pop(name, None)
        for name in list(parts.keys()):
            if name == "xl/workbook.xml":
                payload = self._remover_slicer_ext_workbook_xml(parts[name])
                payload = self._remover_defined_names_orfaos_workbook_xml(payload)
                parts[name] = payload
            elif name == "xl/_rels/workbook.xml.rels":
                parts[name] = self._remover_rels_por_tipo_fragmento(parts[name], "/slicerCache")
            elif name == "[Content_Types].xml":
                parts[name] = self._remover_content_type_slicer_parts(parts[name])
            elif name.startswith("xl/worksheets/") and name.endswith(".xml"):
                parts[name] = self._remover_slicer_ext_worksheet_xml(parts[name])
            elif name.startswith("xl/worksheets/_rels/") and name.endswith(".rels"):
                parts[name] = self._remover_rels_por_tipo_fragmento(parts[name], "/slicer")
            elif name.startswith("xl/drawings/drawing") and name.endswith(".xml"):
                parts[name] = self._remover_slicer_shapes_drawing_xml(parts[name])

        resolved_specs = self._resolver_slicer_specs(parts)
        if not resolved_specs:
            logger.warning("Nenhum slicer resolvido para injeção.")
            return

        sheet_destinos = {s.sheet_destino for s in self._slicer_specs}
        if len(sheet_destinos) != 1:
            raise ValueError("Implementação atual suporta slicers em uma única sheet destino por geração.")
        sheet_destino = next(iter(sheet_destinos))

        part_sheet_map = self._mapear_sheets_partes_from_parts(parts)
        sheet_part = part_sheet_map.get(sheet_destino)
        if not sheet_part:
            raise ValueError(f"Sheet destino '{sheet_destino}' não encontrada no workbook final.")
        sheet_rels_part = self._path_rel_sheet(sheet_part)

        if sheet_rels_part not in parts:
            parts[sheet_rels_part] = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>'
            ).encode("utf-8")

        part_names = set(parts.keys())
        slicer_idx = self._next_part_index(part_names, "xl/slicers/slicer", ".xml")
        slicer_part = f"xl/slicers/slicer{slicer_idx}.xml"
        parts[slicer_part] = build_slicers_xml(resolved_specs)

        cache_parts: list[str] = []
        next_cache_idx = self._next_part_index(part_names, "xl/slicerCaches/slicerCache", ".xml")
        for spec in resolved_specs:
            cache_part = f"xl/slicerCaches/slicerCache{next_cache_idx}.xml"
            next_cache_idx += 1
            cache_parts.append(cache_part)
            parts[cache_part] = build_slicer_cache_xml(spec)

        # workbook rels: adiciona rel para cada slicerCache.
        wb_rels_root = ET.fromstring(parts["xl/_rels/workbook.xml.rels"])
        existing_ids = {
            rel.attrib.get("Id")
            for rel in wb_rels_root.findall(f"{{{self.NS_REL_PKG}}}Relationship")
            if rel.attrib.get("Id")
        }
        cache_rids: list[str] = []
        for cache_part in cache_parts:
            rid = self._next_rid(existing_ids)
            existing_ids.add(rid)
            rel = ET.Element(f"{{{self.NS_REL_PKG}}}Relationship")
            rel.attrib.update(
                {
                    "Id": rid,
                    "Type": self.REL_SLICER_CACHE,
                    "Target": cache_part.removeprefix("xl/"),
                }
            )
            wb_rels_root.append(rel)
            cache_rids.append(rid)
        parts["xl/_rels/workbook.xml.rels"] = self._serializar_rels(wb_rels_root)

        workbook_xml = parts["xl/workbook.xml"].decode("utf-8", errors="ignore")
        workbook_xml = self._injetar_slicer_caches_workbook_xml(
            workbook_xml,
            cache_rids=cache_rids,
            cache_names=[s.cache_name for s in resolved_specs],
        )
        parts["xl/workbook.xml"] = workbook_xml.encode("utf-8")

        # sheet destino rels: adiciona rel para slicer part.
        sheet_rels_root = ET.fromstring(parts[sheet_rels_part])
        existing_sheet_ids = {
            rel.attrib.get("Id")
            for rel in sheet_rels_root.findall(f"{{{self.NS_REL_PKG}}}Relationship")
            if rel.attrib.get("Id")
        }
        rid_slicer = self._next_rid(existing_sheet_ids)
        rel_slicer = ET.Element(f"{{{self.NS_REL_PKG}}}Relationship")
        rel_slicer.attrib.update(
            {
                "Id": rid_slicer,
                "Type": self.REL_SLICER,
                "Target": f"../slicers/{Path(slicer_part).name}",
            }
        )
        sheet_rels_root.append(rel_slicer)
        parts[sheet_rels_part] = self._serializar_rels(sheet_rels_root)

        # sheet destino xml: extLst slicerList.
        sheet_xml = parts[sheet_part].decode("utf-8", errors="ignore")
        sheet_xml = self._injetar_slicer_list_sheet_xml(sheet_xml, rid_slicer=rid_slicer)
        parts[sheet_part] = sheet_xml.encode("utf-8")

        # Drawing da sheet destino (se existir): adiciona anchors visuais.
        drawing_target = None
        for rel in sheet_rels_root.findall(f"{{{self.NS_REL_PKG}}}Relationship"):
            if rel.attrib.get("Type", "").endswith("/drawing"):
                drawing_target = rel.attrib.get("Target")
                break
        if drawing_target:
            drawing_part = self._resolver_rel_target(sheet_rels_part, drawing_target)
            if drawing_part in parts:
                drawing_xml = parts[drawing_part].decode("utf-8", errors="ignore")
                drawing_xml = self._injetar_anchors_slicer_drawing(drawing_xml, resolved_specs)
                parts[drawing_part] = drawing_xml.encode("utf-8")
        else:
            logger.warning(
                "Sheet '%s' sem drawing associado; slicers serão injetados sem anchors visuais.",
                sheet_destino,
            )

        # Content types dos novos parts.
        ct_xml = parts["[Content_Types].xml"].decode("utf-8", errors="ignore")
        for cache_part in cache_parts:
            ct_xml = self._append_content_type_override(
                ct_xml,
                part_name=cache_part,
                content_type="application/vnd.ms-excel.slicerCache+xml",
            )
        ct_xml = self._append_content_type_override(
            ct_xml,
            part_name=slicer_part,
            content_type="application/vnd.ms-excel.slicer+xml",
        )
        parts["[Content_Types].xml"] = ct_xml.encode("utf-8")

        tmp_out = output_path.with_name(output_path.stem + "_with_slicers_tmp.xlsx")
        with zipfile.ZipFile(tmp_out, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for name in sorted(parts.keys()):
                zout.writestr(name, parts[name])
        tmp_out.replace(output_path)

    def _listar_tabelas(self) -> list[str]:
        if not self._wb:
            return []

        tabelas = []
        for ws in self._wb.worksheets:
            tabelas.extend(list(ws.tables.keys()))
        return tabelas

    def validar_integridade(self) -> list[str]:
        """Valida que a estrutura do template foi preservada após escrita.

        Returns:
            Lista de problemas encontrados (vazia = OK)
        """
        if not self._wb:
            raise RuntimeError("Template não aberto.")

        problemas = []

        # Verificar sheets preservadas
        sheets_atuais = list(self._wb.sheetnames)
        for sheet in self._original_sheets:
            if sheet not in sheets_atuais:
                problemas.append(f"Sheet '{sheet}' removida do template")

        # Verificar defined names preservados
        names_atuais = list(self._wb.defined_names.keys())
        for name in self._original_defined_names:
            if name not in names_atuais:
                problemas.append(f"Defined name '{name}' removido do template")

        # Verificar tabelas preservadas
        tabelas_atuais = self._listar_tabelas()
        for tabela in self._original_tables:
            if tabela not in tabelas_atuais:
                problemas.append(f"Table '{tabela}' removida do template")

        if problemas:
            logger.warning(f"Problemas de integridade: {problemas}")
        else:
            logger.info("Integridade do template verificada: OK")

        return problemas

    def salvar(self, output_path: Path) -> Path:
        """Salva o workbook editado em novo arquivo (nunca sobrescreve o template).

        Args:
            output_path: caminho do arquivo de saída

        Returns:
            Path do arquivo salvo
        """
        if not self._wb:
            raise RuntimeError("Template não aberto.")

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        settings.temp_dir.mkdir(parents=True, exist_ok=True)

        with tempfile.NamedTemporaryFile(
            delete=False,
            suffix=".xlsx",
            dir=str(settings.temp_dir),
        ) as tmp:
            tmp_path = Path(tmp.name)

        try:
            self._aplicar_table_specs_workbook()

            # Força recálculo completo ao abrir no Excel para evitar exibição
            # vazia em arquivos com fórmulas e dados recém-escritos.
            calc = getattr(self._wb, "calculation", None)
            if calc is not None:
                calc.fullCalcOnLoad = True
                calc.forceFullCalc = True
                if not calc.calcMode:
                    calc.calcMode = "auto"

            self._wb.save(str(tmp_path))
            self._salvar_com_preservacao_visual(tmp_path, output_path)
            if self._slicers_novos_ativos:
                self._injetar_slicers_reconstruidos(output_path)
        finally:
            if tmp_path.exists():
                tmp_path.unlink()

        logger.info(f"Arquivo salvo: {output_path}")

        return output_path

    def fechar(self) -> None:
        """Fecha o workbook."""
        if self._wb:
            self._wb.close()
            self._wb = None

    def __enter__(self):
        self.abrir()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.fechar()
        return False
