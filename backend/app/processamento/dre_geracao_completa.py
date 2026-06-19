"""Serviço de geração completa de DRE com DB como fonte de verdade.

Fluxo:
  1. Determinar meses disponíveis conforme estratégia solicitada
  2. Buscar lançamentos YTD apenas para meses efetivos
  3. Preencher BD_FLUXO1 com os campos derivados materializados
  4. Ajustar o range da tabela fonte do pivot cache
  5. Controlar visibilidade de colunas na aba DRE (exibe somente meses com dados)
  6. Gerar aba detalhada mensal com visão analítica e granular dos lançamentos
"""

import logging
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl.styles import Font, PatternFill
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries

from ..config import settings
from ..contracts.persistence import DRELancamentoDB
from ..db.connection import DatabaseConnection
from ..repository.dre_repository import DRERepository
from ..templates.writer import TemplateWriter

logger = logging.getLogger(__name__)

# Nomes abreviados dos 12 meses (índice 0 = Janeiro)
_MESES_NOMES = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
_EMPRESA_PADRAO = "AIDEAL"

# Conta pai que identifica a Receita Bruta no PLANO_CONTAS.
# Receita Bruta soma valores brutos de entrada; Faturamento soma o líquido.
_RECEITA_BRUTA_CONTA_PAI = "(=)Receita Bruta"
_FATURAMENTO_ROTULO = "Faturamento"
_RECEITA_LIQUIDA_ROTULO = "(=)Receita Líquida"
_RESULTADO_LIQUIDO_ROTULO = "(=)RESULTADO LÍQUIDO"
_RESULTADO_GERENCIAL_ROTULO = "(=)RESULTADO GERENCIAL"
_PLANO_CONTAS_CODIGO_OVERRIDES = {
    "12.100": {
        "rubrica": "PREVISAO FÉRIAS",
        "conta_filho": "Despesas com Pessoal",
        "conta_pai": "(-)Gastos Fixos",
        "cod": 4,
    },
    "12.101": {
        "rubrica": "13° PREVISAO",
        "conta_filho": "Despesas com Pessoal",
        "conta_pai": "(-)Gastos Fixos",
        "cod": 4,
    },
}
_PLANO_CONTAS_TEXTO_OVERRIDES = {
    # No template legado, "IR" aparece como dedução sobre vendas. A regra atual
    # do DRE lança IR depois do resultado antes do IR, no bloco IRPJ/CSLL.
    "IR": {
        "rubrica": "IRPJ",
        "conta_filho": "IRPJ/CSLL",
        "conta_pai": "(-)IRPJ/CSLL",
        "cod": 7,
    },
    "IR Retido": {
        "rubrica": "IRPJ",
        "conta_filho": "IRPJ/CSLL",
        "conta_pai": "(-)IRPJ/CSLL",
        "cod": 7,
    },
}
_DRE_COLUNAS_VALOR = (2, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30, 32, 34)
_DRE_COLUNAS_PERCENTUAL = tuple(col + 1 for col in _DRE_COLUNAS_VALOR)
_DRE_COLUNA_ANO = 34
_DRE_COLUNA_VALOR_POR_MES = {
    1: 2,
    2: 4,
    3: 6,
    4: 10,
    5: 12,
    6: 14,
    7: 18,
    8: 20,
    9: 22,
    10: 26,
    11: 28,
    12: 30,
}
_DRE_MES_POR_COLUNA_VALOR = {col: mes for mes, col in _DRE_COLUNA_VALOR_POR_MES.items()}
_DRE_MESES_POR_COLUNA_TRIMESTRE = {
    8: (1, 2, 3),
    16: (4, 5, 6),
    24: (7, 8, 9),
    32: (10, 11, 12),
}
_RUBRICAS_IMPOSTO_SALDO_PAINEL = frozenset(
    (
        "IR",
        "IR Retido",
        "ISS",
        "ISS Retido",
        "INSS",
        "INSS Retido",
        "PIS",
        "COFINS",
        "CSLL",
        "Tarifa de Antecipação",
        "Impostos sobre vendas",
        "Deduções sobre vendas",
        "(-)Deduções sobre vendas",
        "Descontos sobre vendas",
        "Simples Nacional",
    )
)
_CODIGOS_IMPOSTO_SALDO_PAINEL = ("17.2", "17.3", "17.4", "17.5", "17.7", "17.8")
# Usado apenas para recompor Receita Bruta quando bases antigas persistiram
# recebimentos líquidos; a linha do DRE é definida pelo PLANO_CONTAS resolvido.
_IMPOSTOS_RECEITA_ROTULOS = frozenset(
    valor.casefold()
    for valor in (
        "IR",
        "IR Retido",
        "ISS",
        "ISS Retido",
        "INSS",
        "INSS Retido",
        "PIS",
        "COFINS",
        "CSLL",
        "Tarifa de Antecipação",
        "Impostos sobre Vendas",
        "(-)Deduções sobre vendas",
        "Deduções sobre vendas",
        "Descontos sobre vendas",
        "Simples Nacional",
    )
)

# Código gerencial no padrão "N.N[.N...]" no início do rótulo (ex.: "1.1.1 - ...").
_CODIGO_GERENCIAL_RE = re.compile(r"^\s*(\d+(?:\.\d+)+)")
_FORMULA_SUM_RE = re.compile(r"^=SUM\((?P<args>.+)\)$", re.IGNORECASE)
_FORMULA_DIV_IFERROR_RE = re.compile(
    r"^=IFERROR\(\s*(?P<num>\$?[A-Z]{1,3}\$?\d+)\s*/\s*"
    r"(?P<den>\$?[A-Z]{1,3}\$?\d+)\s*,\s*0\s*\)$",
    re.IGNORECASE,
)
_FORMULA_DIV_RE = re.compile(
    r"^=\s*(?P<num>\$?[A-Z]{1,3}\$?\d+)\s*/\s*"
    r"(?P<den>\$?[A-Z]{1,3}\$?\d+)\s*$",
    re.IGNORECASE,
)


def _extrair_codigo_gerencial(texto: Any) -> str | None:
    """Extrai o código gerencial (ex.: '1.1.1') do início de um rótulo de conta."""
    if not texto:
        return None
    match = _CODIGO_GERENCIAL_RE.match(str(texto))
    return match.group(1) if match else None


def _normalizar_rotulo(texto: Any) -> str:
    return re.sub(r"\s+", " ", str(texto or "").strip()).casefold()


def _eh_imposto_saldo_painel(rubrica: Any, natureza_raw: Any) -> bool:
    rubrica_texto = str(rubrica or "").strip()
    natureza_texto = str(natureza_raw or "").strip()
    if rubrica_texto in _RUBRICAS_IMPOSTO_SALDO_PAINEL:
        return True
    return any(
        valor == codigo or valor.startswith(f"{codigo} - ")
        for valor in (rubrica_texto, natureza_texto)
        for codigo in _CODIGOS_IMPOSTO_SALDO_PAINEL
    )


def _parse_competencia(competencia: str) -> tuple[int, int]:
    """Converte 'MM/AAAA' para (ano, mes)."""
    parts = competencia.replace("-", "/").replace("\\", "/").split("/")
    if len(parts) != 2:
        raise ValueError(f"Competência deve estar no formato MM/AAAA: {competencia}")
    mes_str, ano_str = parts
    mes = int(mes_str)
    ano = int(ano_str)
    if mes < 1 or mes > 12:
        raise ValueError(f"Mês da competência inválido: {mes:02d}. Use valores entre 01 e 12.")
    return ano, mes


def _data_lancamento_para_date(data_str: str) -> date:
    """Converte string ISO ou similar para date."""
    try:
        return date.fromisoformat(data_str)
    except ValueError:
        return datetime.strptime(data_str, "%Y-%m-%d %H:%M:%S").date()


class DREGeracaoCompletaService:
    """Serviço para geração completa de DRE a partir do banco de dados."""

    def __init__(self, db: DatabaseConnection | None = None):
        self.db = db or DatabaseConnection()
        self.repository = DRERepository(self.db)
        self.template_path = settings.template_dre_path

    # ------------------------------------------------------------------ #
    # Disponibilidade de meses                                             #
    # ------------------------------------------------------------------ #

    def _get_meses_disponiveis(self, ano: int, mes_alvo: int) -> list[int]:
        """Retorna lista ordenada de meses com upload completed no ano até mes_alvo."""
        meses = self.repository.get_meses_disponiveis(ano)
        return [m for m in meses if m <= mes_alvo]

    @staticmethod
    def _normalizar_meses_incluir(meses_incluir: list[int] | None) -> list[int]:
        """Normaliza lista de meses informados manualmente (1..12)."""
        if not meses_incluir:
            return []

        meses_normalizados = sorted({int(m) for m in meses_incluir})
        meses_invalidos = [m for m in meses_normalizados if m < 1 or m > 12]
        if meses_invalidos:
            raise ValueError(
                f"Meses inválidos em meses_incluir: {meses_invalidos}. Use valores entre 1 e 12."
            )
        return meses_normalizados

    def _resolver_meses_para_geracao(
        self,
        ano: int,
        mes_alvo: int,
        meses_incluir: list[int] | None = None,
        ano_todo: bool = False,
    ) -> tuple[list[int], list[int], str]:
        """Resolve meses efetivos para geração conforme estratégia solicitada.

        Returns:
            Tuple de (meses_disponiveis_ano, meses_efetivos, estrategia)
            estratégia: competencia | meses_incluir | ano_todo
        """
        meses_disponiveis_ano = self.repository.get_meses_disponiveis(ano)
        if not meses_disponiveis_ano:
            raise ValueError(
                f"Não há uploads com status 'completed' para o ano {ano}. "
                "Realize ao menos uma ingestão antes de gerar."
            )

        if ano_todo:
            return meses_disponiveis_ano, meses_disponiveis_ano, "ano_todo"

        meses_solicitados = self._normalizar_meses_incluir(meses_incluir)
        if meses_solicitados:
            disponiveis_set = set(meses_disponiveis_ano)
            meses_indisponiveis = [m for m in meses_solicitados if m not in disponiveis_set]
            if meses_indisponiveis:
                raise ValueError(
                    f"Meses sem upload completed para {ano}: "
                    f"{meses_indisponiveis}. Meses disponíveis: {meses_disponiveis_ano}."
                )
            return meses_disponiveis_ano, meses_solicitados, "meses_incluir"

        meses_efetivos = [m for m in meses_disponiveis_ano if m <= mes_alvo]
        if not meses_efetivos:
            raise ValueError(
                f"Nenhum mês disponível para geração até {mes_alvo:02d}/{ano}. "
                f"Meses disponíveis no ano: {meses_disponiveis_ano}."
            )
        return meses_disponiveis_ano, meses_efetivos, "competencia"

    def _mes_alvo_disponivel(self, ano: int, mes: int) -> bool:
        """Verifica se existe upload completed para o mês alvo."""
        uploads = self.repository.uploads.get_by_competencia(ano, mes)
        return bool(uploads and any(u.status == "completed" for u in uploads))

    # ------------------------------------------------------------------ #
    # Lançamentos                                                          #
    # ------------------------------------------------------------------ #

    def _get_lancamentos_ytd(
        self,
        ano: int,
        meses_disponiveis: list[int],
        centro_custo: str | None = None,
    ) -> list[DRELancamentoDB]:
        """Busca lançamentos somente dos meses disponíveis (não usa <= mes para evitar gaps)."""
        return self.repository.lancamentos.get_by_meses(ano, meses_disponiveis, centro_custo)

    # ------------------------------------------------------------------ #
    # PLANO_CONTAS                                                         #
    # ------------------------------------------------------------------ #

    @staticmethod
    def _ler_plano_contas(
        writer: TemplateWriter,
        *,
        aplicar_overrides_dre_gerado: bool = False,
    ) -> dict[str, dict]:
        """Lê PLANO_CONTAS do template para mapeamento natureza → contas.

        Além das chaves textuais (classificação e rubrica), indexa cada conta
        pelo código gerencial ("@cod:1.1.1") e pela família ("@fam:1.1"). Isso
        permite casar lançamentos cujo rótulo traz sufixos divergentes
        (ex.: "1.1.1 - Recebimento de Clientes - prestação de serviço") e herdar
        a classificação para subcontas irmãs ainda não cadastradas no plano
        (ex.: 1.1.2/1.1.3 herdam de 1.1.1 e somam em Receita Bruta).
        """
        ws = writer._wb["PLANO_CONTAS"]
        plano: dict[str, dict] = {}
        for row in ws.iter_rows(
            min_row=2, max_row=ws.max_row, min_col=1, max_col=5, values_only=True
        ):
            classificacao, rubrica, conta_filho, conta_pai, cod = row
            entry = {
                "rubrica": str(rubrica or "").strip(),
                "conta_filho": str(conta_filho or "").strip(),
                "conta_pai": str(conta_pai or "").strip(),
                "cod": cod,
            }
            for chave_texto in (classificacao, rubrica):
                if chave_texto is None or not str(chave_texto).strip():
                    continue
                plano.setdefault(str(chave_texto).strip(), entry)
                codigo = _extrair_codigo_gerencial(chave_texto)
                if codigo:
                    plano.setdefault(f"@cod:{codigo}", entry)
                    familia = codigo.rsplit(".", 1)[0]
                    plano.setdefault(f"@fam:{familia}", entry)
        for codigo, entry in _PLANO_CONTAS_CODIGO_OVERRIDES.items():
            plano.setdefault(f"@cod:{codigo}", dict(entry))
        if aplicar_overrides_dre_gerado:
            for chave, entry in _PLANO_CONTAS_TEXTO_OVERRIDES.items():
                plano[chave] = dict(entry)
        return plano

    @staticmethod
    def _resolver_conta_pai(
        natureza_raw: str | None,
        rubrica_raw: str | None,
        plano: dict[str, dict],
    ) -> tuple[str, str, str, int | None]:
        """Resolve natureza/rubrica para (rubrica, conta_filho, conta_pai, cod).

        Ordem de tentativas (a primeira que casar vence):
        1. texto exato e texto após o primeiro " - " (cadastro literal);
        2. código gerencial (ex.: "1.1.1") — tolera sufixos no rótulo;
        3. família do código (ex.: "1.1") — subcontas irmãs herdam a
           classificação (1.1.2/1.1.3 somam em Receita Bruta como 1.1.1).
        """
        chaves = [
            (rubrica_raw or "").strip(),
            (natureza_raw or "").strip(),
        ]

        def _resultado(p: dict) -> tuple[str, str, str, int | None]:
            return p["rubrica"], p["conta_filho"], p["conta_pai"], p["cod"]

        # 1. Texto exato e texto após o primeiro " - ".
        for chave in chaves:
            if not chave:
                continue
            tentativas = [chave]
            if " - " in chave:
                tentativas.append(chave.split(" - ", 1)[1].strip())
            for tentativa in tentativas:
                if tentativa and tentativa in plano:
                    return _resultado(plano[tentativa])

        # 2. Código gerencial exato e 3. família (fallback para subcontas irmãs).
        for chave in chaves:
            codigo = _extrair_codigo_gerencial(chave)
            if not codigo:
                continue
            if f"@cod:{codigo}" in plano:
                return _resultado(plano[f"@cod:{codigo}"])
            familia = codigo.rsplit(".", 1)[0]
            if f"@fam:{familia}" in plano:
                return _resultado(plano[f"@fam:{familia}"])

        return "", "", "", None

    # ------------------------------------------------------------------ #
    # APOIO                                                                #
    # ------------------------------------------------------------------ #

    def _agregar_para_apoio(
        self,
        lancamentos: list[DRELancamentoDB],
        plano: dict[str, dict],
    ) -> tuple[list[list], list[int]]:
        """Agrega lançamentos por Conta Pai/Rubrica x Mês para aba APOIO.

        Retorna (linhas, meses_encontrados_ordenados).
        """
        # (cod, label) → {mes: valor}
        agregado: dict[tuple[int | None, str], dict[int, float]] = defaultdict(
            lambda: defaultdict(float)
        )
        receita_bruta_keys: set[tuple[int | None, str]] = set()
        receita_bruta_informada: dict[int, float] = defaultdict(float)
        faturamento_liquido: dict[int, float] = defaultdict(float)
        deducoes_receita: dict[int, float] = defaultdict(float)
        meses_encontrados: set[int] = set()

        for lanc in lancamentos:
            # Usar competência, não data de emissão — mesma regra do BD_FLUXO.
            data = _data_lancamento_para_date(lanc.data_lancamento)
            mes = lanc.competencia_mes or data.month
            meses_encontrados.add(mes)
            credito = float(lanc.credito or 0)
            debito = float(lanc.debito or 0)
            valor = credito - debito
            valor_bruto = float(lanc.valor_bruto or 0)

            rubrica, conta_filho, conta_pai, cod = self._resolver_conta_pai(
                lanc.natureza_raw,
                lanc.rubrica,
                plano,
            )

            eh_receita_bruta = conta_pai == _RECEITA_BRUTA_CONTA_PAI
            if eh_receita_bruta:
                valor_receita_bruta = max(valor_bruto, credito, valor)
                receita_bruta_informada[mes] += valor_receita_bruta
                faturamento_liquido[mes] += valor

                if conta_filho:
                    receita_bruta_keys.add((cod, conta_filho))
                if conta_pai:
                    receita_bruta_keys.add((cod, conta_pai))

                if rubrica:
                    agregado[(cod, rubrica)][mes] += valor
                if conta_filho:
                    agregado[(cod, conta_filho)][mes] += valor_receita_bruta
                if conta_pai:
                    agregado[(cod, conta_pai)][mes] += valor_receita_bruta
                continue

            if debito > 0 and self._eh_deducao_receita(lanc, rubrica, conta_filho, conta_pai):
                deducoes_receita[mes] += debito

            if rubrica:
                agregado[(cod, rubrica)][mes] += valor
            if conta_filho:
                agregado[(cod, conta_filho)][mes] += valor
            if conta_pai:
                agregado[(cod, conta_pai)][mes] += valor

        for mes in meses_encontrados:
            receita_bruta = max(
                receita_bruta_informada[mes],
                faturamento_liquido[mes] + deducoes_receita[mes],
            )
            for key in receita_bruta_keys:
                agregado[key][mes] = receita_bruta

        meses_ordenados = sorted(meses_encontrados)
        linhas: list[list] = []

        for (cod, label), meses_vals in sorted(
            agregado.items(),
            key=lambda x: (x[0][0] if x[0][0] is not None else 99, x[0][1]),
        ):
            # Estrutura fixa Jan..Dez para manter MATCH/VLOOKUP da DRE estável.
            row: list = [cod, label]
            total = 0.0
            for m in range(1, 13):
                val = meses_vals.get(m)
                row.append(val)
                if val:
                    total += val
            row.append(total)
            linhas.append(row)

        return linhas, meses_ordenados

    @staticmethod
    def _eh_deducao_receita(
        lanc: DRELancamentoDB,
        rubrica: str,
        conta_filho: str,
        conta_pai: str,
    ) -> bool:
        rotulos = (
            lanc.natureza_raw,
            lanc.rubrica,
            lanc.conta_pai,
            rubrica,
            conta_filho,
            conta_pai,
        )
        return any(_normalizar_rotulo(rotulo) in _IMPOSTOS_RECEITA_ROTULOS for rotulo in rotulos)

    @staticmethod
    def _saldos_painel_por_mes(lancamentos: list[DRELancamentoDB]) -> dict[int, float]:
        """Calcula o saldo com a mesma regra do Painel DRE.

        Painel: crédito líquido - (débitos - impostos). A classificação de
        imposto aqui replica a regra do painel web, que usa a rubrica persistida.
        """
        saldos: dict[int, float] = defaultdict(float)
        for lanc in lancamentos:
            data = _data_lancamento_para_date(lanc.data_lancamento)
            mes = lanc.competencia_mes or data.month
            credito = float(lanc.credito or 0)
            debito = float(lanc.debito or 0)
            imposto = (
                debito
                if _eh_imposto_saldo_painel(lanc.rubrica, lanc.natureza_raw)
                else 0.0
            )
            saldos[mes] += credito - (debito - imposto)
        return dict(saldos)

    def _escrever_apoio(
        self,
        writer: TemplateWriter,
        lancamentos: list[DRELancamentoDB],
        plano: dict[str, dict],
    ) -> list[int]:
        """Reescreve aba APOIO com dados agregados do banco. Retorna meses_encontrados."""
        linhas_apoio, meses = self._agregar_para_apoio(lancamentos, plano)

        if not linhas_apoio:
            logger.warning("Nenhum dado para escrever na APOIO")
            return meses

        # Cabeçalho fixo Jan..Dez (linha 5) para evitar duplicidade/resíduo.
        header: list = ["Rótulos de Linha", "Conta Pai", *_MESES_NOMES, "Total Geral"]

        writer.escrever_area("APOIO", [header], linha_inicio=5, coluna_inicio=1)

        # Limpar dados antigos (linhas 6..200), cobrindo toda a faixa usada
        # pela DRE (B:N para lookup) e coluna de Total Geral (O).
        writer.limpar_area("APOIO", 6, 200, 1, 2 + 12 + 1)

        # Escrever dados novos
        writer.escrever_area("APOIO", linhas_apoio, linha_inicio=6, coluna_inicio=1)

        logger.info(
            "APOIO reescrita: %d linhas, meses=%s",
            len(linhas_apoio),
            [_MESES_NOMES[m - 1] for m in meses],
        )
        return meses

    @staticmethod
    def _numero_planilha(valor: Any) -> float:
        if valor is None or valor == "":
            return 0.0
        try:
            return float(valor)
        except (TypeError, ValueError):
            return 0.0

    @classmethod
    def _valor_referencia_dre(cls, ws: Any, referencia: str) -> float:
        linha, coluna = coordinate_to_tuple(referencia.replace("$", ""))
        return cls._numero_planilha(ws.cell(row=linha, column=coluna).value)

    @classmethod
    def _somar_intervalo_dre(cls, ws: Any, referencia: str) -> float:
        min_col, min_row, max_col, max_row = range_boundaries(referencia.replace("$", ""))
        total = 0.0
        for linha in range(min_row, max_row + 1):
            for coluna in range(min_col, max_col + 1):
                total += cls._numero_planilha(ws.cell(row=linha, column=coluna).value)
        return total

    @classmethod
    def _avaliar_formula_dre(
        cls,
        *,
        ws: Any,
        formula: str,
        linha: int,
        coluna: int,
        apoio_por_rotulo: dict[str, dict[int, float]],
        apoio_por_rotulo_normalizado: dict[str, dict[int, float]],
    ) -> float | None:
        formula_limpa = formula.strip()
        formula_upper = formula_limpa.upper()

        if formula_upper.startswith("=IFERROR(VLOOKUP("):
            mes = _DRE_MES_POR_COLUNA_VALOR.get(coluna)
            if not mes:
                return 0.0
            rotulo = str(ws.cell(row=linha, column=1).value or "").strip()
            valores = apoio_por_rotulo.get(rotulo)
            if valores is None:
                valores = apoio_por_rotulo_normalizado.get(_normalizar_rotulo(rotulo))
            return valores.get(mes, 0.0) if valores else 0.0

        match_sum = _FORMULA_SUM_RE.match(formula_limpa)
        if match_sum:
            total = 0.0
            for termo in match_sum.group("args").split(","):
                referencia = termo.strip()
                if ":" in referencia:
                    total += cls._somar_intervalo_dre(ws, referencia)
                else:
                    total += cls._valor_referencia_dre(ws, referencia)
            return total

        match_div = _FORMULA_DIV_IFERROR_RE.match(formula_limpa) or _FORMULA_DIV_RE.match(
            formula_limpa
        )
        if match_div:
            numerador = cls._valor_referencia_dre(ws, match_div.group("num"))
            denominador = cls._valor_referencia_dre(ws, match_div.group("den"))
            return numerador / denominador if denominador else 0.0

        return None

    @staticmethod
    def _valor_periodo_dre(
        valores_por_mes: dict[int, float],
        coluna: int,
    ) -> float | None:
        mes = _DRE_MES_POR_COLUNA_VALOR.get(coluna)
        if mes:
            return valores_por_mes.get(mes, 0.0)

        meses_trimestre = _DRE_MESES_POR_COLUNA_TRIMESTRE.get(coluna)
        if meses_trimestre:
            return sum(valores_por_mes.get(mes, 0.0) for mes in meses_trimestre)

        if coluna == _DRE_COLUNA_ANO:
            return sum(valores_por_mes.get(mes, 0.0) for mes in range(1, 13))

        return None

    def _materializar_dre(
        self,
        writer: TemplateWriter,
        saldos_painel_por_mes: dict[int, float] | None = None,
    ) -> dict[str, Any]:
        """Resolve fórmulas do bloco gerencial da DRE para evitar cache antigo no XLSX.

        O arquivo gerado é um relatório estático: os números devem abrir corretos
        mesmo em leitores que não recalculam fórmulas. A Receita Líquida segue a
        estrutura do DRE: Receita Bruta menos deduções sobre vendas; IR é lançado
        depois, no bloco IRPJ/CSLL.
        """
        if (
            not writer._wb
            or "DRE" not in writer._wb.sheetnames
            or "APOIO" not in writer._wb.sheetnames
        ):
            return {"celulas_materializadas": 0}

        ws_dre = writer._wb["DRE"]
        ws_apoio = writer._wb["APOIO"]

        apoio_por_rotulo: dict[str, dict[int, float]] = {}
        apoio_por_rotulo_normalizado: dict[str, dict[int, float]] = {}
        for linha in range(6, ws_apoio.max_row + 1):
            rotulo = str(ws_apoio.cell(row=linha, column=2).value or "").strip()
            if not rotulo:
                continue
            valores = {
                mes: self._numero_planilha(ws_apoio.cell(row=linha, column=2 + mes).value)
                for mes in range(1, 13)
            }
            apoio_por_rotulo.setdefault(rotulo, valores)
            apoio_por_rotulo_normalizado.setdefault(_normalizar_rotulo(rotulo), valores)

        linhas_por_rotulo: dict[str, int] = {}
        for linha in range(1, ws_dre.max_row + 1):
            rotulo = str(ws_dre.cell(row=linha, column=1).value or "").strip()
            if rotulo:
                linhas_por_rotulo.setdefault(rotulo, linha)

        linha_faturamento = linhas_por_rotulo.get(_FATURAMENTO_ROTULO)
        linha_receita_liquida = linhas_por_rotulo.get(_RECEITA_LIQUIDA_ROTULO)
        linha_resultado_liquido = linhas_por_rotulo.get(_RESULTADO_LIQUIDO_ROTULO)
        linha_resultado_gerencial = linhas_por_rotulo.get(_RESULTADO_GERENCIAL_ROTULO)
        saldos_painel = saldos_painel_por_mes or {}
        celulas_materializadas = 0

        for linha in range(6, ws_dre.max_row + 1):
            for coluna in _DRE_COLUNAS_VALOR:
                celula = ws_dre.cell(row=linha, column=coluna)
                valor_atual = celula.value
                valor_calculado: float | None = None

                if linha == linha_resultado_liquido and saldos_painel:
                    valor_calculado = self._valor_periodo_dre(saldos_painel, coluna)
                elif isinstance(valor_atual, str) and valor_atual.startswith("="):
                    valor_calculado = self._avaliar_formula_dre(
                        ws=ws_dre,
                        formula=valor_atual,
                        linha=linha,
                        coluna=coluna,
                        apoio_por_rotulo=apoio_por_rotulo,
                        apoio_por_rotulo_normalizado=apoio_por_rotulo_normalizado,
                    )

                if valor_calculado is not None:
                    celula.value = valor_calculado
                    celulas_materializadas += 1

        for linha in range(6, ws_dre.max_row + 1):
            for coluna in _DRE_COLUNAS_PERCENTUAL:
                celula = ws_dre.cell(row=linha, column=coluna)
                valor_atual = celula.value
                if not isinstance(valor_atual, str) or not valor_atual.startswith("="):
                    continue
                valor_calculado = self._avaliar_formula_dre(
                    ws=ws_dre,
                    formula=valor_atual,
                    linha=linha,
                    coluna=coluna,
                    apoio_por_rotulo=apoio_por_rotulo,
                    apoio_por_rotulo_normalizado=apoio_por_rotulo_normalizado,
                )
                if valor_calculado is not None:
                    celula.value = valor_calculado
                    celulas_materializadas += 1

        writer._modified_sheets.add("DRE")
        logger.info("DRE materializada: %d células calculadas", celulas_materializadas)
        return {
            "celulas_materializadas": celulas_materializadas,
            "linha_faturamento": linha_faturamento,
            "linha_receita_liquida": linha_receita_liquida,
            "linha_resultado_liquido": linha_resultado_liquido,
            "linha_resultado_gerencial": linha_resultado_gerencial,
        }

    # ------------------------------------------------------------------ #
    # BD_FLUXO                                                             #
    # ------------------------------------------------------------------ #

    @staticmethod
    def _converte_linha_bd_fluxo(lanc: DRELancamentoDB) -> list:
        """Converte lançamento para colunas A-G da aba BD_FLUXO."""
        data_val: date | str = lanc.data_lancamento
        if isinstance(data_val, str):
            data_val = _data_lancamento_para_date(data_val)

        credito = float(lanc.credito) if lanc.credito else None
        debito = float(lanc.debito) if lanc.debito else None

        return [
            data_val,  # A - Data
            lanc.historico,  # B - Histórico
            credito if credito else None,  # C - Crédito
            debito if debito else None,  # D - Débito
            None,  # E - Saldo (fórmula externa)
            lanc.natureza_raw or "",  # F - Natureza (C. gerencial para VLOOKUP)
            lanc.centro_custo or "",  # G - Centro de Custo
        ]

    def _converte_linha_bd_fluxo_expandida(
        self,
        lanc: DRELancamentoDB,
        plano: dict[str, dict],
    ) -> list:
        """Converte lançamento para colunas A-R (com campos derivados materializados).

        A materialização de H:R evita depender de recálculo de fórmula para alimentar
        pivots/slicers do template.
        """
        data_val: date | str = lanc.data_lancamento
        if isinstance(data_val, str):
            data_val = _data_lancamento_para_date(data_val)

        credito = float(lanc.credito) if lanc.credito else None
        debito = float(lanc.debito) if lanc.debito else None
        valor = (credito or 0.0) - (debito or 0.0)
        # Ano/Mês do fluxo seguem a competência do lançamento, não a data de emissão.
        # Relatórios mensais podem conter linhas datadas no mês seguinte (ex.:
        # recebimentos tardios) e devem contar para a competência do upload.
        ano_fluxo = lanc.competencia_ano or data_val.year
        mes_fluxo = lanc.competencia_mes or data_val.month
        mes_nome = _MESES_NOMES[mes_fluxo - 1]

        rubrica, conta_filho, conta_pai, cod = self._resolver_conta_pai(
            lanc.natureza_raw,
            lanc.rubrica,
            plano,
        )
        rubrica_final = rubrica or (lanc.rubrica or "").strip() or None
        conta_pai_final = conta_pai or (lanc.conta_pai or "").strip() or None

        return [
            data_val,  # A - Data
            lanc.historico,  # B - Histórico
            credito if credito else None,  # C - Crédito
            debito if debito else None,  # D - Débito
            valor if valor else None,  # E - Saldo
            lanc.natureza_raw or "",  # F - Natureza
            lanc.centro_custo or "",  # G - Centro de Custo - Obra
            ano_fluxo,  # H - Ano Fluxo
            mes_fluxo,  # I - C. Mês
            mes_nome,  # J - Mês
            None,  # K - Banco
            None,  # L - Empresa
            valor if valor else None,  # M - Valor
            rubrica_final,  # N - Rubrica
            conta_filho or None,  # O - Conta Filho
            conta_pai_final,  # P - Conta Pai
            cod,  # Q - Cod
            ano_fluxo,  # R - Ano
        ]

    def _limpar_bd_fluxo(self, writer: TemplateWriter) -> None:
        writer.limpar_area(
            "BD_FLUXO", linha_inicio=2, linha_fim=4964, coluna_inicio=1, coluna_fim=18
        )

    # ------------------------------------------------------------------ #
    # Aba adicional de detalhamento                                        #
    # ------------------------------------------------------------------ #

    def _escrever_aba_detalhamento_mensal(
        self,
        writer: TemplateWriter,
        lancamentos: list[DRELancamentoDB],
        plano: dict[str, dict],
        competencia: str,
        estrategia_meses: str,
        meses_utilizados: list[int],
    ) -> dict[str, Any]:
        """Cria aba analítica com detalhe mensal e granular dos lançamentos."""
        if not writer._wb:
            raise RuntimeError("Template não aberto.")

        wb = writer._wb
        nome_aba = "DETALHE_MENSAL_DB"
        if nome_aba in wb.sheetnames:
            del wb[nome_aba]

        ws = wb.create_sheet(title=nome_aba)
        writer._modified_sheets.add(nome_aba)

        resumo_mensal: dict[int, dict[str, Any]] = defaultdict(
            lambda: {
                "lancamentos": 0,
                "credito": 0.0,
                "debito": 0.0,
                "saldo": 0.0,
                "centros": set(),
                "rubricas": set(),
                "contas_pai": set(),
            }
        )
        detalhe_agrupado: dict[tuple, dict[str, Any]] = defaultdict(
            lambda: {"lancamentos": 0, "credito": 0.0, "debito": 0.0, "saldo": 0.0}
        )
        lancamentos_rows: list[list[Any]] = []

        for lanc in sorted(lancamentos, key=lambda x: (x.data_lancamento, x.historico or "")):
            data_val = _data_lancamento_para_date(lanc.data_lancamento)
            # Competência é a fonte da verdade para agrupamento mensal.
            ano = lanc.competencia_ano or data_val.year
            mes = lanc.competencia_mes or data_val.month
            mes_nome = _MESES_NOMES[mes - 1]
            credito = float(lanc.credito) if lanc.credito else 0.0
            debito = float(lanc.debito) if lanc.debito else 0.0
            saldo = credito - debito

            rubrica_map, conta_filho, conta_pai_map, cod = self._resolver_conta_pai(
                lanc.natureza_raw,
                lanc.rubrica,
                plano,
            )
            rubrica = (rubrica_map or lanc.rubrica or "").strip() or "SEM_RUBRICA"
            conta_filho_final = (conta_filho or "").strip() or "SEM_CONTA_FILHO"
            conta_pai = (conta_pai_map or lanc.conta_pai or "").strip() or "SEM_CONTA_PAI"
            centro_custo = (lanc.centro_custo or "").strip() or "SEM_CENTRO_CUSTO"
            natureza = (lanc.natureza_raw or "").strip() or "SEM_NATUREZA"

            resumo = resumo_mensal[mes]
            resumo["lancamentos"] += 1
            resumo["credito"] += credito
            resumo["debito"] += debito
            resumo["saldo"] += saldo
            resumo["centros"].add(centro_custo)
            resumo["rubricas"].add(rubrica)
            resumo["contas_pai"].add(conta_pai)

            chave = (ano, mes, mes_nome, centro_custo, conta_pai, conta_filho_final, rubrica, cod)
            detalhe = detalhe_agrupado[chave]
            detalhe["lancamentos"] += 1
            detalhe["credito"] += credito
            detalhe["debito"] += debito
            detalhe["saldo"] += saldo

            lancamentos_rows.append(
                [
                    data_val,
                    ano,
                    mes,
                    mes_nome,
                    _EMPRESA_PADRAO,
                    centro_custo,
                    natureza,
                    rubrica,
                    conta_filho_final,
                    conta_pai,
                    cod,
                    lanc.historico,
                    credito if credito else None,
                    debito if debito else None,
                    saldo if saldo else None,
                    lanc.upload_id,
                    lanc.linha_origem,
                    lanc.hash_linha,
                ]
            )

        resumo_rows: list[list[Any]] = []
        for mes in sorted(resumo_mensal.keys()):
            base = resumo_mensal[mes]
            resumo_rows.append(
                [
                    mes,
                    _MESES_NOMES[mes - 1],
                    base["lancamentos"],
                    base["credito"],
                    base["debito"],
                    base["saldo"],
                    len(base["centros"]),
                    len(base["rubricas"]),
                    len(base["contas_pai"]),
                ]
            )

        detalhe_rows: list[list[Any]] = []
        for chave in sorted(
            detalhe_agrupado.keys(),
            key=lambda c: (c[0], c[1], c[3], c[4], c[5], c[6]),
        ):
            ano, mes, mes_nome, centro_custo, conta_pai, conta_filho, rubrica, cod = chave
            base = detalhe_agrupado[chave]
            detalhe_rows.append(
                [
                    ano,
                    mes,
                    mes_nome,
                    centro_custo,
                    conta_pai,
                    conta_filho,
                    rubrica,
                    cod,
                    base["lancamentos"],
                    base["credito"],
                    base["debito"],
                    base["saldo"],
                ]
            )

        title_font = Font(bold=True, size=12)
        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="D9E1F2")

        def _escrever_secao(
            linha_inicio: int,
            titulo: str,
            cabecalho: list[str],
            linhas: list[list[Any]],
        ) -> int:
            ws.cell(row=linha_inicio, column=1, value=titulo).font = title_font
            cab_row = linha_inicio + 1
            for col_idx, col_name in enumerate(cabecalho, start=1):
                cell = ws.cell(row=cab_row, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill

            dados_inicio = cab_row + 1
            for lin_idx, linha in enumerate(linhas, start=dados_inicio):
                for col_idx, valor in enumerate(linha, start=1):
                    ws.cell(row=lin_idx, column=col_idx, value=valor)

            dados_fim = dados_inicio + max(len(linhas) - 1, 0)
            if linhas:
                ws.auto_filter.ref = f"A{cab_row}:{chr(64 + min(len(cabecalho), 26))}{dados_fim}"
            return dados_fim + 2

        ws.cell(
            row=1, column=1, value="DRE - Detalhamento mensal (gerado do banco)"
        ).font = title_font
        ws.cell(row=2, column=1, value=f"Competência: {competencia}")
        ws.cell(row=2, column=4, value=f"Estratégia: {estrategia_meses}")
        ws.cell(
            row=2,
            column=7,
            value=f"Meses utilizados: {', '.join(str(m).zfill(2) for m in meses_utilizados)}",
        )

        linha = 4
        linha = _escrever_secao(
            linha,
            "Resumo mensal",
            [
                "Mes",
                "Mes Nome",
                "Lancamentos",
                "Credito Total",
                "Debito Total",
                "Saldo Liquido",
                "Centros de Custo (qtd)",
                "Rubricas (qtd)",
                "Contas Pai (qtd)",
            ],
            resumo_rows,
        )
        linha = _escrever_secao(
            linha,
            "Detalhe mensal por centro de custo/conta/rubrica",
            [
                "Ano",
                "Mes",
                "Mes Nome",
                "Centro de Custo",
                "Conta Pai",
                "Conta Filho",
                "Rubrica",
                "Cod",
                "Lancamentos",
                "Credito Total",
                "Debito Total",
                "Saldo Liquido",
            ],
            detalhe_rows,
        )
        colunas_granulares = [
            "Data Lancamento",
            "Ano",
            "Mes",
            "Mes Nome",
            "Empresa",
            "Centro de Custo - Obra",
            "Natureza Raw",
            "Rubrica",
            "Conta Filho",
            "Conta Pai",
            "Cod",
            "Historico",
            "Credito",
            "Debito",
            "Saldo Liquido",
            "Upload ID",
            "Linha Origem",
            "Hash Linha",
        ]
        linha_inicio_granular = linha
        linha_cabecalho_granular = linha_inicio_granular + 1
        linha_dados_granular = linha_cabecalho_granular + 1
        linha_ultima_granular = max(
            linha_cabecalho_granular + 1,
            linha_dados_granular + len(lancamentos_rows) - 1,
        )
        _ = _escrever_secao(
            linha_inicio_granular,
            "Lancamentos granulares",
            colunas_granulares,
            lancamentos_rows,
        )

        # Largura mínima para leitura operacional no painel.
        larguras = {
            "A": 14,
            "B": 10,
            "C": 8,
            "D": 12,
            "E": 16,
            "F": 24,
            "G": 28,
            "H": 24,
            "I": 24,
            "J": 28,
            "K": 10,
            "L": 42,
            "M": 14,
            "N": 14,
            "O": 16,
            "P": 40,
            "Q": 12,
            "R": 34,
        }
        for col, width in larguras.items():
            ws.column_dimensions[col].width = width

        ws.freeze_panes = "A6"

        return {
            "aba_detalhamento": nome_aba,
            "linhas_resumo_mensal": len(resumo_rows),
            "linhas_resumo_agrupado": len(detalhe_rows),
            "linhas_lancamentos_granulares": len(lancamentos_rows),
            "tabela_granular_ref": f"A{linha_cabecalho_granular}:R{linha_ultima_granular}",
            "tabela_granular_colunas": colunas_granulares,
        }

    # ------------------------------------------------------------------ #
    # Visibilidade de colunas na aba DRE                                   #
    # ------------------------------------------------------------------ #

    def _controlar_visibilidade_colunas_dre(
        self,
        writer: TemplateWriter,
        meses_com_dados: list[int],
    ) -> dict[str, Any]:
        """Oculta/exibe pares de colunas (valor + %) de meses na aba DRE."""
        if "DRE" not in writer.listar_sheets():
            logger.warning("Aba 'DRE' não encontrada no template; visibilidade de colunas ignorada")
            return {
                "meses_visiveis": [],
                "meses_ocultos": [],
                "colunas_visiveis": [],
                "colunas_ocultas": [],
            }

        meses_com_dados_set = set(meses_com_dados)
        # Mapeamento estável do template DRE: mês -> (col valor, col %)
        pares_mensais: dict[int, tuple[str, str]] = {
            1: ("B", "C"),
            2: ("D", "E"),
            3: ("F", "G"),
            4: ("J", "K"),
            5: ("L", "M"),
            6: ("N", "O"),
            7: ("R", "S"),
            8: ("T", "U"),
            9: ("V", "W"),
            10: ("Z", "AA"),
            11: ("AB", "AC"),
            12: ("AD", "AE"),
        }
        colunas_trimestre = ["H", "I", "P", "Q", "X", "Y", "AF", "AG"]
        colunas_ano = ["AH", "AI"]

        colunas_ocultar: list[str] = []
        colunas_exibir: list[str] = []
        meses_ocultos: list[int] = []
        meses_visiveis: list[int] = []

        for mes in range(1, 13):
            col_val, col_pct = pares_mensais[mes]
            if mes in meses_com_dados_set:
                colunas_exibir.extend([col_val, col_pct])
                meses_visiveis.append(mes)
            else:
                colunas_ocultar.extend([col_val, col_pct])
                meses_ocultos.append(mes)

        # Mantém trimestres ocultos e ANO sempre visível.
        colunas_ocultar.extend(colunas_trimestre)
        colunas_exibir.extend([c for c in colunas_ano if c not in colunas_exibir])

        # Patch direto de colunas no XML da DRE, sem reserializar a aba inteira.
        # Mantém o layout do template e controla visibilidade só pelo atributo hidden.
        writer.ocultar_colunas_xml_patch("DRE", colunas_ocultar, True)
        writer.ocultar_colunas_xml_patch("DRE", colunas_exibir, False)

        logger.info(
            "Visibilidade DRE: meses visíveis=%s, meses ocultos=%s",
            [_MESES_NOMES[m - 1] for m in meses_visiveis],
            [_MESES_NOMES[m - 1] for m in meses_ocultos],
        )

        return {
            "meses_visiveis": meses_visiveis,
            "meses_ocultos": meses_ocultos,
            "colunas_visiveis": colunas_exibir,
            "colunas_ocultas": colunas_ocultar,
        }

    # ------------------------------------------------------------------ #
    # Interface pública                                                    #
    # ------------------------------------------------------------------ #

    def verificar_dados(
        self,
        competencia: str,
        centro_custo: str | None = None,
        meses_incluir: list[int] | None = None,
        ano_todo: bool = False,
    ) -> dict[str, Any]:
        """Verifica se há dados suficientes para geração.

        Comportamentos:
          - padrão: usa meses disponíveis <= competência (sem exigir cumulativo completo)
          - meses_incluir: valida meses selecionados manualmente
          - ano_todo: usa todos os meses disponíveis do ano
        """
        try:
            ano, mes = _parse_competencia(competencia)
        except ValueError as e:
            return {"valido": False, "error": str(e)}

        try:
            meses_disponiveis, meses_utilizados, estrategia = self._resolver_meses_para_geracao(
                ano=ano,
                mes_alvo=mes,
                meses_incluir=meses_incluir,
                ano_todo=ano_todo,
            )
        except ValueError as exc:
            return {
                "valido": False,
                "competencia": competencia,
                "error": str(exc),
            }

        resumo = self.repository.lancamentos.get_resumo_competencia(ano, mes)

        return {
            "valido": True,
            "competencia": competencia,
            "ano": ano,
            "mes": mes,
            "meses_disponiveis": meses_disponiveis,
            "meses_utilizados": meses_utilizados,
            "estrategia_meses": estrategia,
            "ano_todo": ano_todo,
            "meses_incluir": self._normalizar_meses_incluir(meses_incluir),
            "total_lancamentos_mes_alvo": resumo.total_lancamentos if resumo else 0,
            "total_credito_mes_alvo": float(resumo.total_credito) if resumo else 0.0,
            "total_debito_mes_alvo": float(resumo.total_debito) if resumo else 0.0,
        }

    def gerar_arquivo(
        self,
        competencia: str,
        centro_custo: str | None = None,
        output_path: Path | None = None,
        meses_incluir: list[int] | None = None,
        ano_todo: bool = False,
    ) -> dict[str, Any]:
        """Gera arquivo DRE completo preenchendo BD_FLUXO e visibilidade DRE.

        Args:
            competencia: Competência alvo no formato MM/AAAA
            centro_custo: Filtro opcional por obra/centro de custo
            output_path: Caminho de saída (gerado automaticamente se None)

        Returns:
            Dict com metadados da geração incluindo meses_utilizados,
            meses_ocultos, colunas_dre_visiveis e fonte_dados.

        Raises:
            ValueError: Se não há upload completed para o mês alvo.
        """
        ano, mes = _parse_competencia(competencia)

        # 1. Resolver meses conforme estratégia
        meses_disponiveis, meses_utilizados, estrategia_meses = self._resolver_meses_para_geracao(
            ano=ano,
            mes_alvo=mes,
            meses_incluir=meses_incluir,
            ano_todo=ano_todo,
        )
        logger.info(
            "Geração DRE %s — estratégia=%s — meses disponíveis=%s — meses utilizados=%s",
            competencia,
            estrategia_meses,
            meses_disponiveis,
            meses_utilizados,
        )

        # 2. Buscar lançamentos dos meses efetivos
        lancamentos = self._get_lancamentos_ytd(ano, meses_utilizados, centro_custo)
        if not lancamentos:
            raise ValueError(
                f"Nenhum lançamento encontrado para os meses {meses_utilizados} em {ano}."
            )

        # 3. Caminho de saída
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = (
                settings.base_dir
                / "output"
                / f"DRE_AIDEAL_{competencia.replace('/', '-')}_{timestamp}.xlsx"
            )
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # 4. Preencher template
        visibilidade: dict[str, Any] = {}
        detalhe_meta: dict[str, Any] = {}
        dre_meta: dict[str, Any] = {}
        with TemplateWriter(self.template_path) as writer:
            plano = self._ler_plano_contas(writer, aplicar_overrides_dre_gerado=True)

            # 4a. Limpar e escrever BD_FLUXO (A:R) já com campos derivados
            # materializados. Isso remove dependência de recálculo intermediário.
            linhas_bd = [
                self._converte_linha_bd_fluxo_expandida(lanc, plano) for lanc in lancamentos
            ]
            writer.registrar_sheet_data_override(
                "BD_FLUXO",
                linhas_bd,
                linha_inicio=2,
                coluna_inicio=1,
                limpar_area=(2, 4964, 1, 18),
            )
            # Ajusta ref da tabela BD_FLUXO1 para o volume real.
            ultima_linha_bd = 1 + len(linhas_bd)
            writer.ajustar_tabela_range("BD_FLUXO", "BD_FLUXO1", linha_fim=ultima_linha_bd)
            writer.remover_formulas_calculadas_tabela("BD_FLUXO1")
            writer.remover_slicers()
            writer.esvaziar_pivot_cache_records()
            # Modo Excel-safe: remove charts, pivots, drawings de abas hidden
            # e demais artefatos OOXML complexos que causam prompt de reparo no
            # Excel Desktop. Preserva DRE, BD_FLUXO, DETALHE_MENSAL_DB e logo.
            writer.ativar_modo_excel_safe()

            # 4b. Reescrever APOIO para manter a DRE alimentada por VLOOKUP.
            # Sem isso, o cabeçalho de meses em APOIO fica desatualizado (template)
            # e a DRE tende a retornar 0 via IFERROR.
            meses_apoio = self._escrever_apoio(writer, lancamentos, plano)

            # 4c. Meses efetivos visíveis no painel DRE.
            meses_utilizados_planilha = sorted(set(meses_apoio or meses_utilizados))

            # 4d. Preencher filtro de centro de custo no Painel (se fornecido)
            if centro_custo:
                try:
                    writer.escrever_celula("Painel", row=2, col=3, valor=centro_custo)
                except Exception as e:
                    logger.warning("Não foi possível preencher filtro de centro de custo: %s", e)

            # 4e. Controlar visibilidade de colunas na aba DRE
            visibilidade = self._controlar_visibilidade_colunas_dre(
                writer, meses_utilizados_planilha
            )

            # 4f. Materializar a DRE para evitar cache antigo de fórmulas no XLSX.
            dre_meta = self._materializar_dre(
                writer,
                saldos_painel_por_mes=self._saldos_painel_por_mes(lancamentos),
            )

            # 4g. Aba de detalhamento para uso operacional no workspace/painel.
            detalhe_meta = self._escrever_aba_detalhamento_mensal(
                writer=writer,
                lancamentos=lancamentos,
                plano=plano,
                competencia=competencia,
                estrategia_meses=estrategia_meses,
                meses_utilizados=meses_utilizados_planilha,
            )
            tabela_ref = detalhe_meta.get("tabela_granular_ref")
            tabela_colunas = detalhe_meta.get("tabela_granular_colunas")
            if tabela_ref and tabela_colunas:
                writer.registrar_table_ooxml(
                    sheet_name="DETALHE_MENSAL_DB",
                    table_name="DETALHE_MENSAL_DB",
                    ref=tabela_ref,
                    colunas=list(tabela_colunas),
                )

            # 4h. Salvar
            writer.salvar(output_path)

        total_credito = sum(float(lanc.credito) for lanc in lancamentos)
        total_debito = sum(float(lanc.debito) for lanc in lancamentos)

        logger.info(
            "DRE gerado: %s | competência=%s | lançamentos=%d | meses_utilizados=%s",
            output_path.name,
            competencia,
            len(lancamentos),
            [_MESES_NOMES[m - 1] for m in visibilidade.get("meses_visiveis", [])],
        )

        return {
            "success": True,
            "competencia": competencia,
            "ano": ano,
            "mes": mes,
            "arquivo_saida": output_path.name,
            "arquivo_path": str(output_path),
            "registros_reais": len(lancamentos),
            "total_lancamentos": len(lancamentos),
            "total_credito": total_credito,
            "total_debito": total_debito,
            "saldo_liquido": total_credito - total_debito,
            "centro_custo": centro_custo,
            "fonte_dados": "db",
            "meses_disponiveis": meses_disponiveis,
            "meses_utilizados": visibilidade.get("meses_visiveis", []),
            "meses_solicitados": meses_utilizados,
            "estrategia_meses": estrategia_meses,
            "ano_todo": ano_todo,
            "meses_incluir": self._normalizar_meses_incluir(meses_incluir),
            "meses_ocultos": visibilidade.get("meses_ocultos", []),
            "colunas_dre_visiveis": visibilidade.get("colunas_visiveis", []),
            "aba_detalhamento": detalhe_meta.get("aba_detalhamento"),
            "linhas_resumo_mensal": detalhe_meta.get("linhas_resumo_mensal", 0),
            "linhas_resumo_agrupado": detalhe_meta.get("linhas_resumo_agrupado", 0),
            "linhas_lancamentos_granulares": detalhe_meta.get("linhas_lancamentos_granulares", 0),
            "celulas_dre_materializadas": dre_meta.get("celulas_materializadas", 0),
        }
