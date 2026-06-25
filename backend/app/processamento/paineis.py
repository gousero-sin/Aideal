"""Agregações analíticas para os painéis DRE e Fluxo de Caixa."""

from __future__ import annotations

import logging
import re
import unicodedata
from datetime import date
from decimal import Decimal
from typing import Any

from openpyxl.utils import column_index_from_string

from ..config import settings
from ..contracts.fluxo_caixa import FCMovimento, eh_transferencia_classificacao
from ..db.connection import DatabaseConnection
from ..repository.dre_indicadores_manuais import DREIndicadoresManuaisRepository
from ..repository.fluxo_indicadores_manuais import FluxoIndicadoresManuaisRepository
from ..repository.fluxo_repository import FluxoCaixaRepository
from ..templates.writer import TemplateWriter
from .dre_geracao_completa import DREGeracaoCompletaService, _extrair_codigo_gerencial
from .fluxo_caixa import FluxoCaixaProcessamentoService

MESES_LABEL = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]

logger = logging.getLogger(__name__)


def _validar_ano(ano: int) -> None:
    if ano < 2000 or ano > 2100:
        raise ValueError("Ano deve estar entre 2000 e 2100.")


def _validar_meses(meses: list[int]) -> None:
    invalidos = [mes for mes in meses if mes < 1 or mes > 12]
    if invalidos:
        raise ValueError("Mês deve estar entre 1 e 12.")


def _normalizar_lista_texto(valores: list[str] | None) -> list[str]:
    return [valor.strip() for valor in valores or [] if valor and valor.strip()]


def _float(value: Any) -> float:
    return float(value or 0)


def _percent(numerador: float, denominador: float) -> float:
    return (numerador / denominador) * 100 if denominador else 0


def _placeholders(total: int) -> str:
    return ",".join("?" for _ in range(total))


def _append_in(
    where: list[str],
    params: list[Any],
    expression: str,
    values: list[Any],
) -> None:
    if not values:
        return
    where.append(f"{expression} IN ({_placeholders(len(values))})")
    params.extend(values)


def _normalizar_chave_conta(valor: Any) -> str:
    texto = str(valor or "").strip()
    if not texto:
        return ""
    texto = re.sub(r"\([^)]*%\)\s*;?", "", texto)
    texto = re.sub(r"^\s*\d+(?:\.\d+)+(?:\s*[-.]?\s*|\s+)", "", texto)
    texto = re.sub(r"^\s*\d+\s*-\s*", "", texto)
    texto = texto.replace("°", "")
    sem_acento = unicodedata.normalize("NFKD", texto)
    sem_acento = "".join(ch for ch in sem_acento if not unicodedata.combining(ch))
    sem_pontuacao = re.sub(r"[^A-Za-z0-9]+", " ", sem_acento)
    return re.sub(r"\s+", " ", sem_pontuacao).strip().upper()


def _alias_set(valores: list[str]) -> set[str]:
    return {_normalizar_chave_conta(valor) for valor in valores if _normalizar_chave_conta(valor)}


def _competencia_label(ano: int, mes: int) -> str:
    return f"{mes:02d}/{ano}"


def _competencia_curta(ano: int, mes: int) -> str:
    return f"{MESES_LABEL[mes - 1]}/{str(ano)[-2:]}"


def _indicador_calculado(
    indicador_id: str,
    nome: str,
    valor: float | None,
    percentual: float | None = None,
    componentes: dict[str, float] | None = None,
    ideal: str | None = None,
) -> dict[str, Any]:
    return {
        "id": indicador_id,
        "nome": nome,
        "status": "calculado",
        "valor": valor,
        "percentual": percentual,
        "ideal": ideal,
        "componentes": componentes or {},
        "componentes_faltantes": [],
    }


def _indicador_indisponivel(
    indicador_id: str,
    nome: str,
    componentes_faltantes: list[str],
    componentes: dict[str, float] | None = None,
    ideal: str | None = None,
) -> dict[str, Any]:
    return {
        "id": indicador_id,
        "nome": nome,
        "status": "indisponivel",
        "valor": None,
        "percentual": None,
        "ideal": ideal,
        "componentes": componentes or {},
        "componentes_faltantes": componentes_faltantes,
    }


class _PainelBaseService:
    def __init__(self, db: DatabaseConnection | None = None) -> None:
        self.db = db or DatabaseConnection()

    def _ano_mais_recente(self, table: str) -> int:
        with self.db.get_connection() as conn:
            row = conn.execute(f"SELECT MAX(competencia_ano) AS ano FROM {table}").fetchone()
        return int(row["ano"]) if row and row["ano"] is not None else date.today().year

    @staticmethod
    def _periodo_payload(
        ano: int,
        meses_aplicados: list[int],
        meses_disponiveis: list[int],
    ) -> dict[str, Any]:
        meses_periodo = meses_aplicados or meses_disponiveis
        if meses_periodo:
            primeiro_mes = MESES_LABEL[meses_periodo[0] - 1]
            ultimo_mes = MESES_LABEL[meses_periodo[-1] - 1]
            label = f"{primeiro_mes}-{ultimo_mes}/{ano}"
            if len(meses_periodo) == 1:
                label = f"{MESES_LABEL[meses_periodo[0] - 1]}/{ano}"
        else:
            label = str(ano)
        return {
            "ano": ano,
            "meses": meses_periodo,
            "meses_disponiveis": meses_disponiveis,
            "label": label,
        }

    @staticmethod
    def _periodo_intervalo_payload(
        ano_ref: int,
        competencias: list[tuple[int, int]],
    ) -> dict[str, Any]:
        if not competencias:
            return {
                "ano": ano_ref,
                "meses": [],
                "meses_disponiveis": [],
                "label": str(ano_ref),
                "escopo": "projeto_completo",
                "inicio": None,
                "fim": None,
                "meses_com_dados": [],
            }

        ordenadas = sorted(set(competencias))
        primeiro_ano, primeiro_mes = ordenadas[0]
        ultimo_ano, ultimo_mes = ordenadas[-1]
        inicio = _competencia_label(primeiro_ano, primeiro_mes)
        fim = _competencia_label(ultimo_ano, ultimo_mes)
        label = _competencia_curta(primeiro_ano, primeiro_mes)
        if (primeiro_ano, primeiro_mes) != (ultimo_ano, ultimo_mes):
            label = f"{label}-{_competencia_curta(ultimo_ano, ultimo_mes)}"

        return {
            "ano": ano_ref,
            "meses": [mes for _, mes in ordenadas],
            "meses_disponiveis": [mes for _, mes in ordenadas],
            "label": label,
            "escopo": "projeto_completo",
            "inicio": inicio,
            "fim": fim,
            "meses_com_dados": [
                {
                    "ano": ano,
                    "mes": mes,
                    "periodo": f"{ano}-{mes:02d}",
                    "label": _competencia_curta(ano, mes),
                }
                for ano, mes in ordenadas
            ],
        }

    @staticmethod
    def _series_rows(rows: list[Any], count_key: str) -> list[dict[str, Any]]:
        series = []
        for row in rows:
            ano = int(row["ano"]) if "ano" in row.keys() and row["ano"] is not None else None
            mes = int(row["mes"])
            debito = _float(row["debito"])
            item = {
                "mes": int(row["mes"]),
                "mes_label": _competencia_curta(ano, mes) if ano else MESES_LABEL[mes - 1],
                "credito": _float(row["credito"]),
                "debito": debito,
                "saldo": _float(row["saldo"]),
                count_key: int(row["total"] or 0),
            }
            # Painel DRE: credito já representa receita líquida; impostos ficam
            # separados e saem das despesas operacionais para evitar dupla baixa.
            if "impostos" in row.keys() and row["impostos"] is not None:
                impostos = _float(row["impostos"])
                item["impostos"] = impostos
                item["receita_liquida"] = (
                    _float(row["receita_liquida"])
                    if "receita_liquida" in row.keys()
                    else _float(row["credito"])
                )
                item["saidas_liquidas"] = (
                    _float(row["saidas_liquidas"])
                    if "saidas_liquidas" in row.keys()
                    else debito - impostos
                )
            if ano:
                item["ano"] = ano
                item["periodo"] = f"{ano}-{mes:02d}"
            series.append(item)
        return series

    @staticmethod
    def _ranking_rows(rows: list[Any], count_key: str) -> list[dict[str, Any]]:
        return [
            {
                "nome": row["nome"],
                "credito": _float(row["credito"]),
                "debito": _float(row["debito"]),
                "saldo": _float(row["saldo"]),
                count_key: int(row["total"] or 0),
            }
            for row in rows
        ]

    @staticmethod
    def _filter_option_rows(rows: list[Any]) -> list[dict[str, Any]]:
        return [
            {
                "value": row["value"],
                "label": row["value"],
                "total": int(row["total"] or 0),
                "saldo": _float(row["saldo"]),
            }
            for row in rows
            if row["value"]
        ]


class PainelDREService(_PainelBaseService):
    """Agrega KPIs, séries e slicers web para DRE."""

    def __init__(self, db: DatabaseConnection | None = None) -> None:
        super().__init__(db)
        self.indicadores_manuais = DREIndicadoresManuaisRepository(self.db)
        self._plano_contas_gerado_cache: dict[str, dict] | None = None

    CENTRO_CUSTO_EXPR = "COALESCE(NULLIF(TRIM(centro_custo), ''), 'Sem obra')"
    NATUREZA_EXPR = (
        "COALESCE(NULLIF(TRIM(natureza_norm), ''), NULLIF(TRIM(natureza_raw), ''), 'Sem natureza')"
    )
    # Rubricas tratadas como impostos/deduções (não compõem a saída operacional).
    RUBRICAS_IMPOSTO = (
        "IR",
        "IR Retido",
        "ISS",
        "ISS Retido",
        "INSS",
        "INSS Retido",
        "PIS",
        "COFINS",
        "CSLL",
        "Tarifa de Antecipação",
        "Impostos sobre vendas",
        "Deduções sobre vendas",
        "(-)Deduções sobre vendas",
        "Descontos sobre vendas",
        "Simples Nacional",
    )
    IMPOSTO_DEBITO_EXPR = (
        "CASE WHEN TRIM(rubrica) IN "
        "('IR','IR Retido','ISS','ISS Retido','INSS','INSS Retido','PIS','COFINS','CSLL',"
        "'Tarifa de Antecipação','Impostos sobre vendas','Deduções sobre vendas',"
        "'(-)Deduções sobre vendas','Descontos sobre vendas','Simples Nacional') "
        "OR TRIM(rubrica) IN ('17.2','17.3','17.4','17.5','17.7','17.8') "
        "OR TRIM(natureza_raw) IN ('17.2','17.3','17.4','17.5','17.7','17.8') "
        "OR TRIM(rubrica) LIKE '17.2 - %' "
        "OR TRIM(rubrica) LIKE '17.3 - %' "
        "OR TRIM(rubrica) LIKE '17.4 - %' "
        "OR TRIM(rubrica) LIKE '17.5 - %' "
        "OR TRIM(rubrica) LIKE '17.7 - %' "
        "OR TRIM(rubrica) LIKE '17.8 - %' "
        "OR TRIM(natureza_raw) LIKE '17.2 - %' "
        "OR TRIM(natureza_raw) LIKE '17.3 - %' "
        "OR TRIM(natureza_raw) LIKE '17.4 - %' "
        "OR TRIM(natureza_raw) LIKE '17.5 - %' "
        "OR TRIM(natureza_raw) LIKE '17.7 - %' "
        "OR TRIM(natureza_raw) LIKE '17.8 - %' "
        "THEN debito ELSE 0 END"
    )
    SAIDAS_LIQUIDAS_EXPR = f"(debito - ({IMPOSTO_DEBITO_EXPR}))"
    SALDO_DRE_EXPR = f"(credito - {SAIDAS_LIQUIDAS_EXPR})"
    ESCOPO_PROJETO_COMPLETO = "projeto_completo"
    COMPONENTES_DRE_COM_CODIGO_OBRIGATORIO = {
        "deducoes",
        "custos_variaveis",
        "gastos_fixos",
        "depreciacao",
        "investimentos_gerencial",
        "investimentos",
        "folha_pagamento",
        "total_imposto_retido",
        "estoques",
        "fornecedores",
        "passivos_operacionais",
    }
    COMPONENTES_DRE_CONTA_PAI_ESTRITA = {
        "deducoes": _alias_set(["(-)Deduções sobre vendas"]),
        "custos_variaveis": _alias_set(["(-)Custos Variavéis", "(-)Custos Variáveis"]),
        "gastos_fixos": _alias_set(["(-)Gastos Fixos"]),
        "depreciacao": _alias_set(["(-)Depreciação Imobilizado"]),
        "investimentos_gerencial": _alias_set(["(-)Investimentos"]),
        "folha_pagamento": _alias_set(["(-)Gastos Fixos"]),
    }
    COMPONENTES_DRE_CODIGOS_EXPLICITOS = {
        "folha_pagamento": frozenset(("12.100", "12.101")),
    }
    RECEITA_BRUTA_RECOMPOSICAO_ALIASES = _alias_set(
        [
            "(-)Deduções sobre vendas",
            "Impostos sobre vendas",
            "Descontos sobre vendas",
            "Simples Nacional",
            "IR",
            "IR Retido",
            "ISS",
            "ISS Retido",
            "INSS",
            "INSS Retido",
            "Tarifa de Antecipação",
            "PIS",
            "COFINS",
            "CSLL",
            "ICMS",
        ]
    )
    COMPONENTES_DRE = {
        "receita_bruta": _alias_set(
            ["(=)Receita Bruta", "Faturamento", "Recebimento de Clientes", "Receita Bruta"]
        ),
        "deducoes": _alias_set(
            [
                "(-)Deduções sobre vendas",
                "Impostos sobre vendas",
                "Descontos sobre vendas",
                "Simples Nacional",
                "IR",
                "IR Retido",
                "ISS",
                "ISS Retido",
                "INSS",
                "INSS Retido",
                "Tarifa de Antecipação",
                "PIS",
                "COFINS",
                "CSLL",
            ]
        ),
        "receita_liquida": _alias_set(["(=)Receita Líquida", "Receita Líquida"]),
        "custos_variaveis": _alias_set(
            [
                "(-)Custos Variavéis",
                "(-)Custos Variáveis",
                "Fornecedores",
                "FRETE MATERIAIS E EQUIPAMENTOS",
                "FRETE ABRASIVOS",
                "FRETE TINTAS E SOLVENTE",
                "FRETE TRANSPORTADORA",
                "TINTAS E SOLVENTES",
                "ABRASIVOS",
                "OUTROS MATERIAIS DE APLICAÇÃO",
                "Gastos Comerciais",
                "COMISSÃO",
            ]
        ),
        "margem_contribuicao": _alias_set(
            ["(=)MARGEM DE CONTRIBUIÇÃO", "MARGEM DE CONTRIBUIÇÃO"]
        ),
        "gastos_fixos": _alias_set(
            [
                "(-)Gastos Fixos",
                "Despesas com Pessoal",
                "SALARIO",
                "13 SALARIO",
                "13 PREVISAO",
                "FERIAS",
                "PREVISAO FÉRIAS",
                "MULTA RESCISORIAS FGTS",
                "REFEIÇÕES COLABORADORES",
                "REFEIÇÕES FUNCIONARIOS",
                "ALIMENTAÇÃO ADM",
                "ACERTO RESCISORIOS",
                "TRANSPORTE DE COLABORADORES",
                "TRANSPORTE FUNCIONARIOS",
                "SINDICATO",
                "PASSAGENS COLABORADORES",
                "PASSAGENS FUNCIONARIOS",
                "PLANO DE SAUDE",
                "ASSISTENCIA MEDICA ADM",
                "SEGUROS DE VIDA",
                "VALE ALIMENTAÇÃO",
                "FGTS FUNCIONARIOS",
                "DCTFWEB",
                "PLR",
                "BONIFICAÇOES",
                "EPIS",
                "UNIFORMES",
                "EXAMES MÉDICOS",
                "EXAMES",
                "CURSOS / TREINAMENTOS",
                "PROGRAMA DE SEGURANÇA DO TRABALHO",
                "Serviços de Terceiros",
                "SERVIÇO MECANICO",
                "SERVIÇO ELETRICO",
                "SERVIÇO SERRALHEIRO",
                "SERVIÇO LAUDO TECNICO",
                "Despesas Administrativas",
                "SISTEMA / SERVIDOR",
                "SERVIÇOS DE CONSULTORIA",
                "HONORÁRIOS ADVOCATÍCIOS",
                "HONORÁRIOS CONTÁBEIS",
                "CONTABILIDADE",
                "DESPESA CARTAO DE CREDITO",
                "MATERIAIS DE INFORMATICA / COMPUTADOR",
                "SERVIÇO SERASA",
                "SERVIÇO GOOGLE ADS",
                "SERVIÇO MOTOBOY / UBER",
                "SERVIÇO GRAFICA / IMPRESSAO",
                "SERVIÇO ART",
                "SERVIÇO ABRACO ANUIDADE",
                "SERVIÇO CARTORIO",
                "SERVIÇO TELEMETRIA / FADIGA",
                "MATERIAS DE ESCRITORIO / PAPELARIA",
                "MATERIAS DE ESCRITORIO",
                "SERVIÇO CREA ANUIDADE",
                "SERVIÇO PLOTAGEM / ADESIVO / PLACAS / CRACHAS",
                "SERVIÇO E-MAILS",
                "SERVIÇO ALVARAS E LICENCIAMENTO",
                "SERVIÇO BRINDES E ACESSORIOS",
                "Despesas com Veículos e Equipamentos",
                "IPVA",
                "MULTAS DE TRANSITO",
                "MANUTENÇÃO/PEÇAS  VEICULOS",
                "MANUTENÇÃO/PEÇAS VEICULOS",
                "COMBUSTIVEL VEICULOS",
                "SEGUROS DE VEICULOS",
                "PEDAGIOS",
                "DESPACHANTE",
                "CUSTOS E GUIAS DE TRANSFERENCIA VEICULOS",
                "GUINCHO",
                "SERVIÇO LAVA JATO",
                "Despesas com Locações",
                "LOCAÇÃO MUNCK / GUINDASTES",
                "LOCAÇÃO COMPRESSORES",
                "LOCAÇÃO GERADOR",
                "LOCAÇÃO EQUIPAMENTOS",
                "LOCAÇÃO EMPILHADEIRA",
                "LOCAÇÃO CONTAINER",
                "LOCAÇÃO ANDAIMES",
                "LOCAÇAO VEICULOS",
                "LOCAÇAO CAÇAMBA P/ RESIDUOS",
                "LOCAÇAO RADIOS DE COMUNICAÇÃO",
                "LOCAÇAO BANHEIRO QUIMICO",
                "LOCAÇAO PLATAFORMA ELEVATORIA",
                "LOCAÇÃO AMBULANCIA / PRIMEIROS SOCORROS",
                "Despesas com Locações",
                "Despesas com Máquinas e Equipamentos",
                "MAQUINAS / FERRAMENTAS COMBUSTÃO",
                "MAQUINAS / FERRAMENTAS ELETRICAS",
                "MAQUINAS / FERRAMENTAS PNEUMATICAS",
                "MAQUINAS / EQUIPAMENTOS",
                "MATERIAIS / FERRAMENTAS  MANUAIS",
                "MANGUEIRAS / CONEXÕES",
                "MATERIAL ELETRICO",
                "MATERIAIS DE CONSUMO EM OBRAS",
                "MATERIAIS PINCEL / ROLOS / LIXAS",
                "MATERIAIS DE LIMPEZA / HIGIENE",
                "EQUIPAMENTOS DE AFERIÇÃO",
                "COMBUSTIVEL EQUIPAMENTOS",
                "MANUTENÇÃO/PEÇAS MAQUINAS E EQUIPAMENTOS",
                "Despesas com Viagens e Hospedagens",
                "HOTEL COLABORADORES",
                "HOTEL FUNCIONARIOS",
                "LAVANDERIA",
                "Despesas Financeiras",
                "TARIFAS BANCARIAS",
                "Despesas com Infraestrutura",
                "ALUGUEL ALOJAMENTO",
                "AGUA ALOJAMENTO",
                "ENERGIA ALOJAMENTO",
                "MOVEIS OBRAS",
                "SERVIÇO LIMPEZA ALOJAMENTO/ESCRITORIO",
                "ALUGUEIS ADM",
                "AGUA ADM",
                "ENERGIA ADM",
                "IPTU",
                "INTERNET / TELEFONE ADM",
                "INTERNET OBRAS",
                "Serviços de Terceiros",
            ]
        ),
        "resultado_operacional": _alias_set(
            ["(=)RESULTADO OPERACIONAL", "RESULTADO OPERACIONAL"]
        ),
        "depreciacao": _alias_set(["(-)Depreciação Imobilizado", "Depreciação Imobilizado"]),
        "resultado_liquido": _alias_set(["(=)RESULTADO LÍQUIDO", "RESULTADO LÍQUIDO"]),
        "investimentos_gerencial": _alias_set(["(-)Investimentos"]),
        "investimentos": _alias_set(
            [
                "Compra de veiculos",
                "Compra de veículos",
                "Aquisição de Máquinas e Equipamentos",
                "Aquisição de Maquinas e Equipamentos",
                "Equipamentos de Aferição",
                "Materiais de Consumo e Obra",
                "MATERIAIS DE CONSUMO EM OBRAS",
                "Material Elétrico",
                "MANGUEIRAS / CONEXÕES",
                "Ferramentas Manuais",
                "MATERIAIS / FERRAMENTAS  MANUAIS",
                "Despesas Administrativas",
                "Programas de Segurança do Trabalho",
                "PROGRAMA DE SEGURANÇA DO TRABALHO",
                "Curso e Treinamento",
                "CURSOS / TREINAMENTOS",
                "Exames Médicos",
                "EXAMES MÉDICOS",
                "EXAMES",
                "Uniformes",
                "UNIFORMES",
                "EPI´s",
                "EPIS",
                "Fornecedores",
                "Fornecedores Total",
                "Empréstimos de Terceiros",
                "Empréstimos Bancários",
            ]
        ),
        "resultado_gerencial": _alias_set(["(=)RESULTADO GERENCIAL", "RESULTADO GERENCIAL"]),
        "folha_pagamento": _alias_set(
            [
                "SALÁRIO",
                "SALARIO",
                "130 SALÁRIO",
                "130 SALARIO",
                "13 SALÁRIO",
                "13 SALARIO",
                "PREVISÃO 13",
                "PREVISAO 13",
                "PREVISÃO 13°",
                "PREVISAO 13°",
                "12.3 - PREVISÃO 13°",
                "FÉRIAS",
                "FERIAS",
                "PREVISÃO FÉRIAS",
                "PREVISAO FERIAS",
                "12.70 - PREVISÃO FÉRIAS",
                "MULTA RESCISÓRIAS FGTS",
                "MULTA RESCISORIAS FGTS",
                "ACERTO RESCISÓRIOS",
                "ACERTO RESCISORIOS",
                "FGTS FUNCIONÁRIOS",
                "FGTS FUNCIONARIOS",
            ]
        ),
        "total_imposto_retido": _alias_set(
            [
                "TOTAL DE IMPOSTO RETIDO",
                "TOTAL DE IMPOSTOS RETIDOS",
                "TOTAL IMPOSTO RETIDO",
                "TOTAL IMPOSTOS RETIDOS",
            ]
        ),
        "contas_receber": _alias_set(["Contas a Receber"]),
        "contas_pagar": _alias_set(["Contas a Pagar"]),
        "estoques": _alias_set(["Estoques"]),
        "fornecedores": _alias_set(["Fornecedores"]),
        "passivos_operacionais": _alias_set(["Passivos Operacionais"]),
    }

    def obter_painel(
        self,
        ano: int | None = None,
        meses: list[int] | None = None,
        centro_custo: list[str] | None = None,
        natureza: list[str] | None = None,
        escopo_periodo: str | None = None,
    ) -> dict[str, Any]:
        ano_ref = ano if ano is not None else self._ano_mais_recente("dre_lancamentos")
        meses_ref = sorted(set(meses or []))
        centro_custo_ref = _normalizar_lista_texto(centro_custo)
        natureza_ref = _normalizar_lista_texto(natureza)
        escopo_ref = (escopo_periodo or "").strip()
        projeto_completo = (
            escopo_ref == self.ESCOPO_PROJETO_COMPLETO and len(centro_custo_ref) > 0
        )

        _validar_ano(ano_ref)
        _validar_meses(meses_ref)

        where: list[str] = []
        params: list[Any] = []
        if not projeto_completo:
            where.append("competencia_ano = ?")
            params.append(ano_ref)
            _append_in(where, params, "competencia_mes", meses_ref)
        _append_in(where, params, self.CENTRO_CUSTO_EXPR, centro_custo_ref)
        _append_in(where, params, self.NATUREZA_EXPR, natureza_ref)
        where_sql = " AND ".join(where) if where else "1 = 1"

        period_where = ["competencia_ano = ?"]
        period_params: list[Any] = [ano_ref]
        _append_in(period_where, period_params, "competencia_mes", meses_ref)
        period_where_sql = " AND ".join(period_where)

        with self.db.get_connection() as conn:
            meses_disponiveis = [
                int(row["competencia_mes"])
                for row in conn.execute(
                    """
                    SELECT DISTINCT competencia_mes
                    FROM dre_lancamentos
                    WHERE competencia_ano = ?
                    ORDER BY competencia_mes
                    """,
                    (ano_ref,),
                ).fetchall()
            ]

            kpis = conn.execute(
                f"""
                SELECT
                    COUNT(*) AS total_lancamentos,
                    SUM(credito) AS total_credito,
                    SUM(debito) AS total_debito,
                    SUM({self.IMPOSTO_DEBITO_EXPR}) AS total_impostos,
                    SUM({self.SAIDAS_LIQUIDAS_EXPR}) AS total_saidas_liquidas,
                    SUM({self.SALDO_DRE_EXPR}) AS saldo_liquido,
                    COUNT(DISTINCT {self.CENTRO_CUSTO_EXPR}) AS total_obras,
                    COUNT(DISTINCT {self.NATUREZA_EXPR}) AS total_naturezas
                FROM dre_lancamentos
                WHERE {where_sql}
                """,
                tuple(params),
            ).fetchone()
            series = conn.execute(
                f"""
                SELECT
                    competencia_ano AS ano,
                    competencia_mes AS mes,
                    SUM(credito) AS credito,
                    SUM(debito) AS debito,
                    SUM(credito) AS receita_liquida,
                    SUM({self.IMPOSTO_DEBITO_EXPR}) AS impostos,
                    SUM({self.SAIDAS_LIQUIDAS_EXPR}) AS saidas_liquidas,
                    SUM({self.SALDO_DRE_EXPR}) AS saldo,
                    COUNT(*) AS total
                FROM dre_lancamentos
                WHERE {where_sql}
                GROUP BY competencia_ano, competencia_mes
                ORDER BY competencia_ano, competencia_mes
                """,
                tuple(params),
            ).fetchall()
            ranking_obras = self._ranking(conn, self.CENTRO_CUSTO_EXPR, where_sql, params, limit=14)
            ranking_naturezas = self._ranking(conn, self.NATUREZA_EXPR, where_sql, params, limit=14)
            ultimos = conn.execute(
                f"""
                SELECT
                    id,
                    data_lancamento AS data,
                    historico,
                    credito,
                    debito,
                    {self.SALDO_DRE_EXPR} AS saldo,
                    {self.CENTRO_CUSTO_EXPR} AS centro_custo,
                    {self.NATUREZA_EXPR} AS natureza
                FROM dre_lancamentos
                WHERE {where_sql}
                ORDER BY data_lancamento DESC, id DESC
                LIMIT 12
                """,
                tuple(params),
            ).fetchall()
            componentes = conn.execute(
                f"""
                SELECT
                    competencia_ano AS ano,
                    competencia_mes AS mes,
                    valor_bruto,
                    credito,
                    debito,
                    natureza_raw,
                    natureza_norm,
                    rubrica,
                    conta_pai
                FROM dre_lancamentos
                WHERE {where_sql}
                """,
                tuple(params),
            ).fetchall()

            filtros = {
                "anos": [
                    int(row["competencia_ano"])
                    for row in conn.execute(
                        """
                        SELECT DISTINCT competencia_ano
                        FROM dre_lancamentos
                        ORDER BY competencia_ano DESC
                        """
                    ).fetchall()
                ],
                "meses": meses_disponiveis,
                "centro_custo": self._filter_options(
                    conn, self.CENTRO_CUSTO_EXPR, period_where_sql, period_params
                ),
                "natureza": self._filter_options(
                    conn, self.NATUREZA_EXPR, period_where_sql, period_params
                ),
            }

        total_lancamentos = int(kpis["total_lancamentos"] or 0)
        total_credito = _float(kpis["total_credito"])
        total_debito = _float(kpis["total_debito"])
        total_impostos = _float(kpis["total_impostos"])
        total_saidas_liquidas = _float(kpis["total_saidas_liquidas"])
        saldo_liquido = _float(kpis["saldo_liquido"])
        resultado_liquido = saldo_liquido
        competencias_series = [
            (int(row["ano"]), int(row["mes"])) for row in series if row["ano"] is not None
        ]
        meses_analise = len(competencias_series or meses_ref or meses_disponiveis)
        media_saida_mensal = total_saidas_liquidas / meses_analise if meses_analise else 0
        saldo_medio_mensal = saldo_liquido / meses_analise if meses_analise else 0
        folego_caixa_meses = (
            max(saldo_liquido, 0) / media_saida_mensal if media_saida_mensal else 0
        )
        periodo = (
            self._periodo_intervalo_payload(ano_ref, competencias_series)
            if projeto_completo
            else self._periodo_payload(ano_ref, meses_ref, meses_disponiveis)
        )
        componentes_gerados = self._componentes_compativeis_dre_gerado(componentes)
        competencias_indicadores = self._competencias_indicadores_painel(
            ano_ref,
            meses_ref,
            meses_disponiveis,
            competencias_series,
        )
        indicadores_impostos = self.indicadores_manuais.somar_competencias(
            competencias_indicadores
        )
        indicadores_ncg = self.indicadores_manuais.somar_competencias(
            self._competencias_subsequentes(competencias_indicadores)
        )
        componentes_dre = self._componentes_dre(
            componentes_gerados,
            saldo_liquido,
            receita_liquida_base=total_credito,
        )
        componentes_series = self._componentes_por_competencia(componentes_gerados)
        componentes_dre = self._aplicar_indicadores_manuais(
            componentes_dre,
            indicadores_impostos,
            indicadores_ncg,
        )
        receita_liquida_dre = _float(componentes_dre["receita_liquida"])
        receita_bruta_dre = _float(componentes_dre["receita_bruta"])
        deducoes_vendas = _float(componentes_dre["deducoes"])
        irpj_csll = max(total_impostos - deducoes_vendas, 0)
        series_mensais = self._series_rows(series, "lancamentos")
        series_mensais = self._aplicar_componentes_series(
            series_mensais,
            componentes_series,
            ano_ref,
        )
        return {
            "success": True,
            "periodo": periodo,
            "filtros_disponiveis": filtros,
            "filtros_aplicados": {
                "ano": ano_ref,
                "meses": meses_ref,
                "centro_custo": centro_custo_ref,
                "natureza": natureza_ref,
                "escopo_periodo": (
                    self.ESCOPO_PROJETO_COMPLETO if projeto_completo else "ano"
                ),
            },
            "kpis": {
                "total_lancamentos": total_lancamentos,
                "total_credito": total_credito,
                "receita_bruta": receita_bruta_dre,
                "receita_liquida": receita_liquida_dre,
                "deducoes_vendas": deducoes_vendas,
                "irpj_csll": irpj_csll,
                "total_debito": total_debito,
                "total_saidas_liquidas": total_saidas_liquidas,
                "total_impostos": total_impostos,
                "resultado_liquido": resultado_liquido,
                "saldo_liquido": saldo_liquido,
                "total_obras": int(kpis["total_obras"] or 0),
                "total_naturezas": int(kpis["total_naturezas"] or 0),
                "ticket_medio": saldo_liquido / total_lancamentos if total_lancamentos else 0,
                "media_saida_mensal": media_saida_mensal,
                "saldo_medio_mensal": saldo_medio_mensal,
                "folego_caixa_meses": folego_caixa_meses,
                "margem_resultado_percentual": _percent(saldo_liquido, receita_liquida_dre),
                "pressao_saida_percentual": _percent(total_saidas_liquidas, receita_liquida_dre),
                "meses_analise": meses_analise,
            },
            "indicadores_manuais": self._indicadores_manuais_payload(
                indicadores_impostos,
                indicadores_ncg,
            ),
            "saldos_projeto": (
                self._saldos_projeto(
                    competencias_series,
                    total_credito,
                    total_debito,
                    total_saidas_liquidas,
                    total_impostos,
                    saldo_liquido,
                )
                if projeto_completo
                else None
            ),
            "indicadores_viabilidade": self._indicadores_viabilidade(
                componentes_dre,
            ),
            "objetivos_estrategicos": self._objetivos_estrategicos(componentes_dre),
            "series_mensais": series_mensais,
            "ranking_obras": self._ranking_rows(ranking_obras, "lancamentos"),
            "ranking_naturezas": self._ranking_rows(ranking_naturezas, "lancamentos"),
            "ultimos_lancamentos": [
                {
                    "id": row["id"],
                    "data": row["data"],
                    "historico": row["historico"],
                    "credito": _float(row["credito"]),
                    "debito": _float(row["debito"]),
                    "saldo": _float(row["saldo"]),
                    "centro_custo": row["centro_custo"],
                    "natureza": row["natureza"],
                }
                for row in ultimos
            ],
        }

    def _plano_contas_gerado(self) -> dict[str, dict]:
        if self._plano_contas_gerado_cache is not None:
            return self._plano_contas_gerado_cache

        try:
            with TemplateWriter(settings.template_dre_path) as writer:
                self._plano_contas_gerado_cache = DREGeracaoCompletaService._ler_plano_contas(
                    writer,
                    aplicar_overrides_dre_gerado=True,
                )
        except Exception as exc:
            logger.warning("Não foi possível ler PLANO_CONTAS do DRE gerado: %s", exc)
            self._plano_contas_gerado_cache = {}
        return self._plano_contas_gerado_cache

    def _componentes_compativeis_dre_gerado(self, rows: list[Any]) -> list[dict[str, Any]]:
        plano = self._plano_contas_gerado()
        if not plano:
            return [self._row_to_dict(row) for row in rows]

        linhas: list[dict[str, Any]] = []
        for row in rows:
            base = self._row_to_dict(row)
            codigo_conta = self._codigo_conta_row(base)
            rubrica, conta_filho, conta_pai, _cod = DREGeracaoCompletaService._resolver_conta_pai(
                base.get("natureza_raw"),
                base.get("rubrica"),
                plano,
                _float(base.get("credito")) - _float(base.get("debito")),
            )
            linhas.append(
                {
                    **base,
                    "rubrica": rubrica or base.get("rubrica"),
                    "conta_filho": conta_filho or base.get("conta_filho"),
                    "conta_pai": conta_pai or base.get("conta_pai"),
                    "codigo_conta": codigo_conta,
                    "_plano_contas_resolvido": bool(rubrica or conta_filho or conta_pai),
                    "_plano_contas_cod": _cod,
                }
            )
        return linhas

    def _componentes_por_competencia(
        self,
        rows: list[dict[str, Any]],
    ) -> dict[tuple[int, int], dict[str, Any]]:
        grupos: dict[tuple[int, int], list[dict[str, Any]]] = {}
        creditos: dict[tuple[int, int], float] = {}
        saldos: dict[tuple[int, int], float] = {}

        for row in rows:
            ano = self._row_get(row, "ano")
            mes = self._row_get(row, "mes")
            if ano is None or mes is None:
                continue
            key = (int(ano), int(mes))
            grupos.setdefault(key, []).append(row)
            credito = _float(self._row_get(row, "credito"))
            debito = _float(self._row_get(row, "debito"))
            creditos[key] = creditos.get(key, 0.0) + credito
            saldos[key] = saldos.get(key, 0.0) + (credito - debito)

        return {
            key: self._componentes_dre(
                grupo,
                saldos.get(key, 0.0),
                receita_liquida_base=creditos.get(key, 0.0),
            )
            for key, grupo in grupos.items()
        }

    @staticmethod
    def _aplicar_componentes_series(
        series: list[dict[str, Any]],
        componentes_por_competencia: dict[tuple[int, int], dict[str, Any]],
        ano_ref: int,
    ) -> list[dict[str, Any]]:
        atualizadas: list[dict[str, Any]] = []
        for item in series:
            key = (int(item.get("ano") or ano_ref), int(item["mes"]))
            componentes = componentes_por_competencia.get(key)
            if not componentes:
                atualizadas.append(item)
                continue
            atualizadas.append(
                {
                    **item,
                    "receita_bruta": _float(componentes.get("receita_bruta")),
                    "receita_liquida": _float(componentes.get("receita_liquida")),
                    "deducoes_vendas": _float(componentes.get("deducoes")),
                }
            )
        return atualizadas

    @staticmethod
    def _row_to_dict(row: Any) -> dict[str, Any]:
        if isinstance(row, dict):
            return dict(row)
        keys = row.keys() if hasattr(row, "keys") else []
        return {key: row[key] for key in keys}

    @staticmethod
    def _codigo_conta_row(row: Any) -> str | None:
        for key in ("codigo_conta", "natureza_raw", "rubrica", "conta_filho", "conta_pai"):
            codigo = _extrair_codigo_gerencial(PainelDREService._row_get(row, key))
            if codigo:
                return codigo
        return None

    @staticmethod
    def _competencias_indicadores_painel(
        ano: int,
        meses: list[int],
        meses_disponiveis: list[int],
        competencias_series: list[tuple[int, int]],
    ) -> list[tuple[int, int]]:
        if competencias_series:
            return sorted(set(competencias_series))
        meses_ref = meses or meses_disponiveis
        return [(ano, mes) for mes in meses_ref]

    @staticmethod
    def _competencias_subsequentes(
        competencias: list[tuple[int, int]],
    ) -> list[tuple[int, int]]:
        subsequentes: set[tuple[int, int]] = set()
        for ano, mes in competencias:
            if mes == 12:
                subsequentes.add((ano + 1, 1))
            else:
                subsequentes.add((ano, mes + 1))
        return sorted(subsequentes)

    @staticmethod
    def _indicadores_periodo_payload(indicadores: dict[str, Any]) -> dict[str, Any]:
        return {
            "existe": bool(indicadores.get("existe")),
            "ano": indicadores.get("ano"),
            "meses": indicadores.get("meses", []),
            "competencias": indicadores.get("competencias", []),
            "total_registros": int(indicadores.get("total_registros") or 0),
        }

    @staticmethod
    def _indicadores_manuais_payload(
        indicadores_impostos: dict[str, Any],
        indicadores_ncg: dict[str, Any],
    ) -> dict[str, Any]:
        return {
            "existe": bool(
                indicadores_impostos.get("existe") or indicadores_ncg.get("existe")
            ),
            "ano": indicadores_impostos.get("ano"),
            "meses": indicadores_impostos.get("meses", []),
            "total_registros": (
                int(indicadores_impostos.get("total_registros") or 0)
                + int(indicadores_ncg.get("total_registros") or 0)
            ),
            "contas_pagar": float(indicadores_ncg.get("contas_pagar") or 0),
            "contas_receber": float(indicadores_ncg.get("contas_receber") or 0),
            "total_impostos_retidos_acima_meta": float(
                indicadores_impostos.get("total_impostos_retidos_acima_meta") or 0
            ),
            "total_impostos_retidos": float(
                indicadores_impostos.get("total_impostos_retidos") or 0
            ),
            "periodo_dre": PainelDREService._indicadores_periodo_payload(
                indicadores_impostos
            ),
            "periodo_ncg": PainelDREService._indicadores_periodo_payload(indicadores_ncg),
        }

    @staticmethod
    def _aplicar_indicadores_manuais(
        componentes: dict[str, Any],
        indicadores_impostos: dict[str, Any],
        indicadores_ncg: dict[str, Any],
    ) -> dict[str, Any]:
        atualizados = dict(componentes)

        if indicadores_impostos.get("existe"):
            atualizados = {
                **atualizados,
                "total_imposto_retido": float(
                    indicadores_impostos.get("total_impostos_retidos") or 0
                ),
                "total_impostos_retidos_acima_meta": float(
                    indicadores_impostos.get("total_impostos_retidos_acima_meta") or 0
                ),
                "_tem_total_imposto_retido": True,
                "_tem_total_impostos_retidos_acima_meta": True,
            }

        if indicadores_ncg.get("existe"):
            atualizados = {
                **atualizados,
                "contas_pagar": float(indicadores_ncg.get("contas_pagar") or 0),
                "contas_receber": float(indicadores_ncg.get("contas_receber") or 0),
                "_tem_contas_pagar": True,
                "_tem_contas_receber": True,
            }

        return atualizados

    def _saldos_projeto(
        self,
        competencias: list[tuple[int, int]],
        credito: float,
        debito: float,
        saidas_liquidas: float,
        impostos: float,
        saldo: float,
    ) -> dict[str, Any]:
        ordenadas = sorted(set(competencias))
        primeira = ordenadas[0] if ordenadas else None
        ultima = ordenadas[-1] if ordenadas else None
        return {
            "credito": credito,
            "debito": debito,
            "saidas_liquidas": saidas_liquidas,
            "impostos": impostos,
            "saldo": saldo,
            "primeira_competencia": (
                _competencia_label(primeira[0], primeira[1]) if primeira else None
            ),
            "ultima_competencia": _competencia_label(ultima[0], ultima[1]) if ultima else None,
            "meses_com_dados": [
                {
                    "ano": ano,
                    "mes": mes,
                    "periodo": f"{ano}-{mes:02d}",
                    "label": _competencia_curta(ano, mes),
                }
                for ano, mes in ordenadas
            ],
        }

    def _indicadores_viabilidade(
        self,
        componentes: dict[str, Any],
    ) -> list[dict[str, Any]]:
        receita_liquida = componentes["receita_liquida"]
        custos_variaveis = componentes["custos_variaveis"]
        gastos_fixos = componentes["gastos_fixos"]
        ebit = componentes["resultado_operacional"]
        fcl = componentes["resultado_gerencial"]
        investimento_total = componentes["investimentos"]
        lucro_liquido = componentes["resultado_liquido"]

        mcl = (
            componentes["margem_contribuicao"]
            if componentes["_tem_margem_contribuicao"]
            else receita_liquida - custos_variaveis
        )
        margem_mcl = _percent(mcl, receita_liquida)
        margem_mcl_ratio = mcl / receita_liquida if receita_liquida else 0
        ebitda = ebit

        indicadores = [
            _indicador_calculado(
                "mcl",
                "Margem de Contribuição Líquida",
                mcl,
                margem_mcl,
                {
                    "receita_liquida": receita_liquida,
                    "custos_despesas_variaveis": custos_variaveis,
                },
            )
        ]

        if receita_liquida > 0 and mcl > 0 and margem_mcl_ratio > 0 and gastos_fixos > 0:
            indicadores.append(
                _indicador_calculado(
                    "pel",
                    "Ponto de Equilíbrio Líquido",
                    gastos_fixos / margem_mcl_ratio,
                    None,
                    {"custos_fixos": gastos_fixos, "margem_mcl_decimal": margem_mcl_ratio},
                )
            )
        else:
            faltantes_pe = []
            if receita_liquida <= 0:
                faltantes_pe.append("Receita Líquida positiva")
            if mcl <= 0:
                faltantes_pe.append("Margem de Contribuição Líquida positiva")
            if margem_mcl_ratio <= 0:
                faltantes_pe.append("Margem de Contribuição Líquida positiva")
            if not gastos_fixos:
                faltantes_pe.append("Custos Fixos")
            faltantes_pe = list(dict.fromkeys(faltantes_pe))
            indicadores.append(
                _indicador_indisponivel(
                    "pel",
                    "Ponto de Equilíbrio Líquido",
                    faltantes_pe,
                    {"custos_fixos": gastos_fixos, "margem_mcl_decimal": margem_mcl_ratio},
                )
            )

        indicadores.append(
            _indicador_calculado(
                "ebitda",
                "EBITDA",
                ebitda,
                _percent(ebitda, receita_liquida),
                {
                    "resultado_operacional": ebit,
                    "receita_liquida": receita_liquida,
                },
            )
        )
        indicadores.append(
            _indicador_calculado(
                "fcl",
                "Fluxo de Caixa Livre",
                fcl,
                _percent(fcl, receita_liquida),
                {"resultado_gerencial": fcl, "receita_liquida": receita_liquida},
            )
        )

        tem_resultado_liquido = bool(componentes.get("_tem_resultado_liquido"))
        tem_investimento_total = bool(componentes.get("_tem_investimentos")) and (
            investimento_total > 0
        )

        if tem_resultado_liquido and tem_investimento_total:
            indicadores.append(
                _indicador_calculado(
                    "roi",
                    "ROI",
                    _percent(lucro_liquido, investimento_total),
                    None,
                    {
                        "lucro_liquido": lucro_liquido,
                        "investimento_total": investimento_total,
                    },
                )
            )
        else:
            faltantes_roi = []
            if not tem_resultado_liquido:
                faltantes_roi.append("Resultado Líquido")
            if not tem_investimento_total:
                faltantes_roi.append("Investimento Total")
            indicadores.append(
                _indicador_indisponivel(
                    "roi",
                    "ROI",
                    faltantes_roi,
                    {"lucro_liquido": lucro_liquido, "investimento_total": investimento_total},
                )
            )

        indicadores.append(self._indicador_ncg(componentes, receita_liquida))
        return indicadores

    def _componentes_dre(
        self,
        rows: list[Any],
        saldo_liquido: float,
        receita_liquida_base: float | None = None,
    ) -> dict[str, Any]:
        receita_bruta, tem_receita_bruta = self._somar_receita_bruta(rows)
        deducoes, _ = self._somar_componente(rows, "deducoes", absoluto=True)
        receita_liquida_explicita, tem_receita_liquida = self._somar_componente(
            rows, "receita_liquida"
        )
        if tem_receita_liquida:
            receita_liquida = receita_liquida_explicita
        elif tem_receita_bruta:
            receita_liquida = receita_bruta - deducoes
        elif receita_liquida_base is not None:
            receita_liquida = _float(receita_liquida_base)
        else:
            receita_liquida = _float(saldo_liquido)
        custos_variaveis, _ = self._somar_componente(rows, "custos_variaveis", absoluto=True)
        margem_contribuicao, tem_margem_contribuicao = self._somar_componente(
            rows, "margem_contribuicao"
        )
        gastos_fixos, _ = self._somar_componente(rows, "gastos_fixos", absoluto=True)
        depreciacao, _ = self._somar_componente(rows, "depreciacao", absoluto=True)
        resultado_operacional_explicito, tem_resultado_operacional = self._somar_componente(
            rows, "resultado_operacional"
        )
        resultado_operacional_calculado = (
            margem_contribuicao if tem_margem_contribuicao else receita_liquida - custos_variaveis
        ) - gastos_fixos
        resultado_operacional = (
            resultado_operacional_explicito
            if tem_resultado_operacional
            else resultado_operacional_calculado
        )
        resultado_liquido_explicito, tem_resultado_liquido = self._somar_componente(
            rows, "resultado_liquido"
        )
        resultado_liquido = (
            resultado_liquido_explicito if tem_resultado_liquido else _float(saldo_liquido)
        )
        tem_resultado_liquido_calculado = tem_resultado_liquido or bool(rows)
        investimentos, tem_investimentos = self._somar_componente(
            rows, "investimentos", absoluto=True
        )
        investimentos_gerencial, _ = self._somar_componente(rows, "investimentos_gerencial")
        resultado_gerencial_explicito, tem_resultado_gerencial = self._somar_componente(
            rows, "resultado_gerencial"
        )
        resultado_gerencial = (
            resultado_gerencial_explicito
            if tem_resultado_gerencial
            else resultado_liquido + investimentos_gerencial
        )
        folha_pagamento, tem_folha_pagamento = self._somar_componente(
            rows, "folha_pagamento", absoluto=True
        )
        total_imposto_retido, tem_total_imposto_retido = self._somar_componente(
            rows, "total_imposto_retido", absoluto=True
        )
        total_impostos_retidos_acima_meta = 0.0
        tem_total_impostos_retidos_acima_meta = False
        contas_receber, tem_contas_receber = self._somar_componente(
            rows, "contas_receber", absoluto=True
        )
        contas_pagar, tem_contas_pagar = self._somar_componente(
            rows, "contas_pagar", absoluto=True
        )
        estoques, tem_estoques = self._somar_componente(rows, "estoques", absoluto=True)
        fornecedores, tem_fornecedores = self._somar_componente(
            rows, "fornecedores", absoluto=True
        )
        passivos_operacionais, tem_passivos_operacionais = self._somar_componente(
            rows, "passivos_operacionais", absoluto=True
        )

        return {
            "receita_bruta": receita_bruta,
            "deducoes": deducoes,
            "receita_liquida": receita_liquida,
            "custos_variaveis": custos_variaveis,
            "margem_contribuicao": margem_contribuicao,
            "gastos_fixos": gastos_fixos,
            "depreciacao": depreciacao,
            "resultado_operacional": resultado_operacional,
            "resultado_gerencial": resultado_gerencial,
            "resultado_liquido": resultado_liquido,
            "investimentos": investimentos,
            "folha_pagamento": folha_pagamento,
            "total_imposto_retido": total_imposto_retido,
            "total_impostos_retidos_acima_meta": total_impostos_retidos_acima_meta,
            "contas_receber": contas_receber,
            "contas_pagar": contas_pagar,
            "estoques": estoques,
            "fornecedores": fornecedores,
            "passivos_operacionais": passivos_operacionais,
            "_tem_receita_liquida": tem_receita_liquida or tem_receita_bruta,
            "_tem_margem_contribuicao": tem_margem_contribuicao,
            "_tem_resultado_liquido": tem_resultado_liquido_calculado,
            "_tem_investimentos": tem_investimentos,
            "_tem_folha_pagamento": tem_folha_pagamento,
            "_tem_total_imposto_retido": tem_total_imposto_retido,
            "_tem_total_impostos_retidos_acima_meta": tem_total_impostos_retidos_acima_meta,
            "_tem_contas_receber": tem_contas_receber,
            "_tem_contas_pagar": tem_contas_pagar,
            "_tem_estoques": tem_estoques,
            "_tem_fornecedores": tem_fornecedores,
            "_tem_passivos_operacionais": tem_passivos_operacionais,
        }

    def _somar_componente(
        self,
        rows: list[Any],
        componente: str,
        absoluto: bool = False,
    ) -> tuple[float, bool]:
        aliases = self.COMPONENTES_DRE[componente]
        # Receita Bruta = faturamento bruto: soma o crédito, sem abater débitos.
        # Demais componentes seguem o líquido (crédito - débito).
        usar_bruto = componente == "receita_bruta"
        total = 0.0
        encontrado = False
        for row in rows:
            if not self._row_dre_matches(row, aliases, componente):
                continue
            encontrado = True
            valor = (
                _float(row["credito"])
                if usar_bruto
                else _float(row["credito"]) - _float(row["debito"])
            )
            total += abs(valor) if absoluto else valor
        return total, encontrado

    def _somar_receita_bruta(self, rows: list[Any]) -> tuple[float, bool]:
        receita_bruta_informada = 0.0
        faturamento_liquido = 0.0
        deducoes_receita = 0.0
        encontrado = False

        for row in rows:
            credito = _float(self._row_get(row, "credito"))
            debito = _float(self._row_get(row, "debito"))
            valor = credito - debito

            if self._row_dre_matches(row, self.COMPONENTES_DRE["receita_bruta"], "receita_bruta"):
                encontrado = True
                valor_bruto = _float(self._row_get(row, "valor_bruto"))
                receita_bruta_informada += max(valor_bruto, credito, valor)
                faturamento_liquido += valor
                continue

            if debito > 0 and self._row_recompoe_receita_bruta(row):
                deducoes_receita += debito

        receita_bruta_recomposta = faturamento_liquido + deducoes_receita
        return max(receita_bruta_informada, receita_bruta_recomposta), encontrado

    @staticmethod
    def _row_recompoe_receita_bruta(row: Any) -> bool:
        valores = [
            PainelDREService._row_get(row, "conta_pai"),
            PainelDREService._row_get(row, "conta_filho"),
            PainelDREService._row_get(row, "rubrica"),
            PainelDREService._row_get(row, "natureza_raw"),
            PainelDREService._row_get(row, "natureza_norm"),
        ]
        return any(
            _normalizar_chave_conta(valor)
            in PainelDREService.RECEITA_BRUTA_RECOMPOSICAO_ALIASES
            for valor in valores
        )

    @staticmethod
    def _row_dre_matches(row: Any, aliases: set[str], componente: str) -> bool:
        codigo_conta = PainelDREService._codigo_conta_row(row)
        resolvido_plano = bool(PainelDREService._row_get(row, "_plano_contas_resolvido"))

        codigos_explicitos = PainelDREService.COMPONENTES_DRE_CODIGOS_EXPLICITOS.get(
            componente, frozenset()
        )
        if codigo_conta in codigos_explicitos:
            return True

        contas_pai_estritas = PainelDREService.COMPONENTES_DRE_CONTA_PAI_ESTRITA.get(
            componente
        )
        if contas_pai_estritas is not None:
            conta_pai = _normalizar_chave_conta(PainelDREService._row_get(row, "conta_pai"))
            if conta_pai not in contas_pai_estritas:
                return False
            if componente in {
                "deducoes",
                "custos_variaveis",
                "gastos_fixos",
                "depreciacao",
            }:
                return bool(codigo_conta or resolvido_plano)

        if (
            componente in PainelDREService.COMPONENTES_DRE_COM_CODIGO_OBRIGATORIO
            and not codigo_conta
            and not resolvido_plano
        ):
            return False

        valores = [
            PainelDREService._row_get(row, "conta_pai"),
            PainelDREService._row_get(row, "conta_filho"),
            PainelDREService._row_get(row, "rubrica"),
            PainelDREService._row_get(row, "natureza_raw"),
            PainelDREService._row_get(row, "natureza_norm"),
        ]
        return any(_normalizar_chave_conta(valor) in aliases for valor in valores)

    @staticmethod
    def _row_get(row: Any, key: str) -> Any:
        try:
            return row[key]
        except (KeyError, IndexError):
            return None

    def _indicador_ncg(
        self,
        componentes: dict[str, Any],
        receita_liquida: float,
    ) -> dict[str, Any]:
        obrigatorios = [
            ("_tem_contas_receber", "Contas a Receber"),
            ("_tem_contas_pagar", "Contas a Pagar"),
        ]
        faltantes = [label for key, label in obrigatorios if not componentes.get(key)]
        valores = {
            "contas_receber": componentes["contas_receber"],
            "contas_pagar": componentes["contas_pagar"],
            "receita_liquida": receita_liquida,
        }
        if faltantes:
            return _indicador_indisponivel(
                "ncg",
                "Necessidade de Capital de Giro",
                faltantes,
                valores,
            )
        ncg = componentes["contas_receber"] - componentes["contas_pagar"]
        return _indicador_calculado(
            "ncg",
            "Necessidade de Capital de Giro",
            ncg,
            _percent(ncg, receita_liquida),
            valores,
        )

    def _objetivos_estrategicos(self, componentes: dict[str, Any]) -> list[dict[str, Any]]:
        receita_liquida = componentes["receita_liquida"]
        folha = componentes["folha_pagamento"]
        imposto_retido = componentes["total_imposto_retido"]
        imposto_retido_acima_meta = componentes.get("total_impostos_retidos_acima_meta", 0.0)
        tem_receita = componentes["_tem_receita_liquida"] and receita_liquida > 0
        tem_folha = componentes["_tem_folha_pagamento"]
        tem_imposto = componentes["_tem_total_imposto_retido"]
        tem_imposto_acima_meta = componentes["_tem_total_impostos_retidos_acima_meta"]

        return [
            self._objetivo_percentual(
                "ifsrl",
                "Índice da Folha sobre a Receita Líquida",
                "IFSRL",
                folha,
                receita_liquida,
                "≤ 30%",
                lambda valor: valor <= 30,
                {"folha_pagamento": folha, "receita_liquida": receita_liquida},
                self._faltantes_objetivo(
                    [
                        (tem_folha, "Custo Total da Folha Pagamento"),
                        (tem_receita, "Receita Líquida positiva"),
                    ]
                ),
            ),
            self._objetivo_razao(
                "iefp",
                "Índice de Eficiência da Folha de Pagamento",
                "IEFP",
                receita_liquida,
                folha,
                "> 2,5",
                lambda valor: valor > 2.5,
                {"receita_liquida": receita_liquida, "folha_pagamento": folha},
                self._faltantes_objetivo(
                    [
                        (tem_receita, "Receita Líquida positiva"),
                        (tem_folha and folha > 0, "Custo Total da Folha Pagamento positivo"),
                    ]
                ),
            ),
            self._objetivo_percentual(
                "iirrl",
                "Total de Impostos Retidos Acima da Meta sobre Receita Líquida",
                "IIRRL",
                imposto_retido_acima_meta,
                receita_liquida,
                "≤ 10%",
                lambda valor: valor <= 10,
                {
                    "total_imposto_retido": imposto_retido,
                    "total_impostos_retidos_acima_meta": imposto_retido_acima_meta,
                    "receita_liquida": receita_liquida,
                },
                self._faltantes_objetivo(
                    [
                        (
                            tem_imposto_acima_meta,
                            "Total de Imposto Retido Acima da Meta",
                        ),
                        (tem_receita, "Receita Líquida positiva"),
                    ]
                ),
            ),
            self._objetivo_valor(
                "itmir",
                "Teto Máximo para o Total de Impostos Retidos",
                "ITMIR",
                imposto_retido,
                "R$",
                "< 7 MM",
                lambda valor: valor < 7_000_000,
                {
                    "total_imposto_retido": imposto_retido,
                    "total_impostos_retidos_acima_meta": imposto_retido_acima_meta,
                },
                self._faltantes_objetivo([(tem_imposto, "Total de Imposto Retido")]),
            ),
        ]

    @staticmethod
    def _faltantes_objetivo(regras: list[tuple[bool, str]]) -> list[str]:
        return [label for disponivel, label in regras if not disponivel]

    def _objetivo_percentual(
        self,
        objetivo_id: str,
        nome: str,
        sigla: str,
        numerador: float,
        denominador: float,
        meta: str,
        atende_meta: Any,
        componentes: dict[str, float],
        faltantes: list[str],
    ) -> dict[str, Any]:
        if faltantes:
            return self._objetivo_indisponivel(
                objetivo_id, nome, sigla, "%", meta, componentes, faltantes
            )
        valor = _percent(numerador, denominador)
        return self._objetivo_calculado(
            objetivo_id, nome, sigla, valor, "%", meta, atende_meta(valor), componentes
        )

    def _objetivo_razao(
        self,
        objetivo_id: str,
        nome: str,
        sigla: str,
        numerador: float,
        denominador: float,
        meta: str,
        atende_meta: Any,
        componentes: dict[str, float],
        faltantes: list[str],
    ) -> dict[str, Any]:
        if faltantes:
            return self._objetivo_indisponivel(
                objetivo_id, nome, sigla, "x", meta, componentes, faltantes
            )
        valor = numerador / denominador
        return self._objetivo_calculado(
            objetivo_id, nome, sigla, valor, "x", meta, atende_meta(valor), componentes
        )

    def _objetivo_valor(
        self,
        objetivo_id: str,
        nome: str,
        sigla: str,
        valor: float,
        unidade: str,
        meta: str,
        atende_meta: Any,
        componentes: dict[str, float],
        faltantes: list[str],
    ) -> dict[str, Any]:
        if faltantes:
            return self._objetivo_indisponivel(
                objetivo_id, nome, sigla, unidade, meta, componentes, faltantes
            )
        return self._objetivo_calculado(
            objetivo_id, nome, sigla, valor, unidade, meta, atende_meta(valor), componentes
        )

    @staticmethod
    def _objetivo_calculado(
        objetivo_id: str,
        nome: str,
        sigla: str,
        valor: float,
        unidade: str,
        meta: str,
        dentro_meta: bool,
        componentes: dict[str, float],
    ) -> dict[str, Any]:
        return {
            "id": objetivo_id,
            "nome": nome,
            "sigla": sigla,
            "status": "calculado",
            "valor": valor,
            "unidade": unidade,
            "meta": meta,
            "meta_status": "ok" if dentro_meta else "alerta",
            "componentes": componentes,
            "componentes_faltantes": [],
        }

    @staticmethod
    def _objetivo_indisponivel(
        objetivo_id: str,
        nome: str,
        sigla: str,
        unidade: str,
        meta: str,
        componentes: dict[str, float],
        faltantes: list[str],
    ) -> dict[str, Any]:
        return {
            "id": objetivo_id,
            "nome": nome,
            "sigla": sigla,
            "status": "indisponivel",
            "valor": None,
            "unidade": unidade,
            "meta": meta,
            "meta_status": "indisponivel",
            "componentes": componentes,
            "componentes_faltantes": faltantes,
        }

    def _ranking(
        self,
        conn: Any,
        expression: str,
        where_sql: str,
        params: list[Any],
        limit: int,
    ) -> list[Any]:
        return conn.execute(
            f"""
            SELECT
                {expression} AS nome,
                SUM(credito) AS credito,
                SUM(debito) AS debito,
                SUM({self.SALDO_DRE_EXPR}) AS saldo,
                COUNT(*) AS total
            FROM dre_lancamentos
            WHERE {where_sql}
            GROUP BY nome
            ORDER BY ABS(SUM({self.SALDO_DRE_EXPR})) DESC, total DESC, nome ASC
            LIMIT ?
            """,
            tuple([*params, limit]),
        ).fetchall()

    def _filter_options(
        self,
        conn: Any,
        expression: str,
        where_sql: str,
        params: list[Any],
    ) -> list[dict[str, Any]]:
        rows = conn.execute(
            f"""
            SELECT
                {expression} AS value,
                COUNT(*) AS total,
                SUM({self.SALDO_DRE_EXPR}) AS saldo
            FROM dre_lancamentos
            WHERE {where_sql}
            GROUP BY value
            ORDER BY total DESC, value ASC
            LIMIT 240
            """,
            tuple(params),
        ).fetchall()
        return self._filter_option_rows(rows)


class PainelFluxoCaixaService(_PainelBaseService):
    """Agrega KPIs, séries e slicers web para Fluxo de Caixa."""

    def __init__(self, db: DatabaseConnection | None = None) -> None:
        super().__init__(db)
        self.indicadores_manuais = FluxoIndicadoresManuaisRepository(self.db)
        self.fluxo_repository = FluxoCaixaRepository(self.db)
        self.fluxo_processamento = FluxoCaixaProcessamentoService()

    BANCO_ARQUIVO_EXPR = (
        "CASE "
        "WHEN INSTR(LOWER(COALESCE(arquivo_origem, '')), '_itau_isolamento_') > 0 "
        "  OR INSTR(LOWER(COALESCE(arquivo_origem, '')), '_itau_isolamento.') > 0 "
        "THEN 'itau_isolamento' "
        "WHEN INSTR(LOWER(COALESCE(arquivo_origem, '')), '_mercantil_') > 0 "
        "  OR INSTR(LOWER(COALESCE(arquivo_origem, '')), '_mercantil.') > 0 "
        "  OR INSTR(LOWER(COALESCE(arquivo_origem, '')), '_marcantil_') > 0 "
        "  OR INSTR(LOWER(COALESCE(arquivo_origem, '')), '_marcantil.') > 0 "
        "THEN 'mercantil' "
        "WHEN INSTR(LOWER(COALESCE(arquivo_origem, '')), '_safra_') > 0 "
        "  OR INSTR(LOWER(COALESCE(arquivo_origem, '')), '_safra.') > 0 "
        "THEN 'safra' "
        "WHEN INSTR(LOWER(COALESCE(arquivo_origem, '')), '_itau_') > 0 "
        "  OR INSTR(LOWER(COALESCE(arquivo_origem, '')), '_itau.') > 0 "
        "THEN 'itau' "
        "WHEN INSTR(LOWER(COALESCE(arquivo_origem, '')), '_cef_') > 0 "
        "  OR INSTR(LOWER(COALESCE(arquivo_origem, '')), '_cef.') > 0 "
        "  OR INSTR(LOWER(COALESCE(arquivo_origem, '')), '_caixa_') > 0 "
        "  OR INSTR(LOWER(COALESCE(arquivo_origem, '')), '_caixa.') > 0 "
        "THEN 'cef' "
        "ELSE NULL END"
    )
    BANCO_EXPR = (
        "CASE WHEN LOWER(COALESCE(NULLIF(TRIM(banco_origem), ''), '')) "
        "NOT IN ('', 'desconhecido', 'sem banco') THEN TRIM(banco_origem) "
        f"WHEN {BANCO_ARQUIVO_EXPR} IS NOT NULL THEN {BANCO_ARQUIVO_EXPR} "
        "ELSE COALESCE(NULLIF(TRIM(banco_origem), ''), 'Sem banco') END"
    )
    TIPO_EXPR = "COALESCE(NULLIF(TRIM(tipo), ''), 'sem_tipo')"
    CLASSIFICACAO_NEUTRA_EXPR = (
        "COALESCE(NULLIF(TRIM(classificacao), ''), NULLIF(TRIM(conta_gerencial), ''), '')"
    )
    CLASSIFICACAO_EXPR = (
        "COALESCE(NULLIF(TRIM(classificacao), ''), "
        "NULLIF(TRIM(conta_gerencial), ''), 'Sem classificação')"
    )
    TRANSFERENCIA_EXPR = (
        f"(tipo = 'transferencia' OR UPPER({CLASSIFICACAO_NEUTRA_EXPR}) LIKE 'TRANSFER%')"
    )
    SALDO_BANCARIO_EXPR = (
        f"(UPPER({CLASSIFICACAO_NEUTRA_EXPR}) LIKE 'SALDO INICIAL %' "
        f"OR UPPER({CLASSIFICACAO_NEUTRA_EXPR}) LIKE 'SALDO FINAL %')"
    )
    NEUTRO_CONSOLIDADO_EXPR = (
        f"({TRANSFERENCIA_EXPR} OR {SALDO_BANCARIO_EXPR})"
    )
    CREDITO_EXPR = (
        f"CASE WHEN NOT {NEUTRO_CONSOLIDADO_EXPR} AND tipo = 'credito' THEN valor ELSE 0 END"
    )
    DEBITO_EXPR = (
        f"CASE WHEN NOT {NEUTRO_CONSOLIDADO_EXPR} AND tipo = 'debito' THEN valor ELSE 0 END"
    )
    SALDO_EXPR = (
        f"CASE WHEN {NEUTRO_CONSOLIDADO_EXPR} THEN 0 "
        "WHEN tipo = 'credito' THEN valor WHEN tipo = 'debito' THEN -valor ELSE valor END"
    )
    BANCO_ROTULOS = {
        "cef": "CEF",
        "itau": "ITAU",
        "itau_isolamento": "ITAU Isolamento",
        "mercantil": "MERCANTIL",
        "safra": "SAFRA",
    }
    BANCO_ALIASES = {
        "caixa": "cef",
        "caixa economica": "cef",
        "itau isolamento": "itau_isolamento",
        "marcantil": "mercantil",
    }
    BANCOS_DESCONHECIDOS = {"", "desconhecido", "sem banco"}
    CONTAS_DESTAQUE = [
        {
            "id": "parcelamento",
            "nome": "Parcelamento",
            "contas": [
                {
                    "label": "17.1 parcelamento",
                    "aliases": _alias_set(["17.1 parcelamento", "Parcelamento"]),
                },
                {
                    "label": "Parcelamento de Impostos Exercício Anterior",
                    "aliases": _alias_set(["Parcelamento de Impostos Exercicio Anterior"]),
                },
            ],
        },
        {
            "id": "folha_pessoal",
            "nome": "Folha Pessoal",
            "contas": [
                {
                    "label": "12.1 salário",
                    "aliases": _alias_set(["12.1 salário", "SALARIO"]),
                },
                {
                    "label": "12.2 13° salário",
                    "aliases": _alias_set(["12.2 13 salário", "13 SALARIO"]),
                },
                {
                    "label": "12.100 previsão férias",
                    "aliases": _alias_set(["12.100 previsão férias", "PREVISAO FERIAS"]),
                },
                {
                    "label": "12.101 previsão 13°",
                    "aliases": _alias_set(["12.101 previsão 13°", "PREVISAO 13"]),
                },
                {"label": "12.3 férias", "aliases": _alias_set(["12.3 férias", "FERIAS"])},
                {
                    "label": "12.4 multa rescisória FGTS",
                    "aliases": _alias_set(
                        ["12.4 multa rescisória FGTS", "MULTA RESCISORIAS FGTS"]
                    ),
                },
                {
                    "label": "12.6 acerto rescisório",
                    "aliases": _alias_set(["12.6 acerto rescisório", "ACERTO RESCISORIOS"]),
                },
                {
                    "label": "12.14 FGTS funcionários",
                    "aliases": _alias_set(["12.14 FGTS funcionários", "FGTS FUNCIONARIOS"]),
                },
            ],
        },
        {
            "id": "locacoes",
            "nome": "Locações",
            "contas": [
                {
                    "label": "7.1 locação Munck",
                    "aliases": _alias_set(
                        ["7.1 locação Munck", "LOCAÇÃO MUNCK / GUINDASTES"]
                    ),
                },
                {
                    "label": "7.2 locação compressor",
                    "aliases": _alias_set(
                        ["7.2 locação compressor", "LOCAÇÃO COMPRESSORES"]
                    ),
                },
                {
                    "label": "7.3 locação gerador",
                    "aliases": _alias_set(["7.3 locação gerador", "LOCAÇÃO GERADOR"]),
                },
                {
                    "label": "7.4 locação equipamentos",
                    "aliases": _alias_set(
                        ["7.4 locação equipamentos", "LOCAÇÃO EQUIPAMENTOS"]
                    ),
                },
                {
                    "label": "7.5 locação empilhadeira",
                    "aliases": _alias_set(
                        ["7.5 locação empilhadeira", "LOCAÇÃO EMPILHADEIRA"]
                    ),
                },
                {
                    "label": "7.6 locação container",
                    "aliases": _alias_set(["7.6 locação container", "LOCAÇÃO CONTAINER"]),
                },
                {
                    "label": "7.7 locação andaimes",
                    "aliases": _alias_set(["7.7 locação andaimes", "LOCAÇÃO ANDAIMES"]),
                },
                {
                    "label": "7.8 locação veículos",
                    "aliases": _alias_set(["7.8 locação veículos", "LOCAÇAO VEICULOS"]),
                },
                {
                    "label": "7.9 locação caçamba p/resíduos",
                    "aliases": _alias_set(
                        ["7.9 locação caçamba p/resíduos", "LOCAÇAO CAÇAMBA P/ RESIDUOS"]
                    ),
                },
                {
                    "label": "7.10 locação rádios de comunicação",
                    "aliases": _alias_set(
                        [
                            "7.10 locação rádios de comunicação",
                            "LOCAÇAO RADIOS DE COMUNICAÇÃO",
                        ]
                    ),
                },
                {
                    "label": "7.11 locação banheiro químico",
                    "aliases": _alias_set(
                        ["7.11 locação banheiro químico", "LOCAÇAO BANHEIRO QUIMICO"]
                    ),
                },
                {
                    "label": "7.12 locação plataforma elevatório",
                    "aliases": _alias_set(
                        [
                            "7.12 locação plataforma elevatório",
                            "LOCAÇAO PLATAFORMA ELEVATORIA",
                        ]
                    ),
                },
                {
                    "label": "7.13 locação ambulância",
                    "aliases": _alias_set(
                        ["7.13 locação ambulância", "LOCAÇÃO AMBULANCIA / PRIMEIROS SOCORROS"]
                    ),
                },
            ],
        },
        {
            "id": "fornecedores",
            "nome": "Fornecedores",
            "contas": [
                {
                    "label": "4.1 tintas e solventes",
                    "aliases": _alias_set(["4.1 tintas e solventes", "TINTAS E SOLVENTES"]),
                },
                {"label": "4.2 abrasivos", "aliases": _alias_set(["4.2 abrasivos", "ABRASIVOS"])},
                {
                    "label": "4.3 outros materiais",
                    "aliases": _alias_set(
                        ["4.3 outros materiais", "OUTROS MATERIAIS DE APLICAÇÃO"]
                    ),
                },
                {
                    "label": "8.9 materiais de consumo em obras",
                    "aliases": _alias_set(
                        [
                            "8.9 materiais de consumo em obras",
                            "MATERIAIS DE CONSUMO EM OBRAS",
                        ]
                    ),
                },
                {"label": "13.1 EPIS", "aliases": _alias_set(["13.1 EPIS", "EPIS"])},
            ],
        },
        {
            "id": "gastos_manutencao",
            "nome": "Gastos com manutenção",
            "contas": [
                {
                    "label": "6.3 manutenção / peças veículos",
                    "aliases": _alias_set(
                        ["6.3 manutenção / peças veículos", "MANUTENÇÃO/PEÇAS VEICULOS"]
                    ),
                },
                {
                    "label": "8.14 manutenção / peças maquinas e equipamentos",
                    "aliases": _alias_set(
                        [
                            "8.14 manutenção / peças maquinas e equipamentos",
                            "MANUTENÇÃO/PEÇAS MAQUINAS E EQUIPAMENTOS",
                        ]
                    ),
                },
                {
                    "label": "15.7 despesa manutenção sede",
                    "aliases": _alias_set(
                        ["15.7 despesa manutenção sede", "DESPESA MANUTENÇÃO SEDE"]
                    ),
                },
            ],
        },
    ]

    def obter_painel(
        self,
        ano: int | None = None,
        meses: list[int] | None = None,
        banco: list[str] | None = None,
        tipo: list[str] | None = None,
        classificacao: list[str] | None = None,
    ) -> dict[str, Any]:
        ano_ref = ano if ano is not None else self._ano_mais_recente("fluxo_movimentos")
        meses_ref = sorted(set(meses or []))
        banco_ref = _normalizar_lista_texto(banco)
        tipo_ref = _normalizar_lista_texto(tipo)
        classificacao_ref = _normalizar_lista_texto(classificacao)

        _validar_ano(ano_ref)
        _validar_meses(meses_ref)

        # Transferências e linhas de saldo movimentam/rastreiam bancos individuais,
        # mas não alteram o caixa consolidado exibido no Fluxo gerado.
        where = ["competencia_ano = ?", f"NOT {self.NEUTRO_CONSOLIDADO_EXPR}"]
        params: list[Any] = [ano_ref]
        _append_in(where, params, "competencia_mes", meses_ref)
        _append_in(where, params, self.BANCO_EXPR, banco_ref)
        _append_in(where, params, self.TIPO_EXPR, tipo_ref)
        _append_in(where, params, self.CLASSIFICACAO_EXPR, classificacao_ref)
        where_sql = " AND ".join(where)

        period_where = ["competencia_ano = ?", f"NOT {self.NEUTRO_CONSOLIDADO_EXPR}"]
        period_params: list[Any] = [ano_ref]
        _append_in(period_where, period_params, "competencia_mes", meses_ref)
        period_where_sql = " AND ".join(period_where)

        # A alocação bancária precisa considerar as transferências, pois elas
        # alteram o saldo de cada banco, mesmo sendo neutras no consolidado.
        saldos_where = ["competencia_ano = ?"]
        saldos_params: list[Any] = [ano_ref]
        _append_in(saldos_where, saldos_params, "competencia_mes", meses_ref)
        _append_in(saldos_where, saldos_params, self.BANCO_EXPR, banco_ref)
        saldos_where_sql = " AND ".join(saldos_where)

        with self.db.get_connection() as conn:
            meses_disponiveis = [
                int(row["competencia_mes"])
                for row in conn.execute(
                    """
                    SELECT DISTINCT competencia_mes
                    FROM fluxo_movimentos
                    WHERE competencia_ano = ?
                    ORDER BY competencia_mes
                    """,
                    (ano_ref,),
                ).fetchall()
            ]

            kpis = conn.execute(
                f"""
                SELECT
                    COUNT(*) AS total_movimentos,
                    SUM({self.CREDITO_EXPR}) AS total_creditos,
                    SUM({self.DEBITO_EXPR}) AS total_debitos,
                    SUM({self.SALDO_EXPR}) AS saldo_liquido,
                    COUNT(DISTINCT {self.BANCO_EXPR}) AS total_bancos,
                    COUNT(DISTINCT {self.CLASSIFICACAO_EXPR}) AS total_classificacoes
                FROM fluxo_movimentos
                WHERE {where_sql}
                """,
                tuple(params),
            ).fetchone()
            series = conn.execute(
                f"""
                SELECT
                    competencia_mes AS mes,
                    SUM({self.CREDITO_EXPR}) AS credito,
                    SUM({self.DEBITO_EXPR}) AS debito,
                    SUM({self.SALDO_EXPR}) AS saldo,
                    COUNT(*) AS total
                FROM fluxo_movimentos
                WHERE {where_sql}
                GROUP BY competencia_mes
                ORDER BY competencia_mes
                """,
                tuple(params),
            ).fetchall()
            ranking_bancos = self._ranking(conn, self.BANCO_EXPR, where_sql, params, limit=14)
            ranking_classificacoes = self._ranking(
                conn, self.CLASSIFICACAO_EXPR, where_sql, params, limit=14
            )
            recentes = conn.execute(
                f"""
                SELECT
                    id,
                    data_movimento AS data,
                    descricao,
                    tipo,
                    valor,
                    {self.SALDO_EXPR} AS saldo_liquido,
                    {self.BANCO_EXPR} AS banco,
                    {self.CLASSIFICACAO_EXPR} AS classificacao
                FROM fluxo_movimentos
                WHERE {where_sql}
                ORDER BY data_movimento DESC, id DESC
                LIMIT 12
                """,
                tuple(params),
            ).fetchall()
            contas_destaque_rows = conn.execute(
                f"""
                SELECT
                    id,
                    NULLIF(TRIM(classificacao), '') AS classificacao,
                    NULLIF(TRIM(conta_gerencial), '') AS conta_gerencial,
                    valor
                FROM fluxo_movimentos
                WHERE {where_sql}
                  AND tipo = 'debito'
                """,
                tuple(params),
            ).fetchall()
            saldos_por_banco_rows = conn.execute(
                f"""
                SELECT
                    id,
                    {self.BANCO_EXPR} AS banco,
                    competencia_mes AS mes,
                    data_movimento AS data,
                    saldo,
                    tipo,
                    {self.CLASSIFICACAO_EXPR} AS classificacao,
                    valor,
                    arquivo_origem,
                    linha_origem
                FROM fluxo_movimentos
                WHERE {saldos_where_sql}
                ORDER BY banco ASC, competencia_mes ASC, data_movimento ASC,
                         COALESCE(linha_origem, 0) ASC, id ASC
                """,
                tuple(saldos_params),
            ).fetchall()

            filtros_banco = self._filter_options(
                conn, self.BANCO_EXPR, period_where_sql, period_params
            )
            for opcao in filtros_banco:
                opcao["label"] = self._rotulo_banco(opcao["value"])

            filtros = {
                "anos": [
                    int(row["competencia_ano"])
                    for row in conn.execute(
                        """
                        SELECT DISTINCT competencia_ano
                        FROM fluxo_movimentos
                        ORDER BY competencia_ano DESC
                        """
                    ).fetchall()
                ],
                "meses": meses_disponiveis,
                "banco": filtros_banco,
                "tipo": self._filter_options(conn, self.TIPO_EXPR, period_where_sql, period_params),
                "classificacao": self._filter_options(
                    conn, self.CLASSIFICACAO_EXPR, period_where_sql, period_params
                ),
            }

            indicadores_manuais = self.indicadores_manuais.get_by_ano(ano_ref, conn)
            saldo_ano_anterior = (
                float(indicadores_manuais.saldo_ano_anterior) if indicadores_manuais else 0
            )
            saldo_final_periodo = self._saldo_final_periodo(
                conn,
                ano_ref,
                meses_ref,
                meses_disponiveis,
                saldo_ano_anterior,
            )
            series_mensais = self._series_rows(series, "movimentos")
            series_mensais = self._aplicar_saldos_finais_series_fluxo(
                conn,
                ano_ref,
                meses_ref,
                meses_disponiveis,
                saldo_ano_anterior,
                series_mensais,
            )

        total_movimentos = int(kpis["total_movimentos"] or 0)
        total_creditos = _float(kpis["total_creditos"])
        total_debitos = _float(kpis["total_debitos"])
        saldo_liquido = _float(kpis["saldo_liquido"])
        contas_destaque = self._contas_destaque(contas_destaque_rows, total_debitos)
        saldos_por_banco = self._saldos_por_banco(saldos_por_banco_rows)
        return {
            "success": True,
            "periodo": self._periodo_payload(ano_ref, meses_ref, meses_disponiveis),
            "filtros_disponiveis": filtros,
            "filtros_aplicados": {
                "ano": ano_ref,
                "meses": meses_ref,
                "banco": banco_ref,
                "tipo": tipo_ref,
                "classificacao": classificacao_ref,
            },
            "kpis": {
                "total_movimentos": total_movimentos,
                "total_creditos": total_creditos,
                "total_debitos": total_debitos,
                "saldo_liquido": saldo_liquido,
                "saldo_ano_anterior": saldo_ano_anterior,
                "saldo_final_periodo": saldo_final_periodo,
                "saldo_com_ano_anterior": saldo_final_periodo,
                "total_bancos": int(kpis["total_bancos"] or 0),
                "total_classificacoes": int(kpis["total_classificacoes"] or 0),
                "ticket_medio": (
                    saldo_liquido / total_movimentos if total_movimentos else 0
                ),
            },
            "series_mensais": series_mensais,
            "ranking_bancos": self._rotular_bancos_payload(
                self._ranking_rows(ranking_bancos, "movimentos")
            ),
            "saldos_por_banco": saldos_por_banco,
            "ranking_classificacoes": self._ranking_rows(ranking_classificacoes, "movimentos"),
            "contas_destaque": contas_destaque,
            "equilibrio_contas_destaque": self._equilibrio_contas_destaque(
                contas_destaque, total_creditos, total_debitos, saldo_liquido
            ),
            "indicadores_manuais": self._indicadores_manuais_payload(
                indicadores_manuais,
                ano_ref,
            ),
            "movimentos_recentes": [
                {
                    "id": row["id"],
                    "data": row["data"],
                    "descricao": row["descricao"],
                    "tipo": row["tipo"],
                    "valor": _float(row["valor"]),
                    "saldo_liquido": _float(row["saldo_liquido"]),
                    "banco": row["banco"],
                    "classificacao": row["classificacao"],
                }
                for row in recentes
            ],
        }

    def _saldo_final_periodo(
        self,
        conn: Any,
        ano: int,
        meses_ref: list[int],
        meses_disponiveis: list[int],
        saldo_ano_anterior: float,
    ) -> float:
        meses_base = meses_ref or meses_disponiveis
        if not meses_base:
            return saldo_ano_anterior

        mes_fechamento = max(meses_base)
        meses_template = self._meses_template_saldo(meses_ref, meses_disponiveis, mes_fechamento)
        if not meses_template:
            return saldo_ano_anterior
        try:
            return self._saldo_final_template(
                conn,
                ano,
                meses_template,
                mes_fechamento,
                saldo_ano_anterior,
            )
        except Exception as exc:
            logger.exception("Falha ao calcular saldo final do painel pelo template: %s", exc)

        placeholders = ",".join("?" * len(meses_template))
        rows = conn.execute(
            f"""
            SELECT
                competencia_mes AS mes,
                SUM({self.SALDO_EXPR}) AS saldo
            FROM fluxo_movimentos
            WHERE competencia_ano = ?
              AND competencia_mes IN ({placeholders})
            GROUP BY competencia_mes
            ORDER BY competencia_mes
            """,
            tuple([ano, *meses_template]),
        ).fetchall()

        saldo = saldo_ano_anterior if 1 in meses_template else 0.0
        if 1 not in meses_template:
            saldo += sum(
                float(valor)
                for valor in self.fluxo_repository.movimentos.get_saldos_finais_anteriores(
                    ano,
                    min(meses_template),
                    conn,
                ).values()
            )
        for row in rows:
            saldo += _float(row["saldo"])
        return saldo

    @staticmethod
    def _meses_template_saldo(
        meses_ref: list[int],
        meses_disponiveis: list[int],
        mes_fechamento: int,
    ) -> list[int]:
        meses_base = meses_ref or [mes for mes in meses_disponiveis if mes <= mes_fechamento]
        return sorted({mes for mes in meses_base if mes <= mes_fechamento})

    def _aplicar_saldos_finais_series_fluxo(
        self,
        conn: Any,
        ano: int,
        meses_ref: list[int],
        meses_disponiveis: list[int],
        saldo_ano_anterior: float,
        series: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        if not series:
            return series

        saldos_finais: dict[int, float] = {}
        for mes in sorted({int(item["mes"]) for item in series}):
            meses_template = self._meses_template_saldo(meses_ref, meses_disponiveis, mes)
            if not meses_template:
                saldos_finais[mes] = saldo_ano_anterior
                continue
            try:
                saldos_finais[mes] = self._saldo_final_template(
                    conn,
                    ano,
                    meses_template,
                    mes,
                    saldo_ano_anterior,
                )
            except Exception as exc:
                logger.exception(
                    "Falha ao calcular saldo final mensal do painel pelo template: %s",
                    exc,
                )
                saldos_finais[mes] = _float(
                    next((item["saldo"] for item in series if int(item["mes"]) == mes), 0)
                )

        return [
            {
                **item,
                "saldo_final": saldos_finais.get(int(item["mes"]), _float(item["saldo"])),
            }
            for item in series
        ]

    def _saldo_final_template(
        self,
        conn: Any,
        ano: int,
        meses_utilizados: list[int],
        mes_fechamento: int,
        saldo_ano_anterior: float,
    ) -> float:
        movimentos_db = self.fluxo_repository.movimentos.get_by_meses(ano, meses_utilizados, conn)
        movimentos = [mov.to_movimento() for mov in movimentos_db]
        saldo_manual = Decimal(str(saldo_ano_anterior)) if 1 in meses_utilizados else Decimal("0")
        saldos_iniciais = self.fluxo_repository.movimentos.get_saldos_finais_anteriores(
            ano,
            min(meses_utilizados),
            conn,
        )

        with TemplateWriter(settings.template_fluxo_path) as writer:
            wb = writer._wb
            if (
                wb is None
                or "Fluxo de Caixa " not in wb.sheetnames
                or "Consolidado" not in wb.sheetnames
            ):
                return saldo_ano_anterior

            fluxo_ws = wb["Fluxo de Caixa "]
            consolidado_ws = wb["Consolidado"]
            classificacoes_template = self.fluxo_processamento._mapear_classificacoes_template(
                wb,
                consolidado_ws,
                consolidado_ws.max_row,
            )
            linhas = self._linhas_consolidadas_fluxo_template(
                movimentos,
                ano,
                saldo_manual,
                saldos_iniciais,
                classificacoes_template,
            )
            if saldo_manual and saldo_manual != Decimal("0"):
                self.fluxo_processamento._garantir_saldo_ano_anterior_no_fluxo(writer)

            agregados = self._agregados_apoio_fluxo(linhas)
            linha_saldo_final = self.fluxo_processamento._encontrar_linha_rotulo(
                fluxo_ws,
                "(=) SALDO FINAL",
                coluna=2,
            )
            if not linha_saldo_final:
                return saldo_ano_anterior

            coluna_mes = 3 + mes_fechamento
            cache: dict[tuple[int, int], float] = {}
            return self._avaliar_celula_fluxo(
                fluxo_ws,
                linha_saldo_final,
                coluna_mes,
                agregados,
                cache,
            )

    def _linhas_consolidadas_fluxo_template(
        self,
        movimentos: list[FCMovimento],
        ano: int,
        saldo_manual: Decimal,
        saldos_iniciais: dict[str, Decimal],
        classificacoes_template: dict[str, str],
    ) -> list[list]:
        linhas: list[list] = []
        if saldo_manual and saldo_manual != Decimal("0"):
            linhas.append(
                self.fluxo_processamento._linha_saldo_ano_anterior(
                    ano,
                    1,
                    saldo_manual,
                    row_number=2,
                )
            )

        itens = self.fluxo_processamento._linhas_movimentos_com_saldos(
            movimentos,
            saldos_iniciais_por_banco=saldos_iniciais,
            incluir_saldo_inicial_derivado=not bool(saldo_manual and saldo_manual != Decimal("0")),
        )
        for item in itens:
            row_number = 2 + len(linhas)
            if isinstance(item, FCMovimento):
                linhas.append(
                    self.fluxo_processamento._converter_movimento_para_linha(
                        item,
                        row_number,
                        classificacoes_template,
                    )
                )
            else:
                linhas.append(
                    self.fluxo_processamento._normalizar_linha_consolidado(item, row_number)
                )
        return linhas

    @staticmethod
    def _agregados_apoio_fluxo(linhas: list[list]) -> dict[str, dict[int, dict[str, float]]]:
        agregados: dict[str, dict[int, dict[str, float]]] = {}
        for linha in linhas:
            data = linha[0]
            classificacao = str(linha[5] or "").strip()
            if (
                not classificacao
                or eh_transferencia_classificacao(classificacao)
                or not getattr(data, "month", None)
            ):
                continue
            mes = int(data.month)
            totais = agregados.setdefault(classificacao, {}).setdefault(
                mes,
                {"credito": 0.0, "debito": 0.0},
            )
            totais["credito"] += _float(linha[2])
            totais["debito"] += _float(linha[3])
        return agregados

    def _avaliar_celula_fluxo(
        self,
        ws: Any,
        row: int,
        col: int,
        agregados: dict[str, dict[int, dict[str, float]]],
        cache: dict[tuple[int, int], float],
    ) -> float:
        chave = (row, col)
        if chave in cache:
            return cache[chave]

        valor = ws.cell(row=row, column=col).value
        if isinstance(valor, (int, float)):
            return float(valor)
        if not isinstance(valor, str) or not valor.startswith("="):
            return 0.0

        cache[chave] = self._avaliar_expressao_fluxo(
            ws,
            valor[1:],
            row,
            col,
            agregados,
            cache,
        )
        return cache[chave]

    def _avaliar_expressao_fluxo(
        self,
        ws: Any,
        expr: str,
        row: int,
        col: int,
        agregados: dict[str, dict[int, dict[str, float]]],
        cache: dict[tuple[int, int], float],
    ) -> float:
        expr = expr.strip().replace("$", "")
        expr_upper = expr.upper()
        mes = self._mes_coluna_fluxo(col)
        if expr_upper.startswith("IFERROR("):
            try:
                return self._avaliar_expressao_fluxo(
                    ws,
                    self._primeiro_argumento_funcao(expr),
                    row,
                    col,
                    agregados,
                    cache,
                )
            except Exception:
                return 0.0
        if "INDEX(APOIO!" in expr_upper:
            return self._valor_apoio_fluxo(ws.cell(row=row, column=3).value, expr, agregados, mes)
        if expr_upper.startswith("SUMIFS(CONSOLIDADO!"):
            return self._valor_apoio_fluxo(ws.cell(row=row, column=3).value, expr, agregados, mes)
        if expr_upper.startswith("SUM("):
            return sum(
                self._avaliar_termo_fluxo(ws, item, col, agregados, cache)
                for item in self._argumentos_funcao(expr)
            )
        if expr_upper.startswith("SUBTOTAL("):
            argumentos = self._argumentos_funcao(expr)
            return sum(
                self._avaliar_termo_fluxo(ws, item, col, agregados, cache)
                for item in argumentos[1:]
            )

        total = 0.0
        sinal = 1.0
        for parte in self._split_soma_subtracao(expr):
            if parte == "+":
                sinal = 1.0
                continue
            if parte == "-":
                sinal = -1.0
                continue
            total += sinal * self._avaliar_termo_fluxo(ws, parte, col, agregados, cache)
        return total

    def _avaliar_termo_fluxo(
        self,
        ws: Any,
        termo: str,
        col: int,
        agregados: dict[str, dict[int, dict[str, float]]],
        cache: dict[tuple[int, int], float],
    ) -> float:
        termo = termo.strip().replace("$", "")
        if not termo:
            return 0.0
        if re.fullmatch(r"-?\d+(?:\.\d+)?", termo):
            return float(termo)
        range_match = re.fullmatch(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", termo)
        if range_match:
            col_ini = column_index_from_string(range_match.group(1))
            row_ini = int(range_match.group(2))
            col_fim = column_index_from_string(range_match.group(3))
            row_fim = int(range_match.group(4))
            return sum(
                self._avaliar_celula_fluxo(ws, r, c, agregados, cache)
                for r in range(row_ini, row_fim + 1)
                for c in range(col_ini, col_fim + 1)
            )
        cell_match = re.fullmatch(r"([A-Z]+)(\d+)", termo)
        if cell_match:
            return self._avaliar_celula_fluxo(
                ws,
                int(cell_match.group(2)),
                column_index_from_string(cell_match.group(1)),
                agregados,
                cache,
            )
        return self._avaliar_expressao_fluxo(ws, termo, 1, col, agregados, cache)

    @staticmethod
    def _mes_coluna_fluxo(col: int) -> int:
        mes = col - 3
        return mes if 1 <= mes <= 12 else 0

    @staticmethod
    def _valor_apoio_fluxo(
        label: Any,
        expr: str,
        agregados: dict[str, dict[int, dict[str, float]]],
        mes: int,
    ) -> float:
        label_texto = str(label or "").strip()
        if not label_texto:
            return 0.0
        chave_valor = "debito" if "+1" in expr or "CONSOLIDADO!D" in expr.upper() else "credito"
        return agregados.get(label_texto, {}).get(mes, {}).get(chave_valor, 0.0)

    @staticmethod
    def _primeiro_argumento_funcao(expr: str) -> str:
        argumentos = PainelFluxoCaixaService._argumentos_funcao(expr)
        return argumentos[0] if argumentos else ""

    @staticmethod
    def _argumentos_funcao(expr: str) -> list[str]:
        inicio = expr.find("(")
        fim = expr.rfind(")")
        if inicio < 0 or fim <= inicio:
            return []
        return PainelFluxoCaixaService._split_top_level(expr[inicio + 1 : fim], ",")

    @staticmethod
    def _split_top_level(texto: str, separador: str) -> list[str]:
        partes: list[str] = []
        nivel = 0
        inicio = 0
        for idx, char in enumerate(texto):
            if char == "(":
                nivel += 1
            elif char == ")":
                nivel -= 1
            elif char == separador and nivel == 0:
                partes.append(texto[inicio:idx].strip())
                inicio = idx + 1
        partes.append(texto[inicio:].strip())
        return partes

    @staticmethod
    def _split_soma_subtracao(expr: str) -> list[str]:
        partes: list[str] = []
        nivel = 0
        inicio = 0
        for idx, char in enumerate(expr):
            if char == "(":
                nivel += 1
            elif char == ")":
                nivel -= 1
            elif char in "+-" and nivel == 0 and idx > inicio:
                partes.append(expr[inicio:idx].strip())
                partes.append(char)
                inicio = idx + 1
        partes.append(expr[inicio:].strip())
        return [parte for parte in partes if parte]

    @classmethod
    def _impacto_bancario_movimento(cls, row: Any) -> float:
        valor = abs(_float(row["valor"]))
        tipo = str(row["tipo"] or "").strip().lower()
        classificacao = cls._classificacao_upper(row)
        if cls._eh_saldo_bancario(row):
            return 0
        if tipo == "transferencia" or classificacao.startswith("TRANSFER"):
            return -valor if "EMITIDA" in classificacao else valor
        return -valor if tipo == "debito" else valor

    @classmethod
    def _normalizar_banco_chave(cls, banco: Any) -> str:
        texto = str(banco or "").strip()
        if not texto:
            return ""
        sem_acento = unicodedata.normalize("NFKD", texto)
        sem_acento = "".join(ch for ch in sem_acento if not unicodedata.combining(ch))
        chave = re.sub(r"[_-]+", " ", sem_acento)
        chave = re.sub(r"\s+", " ", chave).strip().lower()
        return cls.BANCO_ALIASES.get(chave, chave)

    @classmethod
    def _rotulo_banco(cls, banco: Any) -> str:
        texto = str(banco or "").strip()
        chave = cls._normalizar_banco_chave(texto)
        if not chave:
            return "Sem banco"
        return cls.BANCO_ROTULOS.get(chave, texto.upper())

    @classmethod
    def _banco_desconhecido(cls, banco: Any) -> bool:
        return cls._normalizar_banco_chave(banco) in cls.BANCOS_DESCONHECIDOS

    @staticmethod
    def _classificacao_texto(row: Any) -> str:
        return str(row["classificacao"] or "").strip()

    @classmethod
    def _classificacao_upper(cls, row: Any) -> str:
        return cls._classificacao_texto(row).upper()

    @classmethod
    def _eh_saldo_inicial_bancario(cls, row: Any) -> bool:
        return cls._classificacao_upper(row).startswith("SALDO INICIAL ")

    @classmethod
    def _eh_saldo_final_bancario(cls, row: Any) -> bool:
        return cls._classificacao_upper(row).startswith("SALDO FINAL ")

    @classmethod
    def _eh_saldo_bancario(cls, row: Any) -> bool:
        return cls._eh_saldo_inicial_bancario(row) or cls._eh_saldo_final_bancario(row)

    @classmethod
    def _banco_linha_saldo(cls, row: Any) -> str:
        classificacao = cls._classificacao_texto(row)
        upper = classificacao.upper()
        for prefixo in ("SALDO INICIAL ", "SALDO FINAL "):
            if upper.startswith(prefixo):
                banco_classificacao = classificacao[len(prefixo) :].strip()
                if not cls._banco_desconhecido(banco_classificacao):
                    return banco_classificacao
                return str(row["banco"] or "").strip()
        return str(row["banco"] or "").strip()

    @staticmethod
    def _valor_linha_saldo_bancario(row: Any) -> float:
        valor = abs(_float(row["valor"]))
        return valor if valor else abs(_float(row["saldo"]))

    @classmethod
    def _saldo_inicial_banco(cls, movimentos: list[Any], saldos_iniciais: list[Any]) -> float:
        if saldos_iniciais:
            return cls._valor_linha_saldo_bancario(saldos_iniciais[0])
        if not movimentos:
            return 0.0
        primeiro = movimentos[0]
        primeiro_saldo = primeiro["saldo"]
        if primeiro_saldo is None:
            return 0.0
        return _float(primeiro_saldo) - cls._impacto_bancario_movimento(primeiro)

    @classmethod
    def _rotular_bancos_payload(cls, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        return [{**row, "nome": cls._rotulo_banco(row["nome"])} for row in rows]

    @classmethod
    def _saldos_por_banco(cls, rows: list[Any]) -> list[dict[str, Any]]:
        agrupados: dict[str, dict[str, Any]] = {}
        for row in rows:
            saldo_inicial = cls._eh_saldo_inicial_bancario(row)
            saldo_final = cls._eh_saldo_final_bancario(row)
            banco_base = (
                cls._banco_linha_saldo(row) if saldo_inicial or saldo_final else row["banco"]
            )

            if saldo_inicial and cls._banco_desconhecido(banco_base):
                continue

            banco_chave = cls._normalizar_banco_chave(banco_base)
            if not banco_chave:
                banco_chave = str(banco_base or "").strip()

            grupo = agrupados.setdefault(
                banco_chave,
                {
                    "nome": cls._rotulo_banco(banco_base),
                    "movimentos": [],
                    "saldos_iniciais": [],
                    "saldos_finais": [],
                },
            )
            if saldo_inicial:
                grupo["saldos_iniciais"].append(row)
            elif saldo_final:
                grupo["saldos_finais"].append(row)
            else:
                grupo["movimentos"].append(row)

        saldos = []
        for grupo in agrupados.values():
            movimentos = grupo["movimentos"]
            saldos_iniciais = grupo["saldos_iniciais"]
            saldos_finais = grupo["saldos_finais"]
            if not movimentos and not saldos_finais:
                continue

            saldo_inicial = cls._saldo_inicial_banco(movimentos, saldos_iniciais)
            saldo_calculado = saldo_inicial
            saldo_final = saldo_inicial
            for movimento in movimentos:
                saldo_calculado += cls._impacto_bancario_movimento(movimento)
                if movimento["saldo"] is not None:
                    saldo_final = _float(movimento["saldo"])
                    saldo_calculado = saldo_final
                else:
                    saldo_final = saldo_calculado
            if saldos_finais:
                saldo_final = cls._valor_linha_saldo_bancario(saldos_finais[-1])
            saldos.append(
                {
                    "nome": grupo["nome"],
                    "saldo_inicial": saldo_inicial,
                    "saldo_final": saldo_final,
                    "movimentos": len(movimentos) + len(saldos_finais),
                }
            )
        return sorted(saldos, key=lambda saldo: (-abs(saldo["saldo_final"]), saldo["nome"]))

    @staticmethod
    def _indicadores_manuais_payload(
        indicadores: Any | None,
        ano: int,
    ) -> dict[str, Any]:
        return {
            "existe": indicadores is not None,
            "ano": ano,
            "saldo_ano_anterior": (
                float(indicadores.saldo_ano_anterior) if indicadores else 0
            ),
            "created_at": indicadores.created_at.isoformat() if indicadores else None,
            "updated_at": indicadores.updated_at.isoformat() if indicadores else None,
        }

    def _contas_destaque(self, rows: list[Any], total_debitos: float) -> list[dict[str, Any]]:
        acumulado = {
            grupo["id"]: {
                "total": 0.0,
                "movimentos": set(),
                "encontradas": set(),
            }
            for grupo in self.CONTAS_DESTAQUE
        }

        for row in rows:
            valor = _float(row["valor"])
            if not valor:
                continue

            partes = self._partes_conta_destaque(
                row["classificacao"] or row["conta_gerencial"]
            )
            contabilizadas_no_movimento: set[tuple[str, str, str]] = set()
            for parte in partes:
                codigo_parte = parte["codigo"]
                if not codigo_parte:
                    continue
                for grupo in self.CONTAS_DESTAQUE:
                    encontrou = False
                    for conta in grupo["contas"]:
                        codigo_conta = self._codigo_conta_label(conta["label"])
                        if not codigo_conta or codigo_parte != codigo_conta:
                            continue
                        chave = (grupo["id"], conta["label"], codigo_parte)
                        if chave in contabilizadas_no_movimento:
                            encontrou = True
                            break
                        peso = parte["percentual"] if parte["percentual"] is not None else 1.0
                        acumulado[grupo["id"]]["total"] += valor * peso
                        acumulado[grupo["id"]]["movimentos"].add(row["id"])
                        acumulado[grupo["id"]]["encontradas"].add(conta["label"])
                        contabilizadas_no_movimento.add(chave)
                        encontrou = True
                        break
                    if encontrou:
                        break

        destaques = []
        for grupo in self.CONTAS_DESTAQUE:
            labels_encontradas = acumulado[grupo["id"]]["encontradas"]
            encontradas = [
                conta["label"] for conta in grupo["contas"] if conta["label"] in labels_encontradas
            ]
            faltantes = [
                conta["label"]
                for conta in grupo["contas"]
                if conta["label"] not in labels_encontradas
            ]
            total = acumulado[grupo["id"]]["total"]
            destaques.append(
                {
                    "id": grupo["id"],
                    "nome": grupo["nome"],
                    "total": total,
                    "movimentos": len(acumulado[grupo["id"]]["movimentos"]),
                    "participacao_saidas_percentual": _percent(total, total_debitos),
                    "contas_encontradas": encontradas,
                    "contas_faltantes": faltantes,
                }
            )
        return destaques

    @staticmethod
    def _equilibrio_contas_destaque(
        destaques: list[dict[str, Any]],
        total_creditos: float,
        total_debitos: float,
        saldo_liquido: float,
    ) -> dict[str, Any]:
        total_destaques = sum(_float(grupo["total"]) for grupo in destaques)
        return {
            "total_contas_destaque": total_destaques,
            "outras_saidas": max(total_debitos - total_destaques, 0),
            "participacao_saidas_percentual": _percent(total_destaques, total_debitos),
            "cobertura_entradas_percentual": _percent(total_creditos, total_destaques),
            "saldo_liquido": saldo_liquido,
            "saldo_apos_contas_destaque": total_creditos - total_destaques,
            "status": "equilibrado" if saldo_liquido >= 0 else "deficit",
        }

    @staticmethod
    def _partes_conta_destaque(texto: Any) -> list[dict[str, float | str | None]]:
        partes = []
        for parte in re.split(r";|\n+", str(texto or "")):
            trecho = parte.strip()
            if not trecho:
                continue
            percentual = None
            percentual_match = re.search(r"\((\d+(?:[,.]\d+)?)\s*%\)", trecho)
            if percentual_match:
                percentual_texto = percentual_match.group(1)
                if "," in percentual_texto:
                    percentual_texto = percentual_texto.replace(".", "").replace(",", ".")
                try:
                    percentual = float(percentual_texto) / 100
                except ValueError:
                    percentual = None
            nome = re.sub(r"\([^)]*%\)", "", trecho).strip(" -;\t")
            chave = _normalizar_chave_conta(nome)
            codigo = _extrair_codigo_gerencial(nome)
            if chave:
                partes.append({"chave": chave, "codigo": codigo, "percentual": percentual})
        return partes

    @staticmethod
    def _codigo_conta_label(label: str) -> str | None:
        return _extrair_codigo_gerencial(label)

    def _ranking(
        self,
        conn: Any,
        expression: str,
        where_sql: str,
        params: list[Any],
        limit: int,
    ) -> list[Any]:
        return conn.execute(
            f"""
            SELECT
                {expression} AS nome,
                SUM({self.CREDITO_EXPR}) AS credito,
                SUM({self.DEBITO_EXPR}) AS debito,
                SUM({self.SALDO_EXPR}) AS saldo,
                COUNT(*) AS total
            FROM fluxo_movimentos
            WHERE {where_sql}
            GROUP BY nome
            ORDER BY ABS(SUM({self.SALDO_EXPR})) DESC, total DESC, nome ASC
            LIMIT ?
            """,
            tuple([*params, limit]),
        ).fetchall()

    def _filter_options(
        self,
        conn: Any,
        expression: str,
        where_sql: str,
        params: list[Any],
    ) -> list[dict[str, Any]]:
        rows = conn.execute(
            f"""
            SELECT
                {expression} AS value,
                COUNT(*) AS total,
                SUM({self.SALDO_EXPR}) AS saldo
            FROM fluxo_movimentos
            WHERE {where_sql}
            GROUP BY value
            ORDER BY total DESC, value ASC
            LIMIT 240
            """,
            tuple(params),
        ).fetchall()
        return self._filter_option_rows(rows)
