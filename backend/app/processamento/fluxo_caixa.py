"""Serviço de processamento ponta a ponta do Fluxo de Caixa."""

from __future__ import annotations

import logging
import re
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl.utils import get_column_letter

from ..config import settings
from ..consolidacao.merger import FluxoCaixaMerger
from ..contracts.common import (
    ErrorSeverity,
    FlowType,
    ProcessingLog,
    ProcessingStatus,
    ValidationError,
)
from ..contracts.fluxo_caixa import (
    FCLote,
    FCMovimento,
    TipoMovimento,
    eh_transferencia_classificacao,
)
from ..contracts.processamento import DREProcessamentoResponse
from ..exportacao.exporter import Exporter
from ..ingestao.parser import ExcelParser
from ..templates.writer import TemplateWriter
from ..transformacao.engine import FluxoCaixaTransformer
from ..validacao.codigos_gerenciais import extrair_codigo_gerencial
from ..validacao.validators import FluxoCaixaValidator

logger = logging.getLogger(__name__)

_MESES_NOMES = [
    "Janeiro",
    "Fevereiro",
    "Março",
    "Abril",
    "Maio",
    "Junho",
    "Julho",
    "Agosto",
    "Setembro",
    "Outubro",
    "Novembro",
    "Dezembro",
]
_EMPRESA_PADRAO = "A IDEAL"
_SALDO_ANO_ANTERIOR_LABEL = "Saldo do Ano Anterior"
_BANCO_ROTULOS = {
    "cef": "CEF",
    "itau": "ITAU",
    "itau_isolamento": "ITAU Isolamento",
    "mercantil": "MERCANTIL",
    "safra": "SAFRA",
}
_BANCO_TITULOS_SALDO = {
    "cef": "CEF",
    "itau": "Itau",
    "itau_isolamento": "Itau Isolamento",
    "mercantil": "Mercantil",
    "safra": "Safra",
}
_BANCO_ORDEM_CONSOLIDADO = {
    "cef": 0,
    "safra": 1,
    "mercantil": 2,
    "itau_isolamento": 3,
    "itau": 4,
}


class FluxoCaixaProcessamentoService:
    """Orquestra validação, consolidação e escrita do Fluxo de Caixa."""

    def __init__(
        self,
        template_path: Path | None = None,
        output_dir: Path | None = None,
        logs_dir: Path | None = None,
        temp_dir: Path | None = None,
    ):
        self.template_path = Path(template_path) if template_path else settings.template_fluxo_path
        self.exporter = Exporter(
            base_dir=settings.base_dir,
            output_dir=output_dir,
            logs_dir=logs_dir,
            temp_dir=temp_dir,
        )
        self.parser = ExcelParser("fluxo")
        self.validator = FluxoCaixaValidator()
        self.transformer = FluxoCaixaTransformer()
        self.merger = FluxoCaixaMerger()

    def processar_lote(
        self,
        arquivos: list[tuple[Path, str]],
        periodo: str,
    ) -> DREProcessamentoResponse:
        """Processa múltiplos arquivos bancários e gera o Fluxo consolidado."""
        nomes_entrada = [nome for _, nome in arquivos]
        log = self.exporter.criar_log(FlowType.FLUXO_CAIXA, nomes_entrada)
        log.metadata.update(
            {
                "periodo": periodo,
                "template": self.template_path.name,
                "total_arquivos_recebidos": len(arquivos),
            }
        )

        try:
            log.status = ProcessingStatus.VALIDATING
            lotes: list[FCLote] = []
            bancos_identificados: set[str] = set()
            arquivos_ignorados: list[dict[str, str]] = []

            for arquivo_path, arquivo_nome in arquivos:
                dados = self.parser.ler_arquivo(arquivo_path)
                dados["arquivo"] = arquivo_nome

                validacao = self.validator.validar(dados)

                if validacao.erros:
                    if self._deve_ignorar_arquivo_por_estrutura(validacao.erros):
                        arquivos_ignorados.append(
                            {
                                "arquivo": arquivo_nome,
                                "motivo": "arquivo_sem_colunas_de_movimento_bancario",
                            }
                        )
                        log.warnings.extend(validacao.warnings)
                        log.warnings.append(
                            ValidationError(
                                campo="arquivo",
                                mensagem=(
                                    f"Arquivo '{arquivo_nome}' foi ignorado por não possuir "
                                    "estrutura de movimento bancário."
                                ),
                                severidade=ErrorSeverity.WARNING,
                                sugestao=(
                                    "Use relatórios de movimento bancário (.xls/.xlsx) "
                                    "para compor o consolidado."
                                ),
                            )
                        )
                        continue

                    log.erros.extend(validacao.erros)
                    log.warnings.extend(validacao.warnings)
                    continue

                log.warnings.extend(validacao.warnings)
                bancos_identificados.update(validacao.bancos_identificados)
                banco = self.parser.detectar_banco(arquivo_nome) or "desconhecido"
                lote = self.transformer.transformar(dados, banco_origem=banco, periodo=periodo)
                log.warnings.extend(self.transformer.erros)
                lotes.append(lote)
                log.total_registros += lote.total_registros

            if log.tem_bloqueante:
                log.finalizar(ProcessingStatus.ERROR)
                self.exporter.salvar_log(log)
                return self._to_response(log)

            if not lotes:
                log.adicionar_erro(
                    campo="arquivos",
                    mensagem="Nenhum arquivo válido para consolidação do Fluxo de Caixa.",
                    severidade=ErrorSeverity.BLOQUEANTE,
                )
                log.finalizar(ProcessingStatus.ERROR)
                self.exporter.salvar_log(log)
                return self._to_response(log)

            log.status = ProcessingStatus.PROCESSING
            consolidado = self.merger.consolidar(lotes, periodo=periodo)

            if consolidado.total_registros == 0:
                log.adicionar_erro(
                    campo="processamento",
                    mensagem="Nenhum movimento válido encontrado para gerar o Fluxo de Caixa.",
                    severidade=ErrorSeverity.BLOQUEANTE,
                )
                log.finalizar(ProcessingStatus.ERROR)
                self.exporter.salvar_log(log)
                return self._to_response(log)

            output_path = self.exporter.caminho_saida(
                self.exporter.gerar_nome_saida(
                    FlowType.FLUXO_CAIXA,
                    self._periodo_para_nome(periodo),
                )
            )
            periodo_parseado = self._parse_periodo(periodo)
            meses_visiveis = [periodo_parseado[0]] if periodo_parseado else None
            totais_saida = self._escrever_template(
                consolidado,
                output_path,
                meses_visiveis=meses_visiveis,
            )

            log.registros_processados = consolidado.total_registros
            log.total_registros = consolidado.total_registros
            log.metadata.update(
                {
                    "download_url": f"/api/processamentos/{log.id}/download",
                    "output_path": str(output_path),
                    "bancos_identificados": sorted(
                        set(bancos_identificados) | set(consolidado.bancos)
                    ),
                    "total_movimentos": consolidado.total_registros,
                    "total_creditos_movimentos": float(consolidado.total_creditos),
                    "total_debitos_movimentos": float(consolidado.total_debitos),
                    "saldo_liquido_movimentos": float(
                        consolidado.total_creditos - consolidado.total_debitos
                    ),
                    "total_creditos": totais_saida["creditos"],
                    "total_debitos": totais_saida["debitos"],
                    "saldo_liquido": totais_saida["saldo_liquido"],
                    "arquivos_ignorados": arquivos_ignorados,
                }
            )
            log.finalizar(ProcessingStatus.COMPLETED, arquivo_saida=output_path.name)
            self.exporter.salvar_log(log)
            return self._to_response(log)

        except Exception as exc:
            logger.exception("Falha ao processar Fluxo de Caixa: %s", exc)
            if log.fim is None:
                log.adicionar_erro(
                    campo="processamento",
                    mensagem=str(exc),
                    severidade=ErrorSeverity.BLOQUEANTE,
                )
                log.finalizar(ProcessingStatus.ERROR)
                try:
                    self.exporter.salvar_log(log)
                except Exception:
                    logger.exception("Falha ao persistir log de erro do Fluxo")
            raise

    def obter_processamento(self, processamento_id: str) -> DREProcessamentoResponse | None:
        """Carrega o estado de um processamento salvo."""
        log = self.exporter.carregar_log(processamento_id)
        if not log or log.fluxo != FlowType.FLUXO_CAIXA:
            return None
        return self._to_response(log)

    def obter_arquivo_saida(self, processamento_id: str) -> Path | None:
        """Retorna caminho do arquivo final, quando existir."""
        log = self.exporter.carregar_log(processamento_id)
        if not log or not log.arquivo_saida:
            return None

        path = self.exporter.caminho_saida(log.arquivo_saida)
        return path if path.exists() else None

    def _escrever_template(
        self,
        lote: FCLote,
        output_path: Path,
        meses_visiveis: list[int] | None = None,
        preservar_historico: bool = True,
        saldo_ano_anterior: Decimal | None = None,
        saldos_iniciais_por_banco: dict[str, Decimal] | None = None,
    ) -> dict[str, float]:
        with TemplateWriter(self.template_path) as writer:
            ws = writer._wb["Consolidado"]
            limite_fim = ws.max_row

            periodo = self._parse_periodo(lote.periodo)
            classificacoes_template = self._mapear_classificacoes_template(
                writer._wb,
                ws,
                limite_fim,
            )
            linhas_existentes = (
                self._linhas_fora_do_periodo(ws, limite_fim, periodo) if preservar_historico else []
            )
            linhas: list[list] = []

            for linha in linhas_existentes:
                linhas.append(
                    self._normalizar_linha_consolidado(
                        linha,
                        row_number=1 + len(linhas) + 1,
                    )
                )

            if saldo_ano_anterior and saldo_ano_anterior != Decimal("0"):
                linhas.append(
                    self._linha_saldo_ano_anterior(
                        periodo[1] if periodo else date.today().year,
                        1,
                        saldo_ano_anterior,
                        row_number=1 + len(linhas) + 1,
                    )
                )

            for item in self._linhas_movimentos_com_saldos(
                lote.movimentos,
                saldos_iniciais_por_banco=saldos_iniciais_por_banco,
            ):
                row_number = 1 + len(linhas) + 1
                if isinstance(item, FCMovimento):
                    linhas.append(
                        self._converter_movimento_para_linha(
                            item,
                            row_number,
                            classificacoes_template,
                        )
                    )
                else:
                    linhas.append(self._normalizar_linha_consolidado(item, row_number))

            writer.limpar_area("Consolidado", 2, limite_fim, 1, 11)
            writer.escrever_area("Consolidado", linhas, 2, 1)
            linha_final_dados = 1 + len(linhas)
            if not preservar_historico and limite_fim > linha_final_dados:
                ws.delete_rows(linha_final_dados + 1, limite_fim - linha_final_dados)
                writer._modified_sheets.add("Consolidado")

            self._recalcular_aba_apoio_fluxo(writer, linhas)
            if saldo_ano_anterior and saldo_ano_anterior != Decimal("0"):
                self._garantir_saldo_ano_anterior_no_fluxo(writer)
            if meses_visiveis:
                self._aplicar_visibilidade_meses(writer, meses_visiveis)
            self._limpar_marcadores_apresentacao(writer)
            self._proteger_formulas_de_divisao(writer)
            # O relatório usa as linhas do template como fonte visual; slicers
            # legados não fazem parte da saída e podem causar reparos no Excel.
            writer.remover_slicers()
            writer.ajustar_tabela_range(
                sheet_name="Consolidado",
                table_name="FluxoConsol",
                linha_fim=linha_final_dados,
                coluna_fim=11,
            )

            problemas = writer.validar_integridade()
            if problemas:
                raise RuntimeError("; ".join(problemas))

            writer.salvar(output_path)
            if meses_visiveis:
                return self._totais_linhas_meses(linhas, meses_visiveis)
            return self._totais_linhas_periodo(linhas, periodo)

    def _converter_movimento_para_linha(
        self,
        movimento: FCMovimento,
        row_number: int,
        classificacoes_template: dict[str, str] | None = None,
    ) -> list:
        valor_abs = abs(movimento.valor)
        credito = Decimal("0")
        debito = Decimal("0")

        if movimento.eh_transferencia:
            if movimento.transferencia_emitida:
                debito = valor_abs
            else:
                credito = valor_abs
        elif movimento.tipo == TipoMovimento.CREDITO:
            credito = valor_abs
        elif movimento.tipo == TipoMovimento.DEBITO:
            debito = valor_abs
        else:
            if movimento.valor >= 0:
                credito = valor_abs
            else:
                debito = valor_abs

        saldo = movimento.saldo if movimento.saldo is not None else (credito - debito)
        classificacao = self._classificacao_canonica(movimento, classificacoes_template or {})
        descricao = self._formatar_descricao_consolidado(movimento)

        return [
            movimento.data_movimento,  # A - Data
            descricao,  # B - Fornecedor/Histórico
            self._valor_ou_vazio(credito),  # C - Crédito
            self._valor_ou_vazio(debito),  # D - Débito
            float(saldo),  # E - Saldo
            classificacao,  # F - Classificação
            f"=YEAR(A{row_number})",  # G - Ano
            f"=MONTH(A{row_number})",  # H - C.M.
            f"=INDEX(mês[],Consolidado!H{row_number},2)",  # I - Mês
            self._rotulo_banco(movimento.banco_origem),  # J - Banco
            _EMPRESA_PADRAO,  # K - Empresa
        ]

    def _linha_saldo_ano_anterior(
        self,
        ano: int,
        mes: int,
        saldo_ano_anterior: Decimal,
        row_number: int,
    ) -> list:
        return [
            date(ano, mes, 1),
            _SALDO_ANO_ANTERIOR_LABEL,
            float(saldo_ano_anterior),
            None,
            float(saldo_ano_anterior),
            _SALDO_ANO_ANTERIOR_LABEL,
            f"=YEAR(A{row_number})",
            f"=MONTH(A{row_number})",
            f"=INDEX(mês[],Consolidado!H{row_number},2)",
            "MANUAL",
            _EMPRESA_PADRAO,
        ]

    def _classificacao_canonica(
        self,
        movimento: FCMovimento,
        classificacoes_template: dict[str, str],
    ) -> str:
        classificacao = movimento.classificacao or "Sem classificação"
        if classificacao.startswith("Transferência "):
            return classificacao

        for candidata in (movimento.conta_gerencial, movimento.classificacao):
            codigo = extrair_codigo_gerencial(candidata)
            if codigo:
                rotulo_por_codigo = classificacoes_template.get(f"@cod:{codigo}")
                if rotulo_por_codigo:
                    return rotulo_por_codigo

            chave = self._normalizar_chave_classificacao(candidata)
            if chave and chave in classificacoes_template:
                return classificacoes_template[chave]

            sem_codigo = self._remover_codigo_gerencial(candidata)
            chave_sem_codigo = self._normalizar_chave_classificacao(sem_codigo)
            if chave_sem_codigo and chave_sem_codigo in classificacoes_template:
                return classificacoes_template[chave_sem_codigo]

        # Conta ainda não cadastrada no template: o código continua no dado
        # canônico, mas a rubrica visual deve exibir apenas seu nome.
        for candidata in (movimento.conta_gerencial, movimento.classificacao):
            nome = self._remover_codigo_gerencial(candidata)
            if nome and nome != str(candidata or "").strip():
                return nome
        return classificacao

    @staticmethod
    def _valor_ou_vazio(valor: Decimal):
        return None if valor == Decimal("0") else float(valor)

    @staticmethod
    def _formatar_descricao_consolidado(movimento: FCMovimento) -> str:
        descricao = movimento.descricao.strip()
        complemento = (movimento.conta_gerencial or movimento.classificacao or "").strip()
        if not complemento:
            return descricao
        return f"{descricao}-{complemento}"

    @staticmethod
    def _rotulo_banco(banco_origem: str) -> str:
        banco = banco_origem.lower().strip()
        return _BANCO_ROTULOS.get(banco, banco_origem.upper())

    @staticmethod
    def _parse_periodo(periodo: str) -> tuple[int, int] | None:
        match = re.match(r"^\s*(\d{1,2})[/-](\d{4})\s*$", periodo)
        if not match:
            return None
        mes = int(match.group(1))
        ano = int(match.group(2))
        if mes < 1 or mes > 12:
            return None
        return mes, ano

    def _linhas_fora_do_periodo(
        self,
        ws,
        limite_fim: int,
        periodo: tuple[int, int] | None,
    ) -> list[list]:
        linhas = []
        for row in ws.iter_rows(min_row=2, max_row=limite_fim, min_col=1, max_col=11):
            valores = [cell.value for cell in row]
            if not any(valor is not None for valor in valores):
                continue
            if periodo and self._valor_data_no_periodo(valores[0], periodo):
                continue
            linhas.append(valores)
        return linhas

    @staticmethod
    def _valor_data_no_periodo(valor, periodo: tuple[int, int]) -> bool:
        mes, ano = periodo
        if isinstance(valor, datetime):
            return valor.month == mes and valor.year == ano
        if isinstance(valor, date):
            return valor.month == mes and valor.year == ano
        return False

    @staticmethod
    def _normalizar_linha_consolidado(linha: list, row_number: int) -> list:
        valores = (linha + [None] * 11)[:11]
        valores[6] = f"=YEAR(A{row_number})"
        valores[7] = f"=MONTH(A{row_number})"
        valores[8] = f"=INDEX(mês[],Consolidado!H{row_number},2)"
        return valores

    def _recalcular_aba_apoio_fluxo(self, writer: TemplateWriter, linhas: list[list]) -> None:
        """Materializa a base da aba Apoio para todos os 12 meses."""
        wb = writer._wb
        if wb is None or "Apoio" not in wb.sheetnames:
            return

        ws = wb["Apoio"]
        primeira_linha_dados = 6
        ultima_linha = max(ws.max_row, primeira_linha_dados)
        ultima_coluna = 28  # C:Z meses em pares; AA:AB totais.

        agregados: dict[str, dict[int, dict[str, Decimal]]] = {}
        for linha in linhas:
            data = linha[0]
            classificacao = linha[5]
            if (
                not classificacao
                or eh_transferencia_classificacao(classificacao)
                or not getattr(data, "month", None)
            ):
                continue
            mes = int(data.month)
            if mes < 1 or mes > 12:
                continue

            chave = str(classificacao).strip()
            por_mes = agregados.setdefault(chave, {})
            totais = por_mes.setdefault(
                mes,
                {"credito": Decimal("0"), "debito": Decimal("0")},
            )
            totais["credito"] += self._decimal_linha(linha[2])
            totais["debito"] += self._decimal_linha(linha[3])

        linhas_por_classificacao: dict[str, int] = {}
        for row in range(primeira_linha_dados, ultima_linha + 1):
            rotulo = ws.cell(row=row, column=2).value
            if rotulo:
                linhas_por_classificacao[str(rotulo).strip()] = row

        for classificacao in sorted(agregados):
            if classificacao not in linhas_por_classificacao:
                ultima_linha += 1
                linhas_por_classificacao[classificacao] = ultima_linha
                ws.cell(
                    row=ultima_linha,
                    column=1,
                    value=f"=COUNTIF('Fluxo de Caixa '!C:C,B{ultima_linha})",
                )
                ws.cell(row=ultima_linha, column=2, value=classificacao)

        ws.cell(row=1, column=2, value="Ano")
        ws.cell(row=1, column=3, value="(Tudo)")
        ws.cell(row=5, column=2, value="Rótulos de Linha")
        for col in range(3, ultima_coluna + 1):
            for row in range(4, ultima_linha + 1):
                ws.cell(row=row, column=col).value = None

        for mes, nome in enumerate(_MESES_NOMES, start=1):
            credito_col = self._coluna_credito_apoio(mes)
            debito_col = credito_col + 1
            ws.cell(row=4, column=credito_col, value=nome)
            ws.cell(row=5, column=credito_col, value="Soma de Crédito")
            ws.cell(row=5, column=debito_col, value="Soma de Débito")

        ws.cell(row=4, column=27, value="Total Soma de Crédito")
        ws.cell(row=4, column=28, value="Total Soma de Débito")

        for classificacao, por_mes in agregados.items():
            row = linhas_por_classificacao[classificacao]
            total_credito = Decimal("0")
            total_debito = Decimal("0")
            for mes, totais in por_mes.items():
                credito = totais["credito"]
                debito = totais["debito"]
                total_credito += credito
                total_debito += debito
                ws.cell(
                    row=row,
                    column=self._coluna_credito_apoio(mes),
                    value=self._float_ou_vazio(credito),
                )
                ws.cell(
                    row=row,
                    column=self._coluna_credito_apoio(mes) + 1,
                    value=self._float_ou_vazio(debito),
                )
            ws.cell(row=row, column=27, value=self._float_ou_vazio(total_credito))
            ws.cell(row=row, column=28, value=self._float_ou_vazio(total_debito))

        writer._modified_sheets.add("Apoio")

    @staticmethod
    def _coluna_credito_apoio(mes: int) -> int:
        return 3 + ((mes - 1) * 2)

    @staticmethod
    def _float_ou_vazio(valor: Decimal):
        return None if valor == Decimal("0") else float(valor)

    def _garantir_saldo_ano_anterior_no_fluxo(self, writer: TemplateWriter) -> None:
        wb = writer._wb
        if wb is None or "Fluxo de Caixa " not in wb.sheetnames:
            return

        ws = wb["Fluxo de Caixa "]
        linha_grupo = self._encontrar_linha_rotulo(ws, "(+) SALDO INICIAL", coluna=2)
        if not linha_grupo:
            return

        linha_inicio = self._inicio_bloco_saldo_inicial(ws, linha_grupo)
        linha_fim_original = self._fim_bloco_saldo_inicial(ws, linha_grupo, linha_inicio)
        linha_saldo_existente = self._encontrar_linha_rotulo(
            ws,
            _SALDO_ANO_ANTERIOR_LABEL,
            coluna=3,
        )
        linha_saldo = linha_saldo_existente or linha_fim_original + 1
        if linha_saldo_existente is None and not self._linha_vazia(ws, linha_saldo):
            linha_saldo = linha_grupo + 1
        linha_inicio = min(linha_inicio, linha_saldo)
        linha_fim = max(linha_fim_original, linha_saldo)
        linha_saldo_final = self._encontrar_linha_rotulo(ws, "(=) SALDO FINAL", coluna=2)

        ws.cell(row=linha_saldo, column=3, value=_SALDO_ANO_ANTERIOR_LABEL)
        for col in range(4, 16):
            col_letter = get_column_letter(col)
            if col == 4:
                ws.cell(
                    row=linha_saldo,
                    column=col,
                    value=(
                        "=IFERROR("
                        f"INDEX(Apoio!$B:$AB,MATCH($C{linha_saldo},Apoio!$B:$B,0),"
                        f"MATCH('Fluxo de Caixa '!{col_letter}$5,Apoio!$B$4:$AB$4,0))"
                        ",0)"
                    ),
                )
                ws.cell(
                    row=linha_grupo,
                    column=col,
                    value=f"=SUM({col_letter}{linha_inicio}:{col_letter}{linha_fim})",
                )
                continue

            ws.cell(
                row=linha_saldo,
                column=col,
                value="=0",
            )
            if linha_saldo_final:
                coluna_anterior = get_column_letter(col - 1)
                ws.cell(
                    row=linha_grupo,
                    column=col,
                    value=f"={coluna_anterior}{linha_saldo_final}",
                )
            else:
                ws.cell(
                    row=linha_grupo,
                    column=col,
                    value=f"=SUM({col_letter}{linha_inicio}:{col_letter}{linha_fim})",
                )

        ws.cell(row=linha_saldo, column=17, value=f"=SUM(D{linha_saldo}:O{linha_saldo})")
        writer._modified_sheets.add("Fluxo de Caixa ")

    @staticmethod
    def _encontrar_linha_rotulo(ws, rotulo: str, coluna: int) -> int | None:
        chave = FluxoCaixaProcessamentoService._normalizar_chave_classificacao(rotulo)
        for row in range(1, ws.max_row + 1):
            valor = ws.cell(row=row, column=coluna).value
            if FluxoCaixaProcessamentoService._normalizar_chave_classificacao(valor) == chave:
                return row
        return None

    @staticmethod
    def _linha_vazia(ws, linha: int) -> bool:
        return all(
            ws.cell(row=linha, column=coluna).value in (None, "")
            for coluna in range(1, 18)
        )

    @staticmethod
    def _fim_bloco_saldo_inicial(ws, linha_grupo: int, linha_saldo: int) -> int:
        formula = ws.cell(row=linha_grupo, column=4).value
        if isinstance(formula, str):
            match = re.search(r":\$?[A-Z]+\$?(\d+)", formula)
            if match:
                return max(linha_saldo, int(match.group(1)))

        linha_fim = linha_saldo
        for row in range(linha_saldo + 1, min(ws.max_row, linha_saldo + 40) + 1):
            if ws.cell(row=row, column=3).value:
                linha_fim = row
                continue
            break
        return linha_fim

    @staticmethod
    def _inicio_bloco_saldo_inicial(ws, linha_grupo: int) -> int:
        formula = ws.cell(row=linha_grupo, column=4).value
        if isinstance(formula, str):
            match = re.search(r"SUM\(\$?[A-Z]+\$?(\d+):", formula, flags=re.IGNORECASE)
            if match:
                return max(linha_grupo + 1, int(match.group(1)))
        return linha_grupo + 1

    def _aplicar_visibilidade_meses(
        self,
        writer: TemplateWriter,
        meses_visiveis: list[int],
    ) -> None:
        wb = writer._wb
        if wb is None:
            return

        meses = {int(mes) for mes in meses_visiveis if 1 <= int(mes) <= 12}
        if "Fluxo de Caixa " in wb.sheetnames:
            self._ocultar_colunas_mensais(wb["Fluxo de Caixa "], 4, meses)
            writer._modified_sheets.add("Fluxo de Caixa ")
        if "Apresentação GMP" in wb.sheetnames:
            apresentacao = wb["Apresentação GMP"]
            self._ocultar_colunas_mensais(apresentacao, 3, meses)
            apresentacao.column_dimensions["P"].hidden = True
            writer._modified_sheets.add("Apresentação GMP")
        if "Apoio" in wb.sheetnames:
            self._ocultar_pares_mensais_apoio(wb["Apoio"], meses)
            writer._modified_sheets.add("Apoio")

    @staticmethod
    def _ocultar_colunas_mensais(ws, primeira_coluna: int, meses_visiveis: set[int]) -> None:
        for mes in range(1, 13):
            col = get_column_letter(primeira_coluna + mes - 1)
            ws.column_dimensions[col].hidden = mes not in meses_visiveis

    @staticmethod
    def _ocultar_pares_mensais_apoio(ws, meses_visiveis: set[int]) -> None:
        for mes in range(1, 13):
            credito_col = 3 + ((mes - 1) * 2)
            hidden = mes not in meses_visiveis
            ws.column_dimensions[get_column_letter(credito_col)].hidden = hidden
            ws.column_dimensions[get_column_letter(credito_col + 1)].hidden = hidden
        ws.column_dimensions["AA"].hidden = False
        ws.column_dimensions["AB"].hidden = False

    @staticmethod
    def _proteger_formulas_de_divisao(writer: TemplateWriter) -> None:
        wb = writer._wb
        if wb is None:
            return

        for sheet_name in ("Fluxo de Caixa ", "Apresentação GMP"):
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            alterou = False
            for row in ws.iter_rows():
                for cell in row:
                    formula = cell.value
                    if not isinstance(formula, str) or not formula.startswith("="):
                        continue
                    formula_upper = formula.upper()
                    if (
                        "/" not in formula
                        or "IFERROR(" in formula_upper
                        or "SEERRO(" in formula_upper
                    ):
                        continue
                    cell.value = f"=IFERROR({formula[1:]},0)"
                    alterou = True

            if alterou:
                writer._modified_sheets.add(sheet_name)

    @staticmethod
    def _limpar_marcadores_apresentacao(writer: TemplateWriter) -> None:
        wb = writer._wb
        if wb is None or "Apresentação GMP" not in wb.sheetnames:
            return

        ws = wb["Apresentação GMP"]
        alterou = False
        for cell in ws["A"]:
            valor = cell.value
            if isinstance(valor, str) and valor.strip().upper() == "OK":
                cell.value = None
                alterou = True

        if alterou:
            writer._modified_sheets.add("Apresentação GMP")

    def _totais_linhas_periodo(
        self,
        linhas: list[list],
        periodo: tuple[int, int] | None,
    ) -> dict[str, float]:
        creditos = Decimal("0")
        debitos = Decimal("0")
        for linha in linhas:
            if periodo and not self._valor_data_no_periodo(linha[0], periodo):
                continue
            if self._eh_linha_neutra_no_consolidado(linha):
                continue
            creditos += self._decimal_linha(linha[2])
            debitos += self._decimal_linha(linha[3])

        saldo_liquido = creditos - debitos
        return {
            "creditos": float(creditos),
            "debitos": float(debitos),
            "saldo_liquido": float(saldo_liquido),
        }

    def _totais_linhas_meses(self, linhas: list[list], meses: list[int]) -> dict[str, float]:
        meses_set = {int(mes) for mes in meses}
        creditos = Decimal("0")
        debitos = Decimal("0")
        for linha in linhas:
            data = linha[0]
            if not getattr(data, "month", None) or int(data.month) not in meses_set:
                continue
            if self._eh_linha_neutra_no_consolidado(linha):
                continue
            creditos += self._decimal_linha(linha[2])
            debitos += self._decimal_linha(linha[3])

        saldo_liquido = creditos - debitos
        return {
            "creditos": float(creditos),
            "debitos": float(debitos),
            "saldo_liquido": float(saldo_liquido),
        }

    @staticmethod
    def _decimal_linha(valor) -> Decimal:
        if valor in (None, ""):
            return Decimal("0")
        return Decimal(str(valor))

    @staticmethod
    def _eh_linha_neutra_no_consolidado(linha: list) -> bool:
        classificacao = str(linha[5] or "").strip()
        return eh_transferencia_classificacao(classificacao) or classificacao.upper().startswith(
            ("SALDO INICIAL ", "SALDO FINAL ")
        )

    def _mapear_classificacoes_template(self, wb, ws, limite_fim: int) -> dict[str, str]:
        mapa: dict[str, str] = {}
        for rotulo in self._rotulos_visuais_fluxo(wb):
            chave_rotulo = self._normalizar_chave_classificacao(rotulo)
            if chave_rotulo:
                mapa.setdefault(chave_rotulo, rotulo)

        for codigo, rotulo in self._rotulos_template_por_codigo().items():
            if rotulo:
                mapa[f"@cod:{codigo}"] = rotulo

        for row in ws.iter_rows(min_row=2, max_row=limite_fim, min_col=1, max_col=11):
            descricao = row[1].value
            classificacao = row[5].value
            if not classificacao:
                continue

            chave_classificacao = self._normalizar_chave_classificacao(classificacao)
            if chave_classificacao:
                mapa.setdefault(chave_classificacao, classificacao)

            if isinstance(descricao, str) and "-" in descricao:
                chave_descricao = self._normalizar_chave_classificacao(descricao.rsplit("-", 1)[1])
                if chave_descricao:
                    mapa[chave_descricao] = classificacao
        return mapa

    def _rotulos_template_por_codigo(self) -> dict[str, str]:
        saida = self.parser.mapping.get("saida", {})
        rotulos = saida.get("rotulos_template_por_codigo", {})
        if not isinstance(rotulos, dict):
            return {}
        return {str(codigo).strip(): str(rotulo) for codigo, rotulo in rotulos.items() if rotulo}

    @staticmethod
    def _rotulos_visuais_fluxo(wb) -> list[str]:
        if wb is None:
            return []

        rotulos: list[str] = []
        for sheet_name, column, primeira_linha in (
            ("Apoio", 2, 6),
            ("Fluxo de Caixa ", 3, 1),
            ("Naturezas Financeiras", 1, 2),
        ):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for row in range(primeira_linha, ws.max_row + 1):
                valor = ws.cell(row=row, column=column).value
                if isinstance(valor, str) and valor.strip():
                    rotulos.append(valor)
        return rotulos

    @staticmethod
    def _remover_codigo_gerencial(valor) -> str:
        if valor is None:
            return ""
        return re.sub(r"^\s*\d+(?:\.\d+)+\s*[-–—]?\s*", "", str(valor).strip())

    @staticmethod
    def _normalizar_chave_classificacao(valor) -> str:
        if valor is None:
            return ""
        texto = str(valor).upper().strip()
        texto = texto.replace("SAÍDA", "SAIDA")
        texto = re.sub(r"\(\s*100\s*,\s*00\s*%\s*\)\s*;?", "", texto)
        texto = re.sub(r"[^A-Z0-9À-Ü]+", " ", texto)
        return " ".join(texto.split())

    def _linhas_movimentos_com_saldos(
        self,
        movimentos: list[FCMovimento],
        saldos_iniciais_por_banco: dict[str, Decimal] | None = None,
    ) -> list[FCMovimento | list]:
        if not movimentos:
            return []

        por_banco_mes: dict[tuple[str, int, int], list[FCMovimento]] = {}
        for movimento in movimentos:
            banco = movimento.banco_origem.lower().strip()
            chave = (banco, movimento.data_movimento.year, movimento.data_movimento.month)
            por_banco_mes.setdefault(chave, []).append(movimento)

        saldos_anteriores = {
            str(banco).lower().strip(): Decimal(str(saldo))
            for banco, saldo in (saldos_iniciais_por_banco or {}).items()
        }
        saida: list[FCMovimento | list] = []
        for banco, _ano, _mes in sorted(
            por_banco_mes,
            key=lambda item: (_BANCO_ORDEM_CONSOLIDADO.get(item[0], 99), item),
        ):
            grupo = sorted(
                por_banco_mes[(banco, _ano, _mes)],
                key=lambda mov: (
                    mov.data_movimento,
                    mov.linha_origem or 0,
                    mov.descricao,
                ),
            )
            incluir_saldos = banco in saldos_anteriores or any(
                mov.saldo is not None for mov in grupo
            )
            if incluir_saldos:
                saldo_corrente = saldos_anteriores.get(banco)
                if saldo_corrente is None:
                    saldo_corrente = self._saldo_inicial_grupo(grupo)
                saida.append(self._linha_saldo_inicial(grupo[0], saldo_corrente))
                for movimento in grupo:
                    saldo_corrente = self._aplicar_movimento_no_saldo(saldo_corrente, movimento)
                    saida.append(movimento.model_copy(update={"saldo": saldo_corrente}))
                saldo_final = self._saldo_final_grupo(grupo, saldo_corrente)
                saida.append(self._linha_saldo_final(grupo[-1], saldo_final))
                saldos_anteriores[banco] = saldo_final
                continue

            saida.extend(grupo)
        return saida

    @staticmethod
    def _saldo_inicial_grupo(grupo: list[FCMovimento]) -> Decimal:
        primeiro_com_saldo = next((mov for mov in grupo if mov.saldo is not None), None)
        if primeiro_com_saldo is None:
            return Decimal("0")
        return (primeiro_com_saldo.saldo or Decimal("0")) - (
            FluxoCaixaProcessamentoService._impacto_bancario(primeiro_com_saldo)
        )

    @staticmethod
    def _saldo_final_grupo(grupo: list[FCMovimento], saldo_calculado: Decimal) -> Decimal:
        candidatos = [mov for mov in grupo if mov.saldo is not None]
        if not candidatos:
            return saldo_calculado
        movimento_final = candidatos[-1]
        return movimento_final.saldo if movimento_final.saldo is not None else saldo_calculado

    @staticmethod
    def _aplicar_movimento_no_saldo(saldo: Decimal, movimento: FCMovimento) -> Decimal:
        return saldo + FluxoCaixaProcessamentoService._impacto_bancario(movimento)

    @staticmethod
    def _impacto_bancario(movimento: FCMovimento) -> Decimal:
        valor_abs = abs(movimento.valor)
        if movimento.eh_transferencia:
            return -valor_abs if movimento.transferencia_emitida else valor_abs
        if movimento.tipo == TipoMovimento.CREDITO:
            return valor_abs
        if movimento.tipo == TipoMovimento.DEBITO:
            return -valor_abs
        return movimento.valor

    def _linha_saldo_inicial(self, movimento: FCMovimento, saldo_inicial: Decimal) -> list:
        banco = movimento.banco_origem.lower().strip()
        titulo = _BANCO_TITULOS_SALDO.get(banco, self._rotulo_banco(movimento.banco_origem))
        return [
            movimento.data_movimento,
            "saldo inicial ",
            # O template materializa as linhas "Saldo Inicial <Banco>" a partir
            # da coluna de crédito da Apoio. O saldo também fica na coluna E do
            # Consolidado para rastreabilidade do extrato.
            float(saldo_inicial) if saldo_inicial != Decimal("0") else None,
            None,
            float(saldo_inicial),
            f"Saldo Inicial {titulo}",
            None,
            None,
            None,
            self._rotulo_banco(movimento.banco_origem),
            _EMPRESA_PADRAO,
        ]

    def _linha_saldo_final(self, movimento: FCMovimento, saldo_final: Decimal) -> list:
        banco = movimento.banco_origem.lower().strip()
        titulo = _BANCO_TITULOS_SALDO.get(banco, self._rotulo_banco(movimento.banco_origem))
        return [
            movimento.data_movimento,
            "saldo ",
            None,
            float(saldo_final) if saldo_final != Decimal("0") else None,
            0,
            f"Saldo Final {titulo}",
            None,
            None,
            None,
            self._rotulo_banco(movimento.banco_origem),
            _EMPRESA_PADRAO,
        ]

    def _periodo_para_nome(self, periodo: str) -> str:
        return periodo.replace("/", "-").replace("\\", "-").strip()

    def _deve_ignorar_arquivo_por_estrutura(self, erros: list[ValidationError]) -> bool:
        """Permite ignorar arquivos não bancários misturados ao lote de fluxo."""
        if not erros:
            return False

        campos_obrigatorios = {"data_movimento", "descricao", "valor"}
        for erro in erros:
            if erro.severidade != ErrorSeverity.BLOQUEANTE:
                return False
            if erro.campo not in campos_obrigatorios:
                return False
            if "coluna obrigatória" not in erro.mensagem.lower():
                return False
        return True

    def _to_response(self, log: ProcessingLog) -> DREProcessamentoResponse:
        download_url = None
        if log.status == ProcessingStatus.COMPLETED and log.arquivo_saida:
            download_url = f"/api/processamentos/{log.id}/download"
        elif log.metadata.get("download_url"):
            download_url = str(log.metadata["download_url"])

        return DREProcessamentoResponse(
            id=log.id,
            fluxo=log.fluxo,
            status=log.status,
            valido=not log.tem_bloqueante and log.status == ProcessingStatus.COMPLETED,
            arquivo_entrada=log.arquivo_entrada,
            arquivo_saida=log.arquivo_saida,
            download_url=download_url,
            total_registros=log.total_registros,
            registros_processados=log.registros_processados,
            erros=log.erros,
            warnings=log.warnings,
            inicio=log.inicio,
            fim=log.fim,
            metadata=log.metadata,
        )
