import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.config import settings
from app.contracts.common import ProcessingStatus
from app.processamento.dre import DREProcessamentoService


def _zip_read(path: Path, part: str) -> bytes:
    with zipfile.ZipFile(path) as zf:
        return zf.read(part)


def _criar_arquivo_dre_cumulativo(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "RELATORIO"
    ws.append(["metadata"])
    ws.append(
        [
            "Emissão",
            "Descri.",
            "Vlr.bruto (R$)",
            "CLASSIFICAÇÃO",
            "Obra/Centro custo",
        ]
    )
    for mes in range(1, 6):
        ws.append(
            [
                f"01/{mes:02d}/2025",
                f"Lancamento {mes}",
                float(100 * mes),
                "1 - ENTRADA" if mes % 2 else "2 - SAIDA",
                "VLI PINTURA",
            ]
        )
    wb.save(path)


def test_processamento_dre_gera_arquivo_e_log(tmp_path):
    service = DREProcessamentoService(
        output_dir=tmp_path / "output",
        logs_dir=tmp_path / "logs",
        temp_dir=tmp_path / "tmp",
    )

    arquivo = tmp_path / "DRE_CUMULATIVO_01_A_05_2025.xlsx"
    _criar_arquivo_dre_cumulativo(arquivo)
    resultado = service.processar(arquivo, arquivo.name, "05/2025")

    assert resultado.valido is True
    assert resultado.status == ProcessingStatus.COMPLETED
    assert resultado.download_url == f"/api/processamentos/{resultado.id}/download"
    assert resultado.total_registros > 0
    assert resultado.registros_processados == resultado.total_registros
    assert resultado.arquivo_saida is not None
    assert resultado.metadata["bd_fluxo_range_fisico"] == "A1:R4964"
    assert resultado.metadata["bd_fluxo_total_linhas_template"] == 4963
    assert resultado.metadata["bd_fluxo_registros_reais"] == resultado.total_registros
    assert resultado.metadata["bd_fluxo_linhas_entrada_reescritas"] == resultado.total_registros
    assert resultado.metadata["bd_fluxo_limpeza_faixa_aplicada"] is True
    assert resultado.metadata["bd_fluxo_cabecalho_linha"] == 1
    assert resultado.metadata["bd_fluxo_ultima_linha_dados_reais"] == (
        resultado.total_registros + 1
    )
    assert resultado.metadata["bd_fluxo_linhas_sem_lancamento_inicio"] == (
        resultado.total_registros + 2
    )
    assert resultado.metadata["bd_fluxo_linhas_sem_lancamento_fim"] == 4964
    assert resultado.metadata["bd_fluxo_linhas_sem_lancamento_faixa"] == (
        f"{resultado.total_registros + 2}:4964"
    )
    assert resultado.metadata["bd_fluxo_linhas_sem_lancamento_ocultadas"] is False
    assert resultado.metadata["dre_periodo_modo_cumulativo"] is True
    assert resultado.metadata["dre_periodo_competencia"] == "05/2025"
    assert resultado.metadata["dre_periodo_meses_encontrados_ano_competencia"] == [1, 2, 3, 4, 5]
    assert resultado.metadata["dre_periodo_meses_faltantes_ano_competencia"] == []

    output_file = tmp_path / "output" / resultado.arquivo_saida
    assert output_file.exists()

    logs = list((tmp_path / "logs").glob(f"log_{resultado.id}_*.json"))
    assert len(logs) == 1

    wb = load_workbook(output_file, data_only=False)
    assert "BD_FLUXO" in wb.sheetnames
    ws = wb["BD_FLUXO"]
    assert ws["A2"].value is not None
    assert ws[f"A{resultado.total_registros + 1}"].value is not None
    assert ws[f"A{resultado.total_registros + 2}"].value is None
    assert "BD_FLUXO1" in ws.tables.keys()
    assert ws.row_dimensions[resultado.total_registros + 1].hidden in (False, None)
    assert ws.row_dimensions[resultado.total_registros + 2].hidden in (False, None)

    template_wb = load_workbook(settings.template_dre_path, data_only=False)
    assert ws["H2"].value == template_wb["BD_FLUXO"]["H2"].value

    with zipfile.ZipFile(output_file) as zf:
        names = set(zf.namelist())
        assert any(name.startswith("xl/slicers/") for name in names)
        assert any(name.startswith("xl/slicerCaches/") for name in names)

        sheet1_xml = zf.read("xl/worksheets/sheet1.xml").decode("utf-8", errors="ignore")
        assert 'dimension ref="A1:W50"' in sheet1_xml
        assert "extLst" in sheet1_xml

        sheet3_rels = zf.read("xl/worksheets/_rels/sheet3.xml.rels").decode(
            "utf-8", errors="ignore"
        )
        assert "printerSettings" in sheet3_rels
        assert "relationships/table" in sheet3_rels

        workbook_xml = zf.read("xl/workbook.xml").decode("utf-8", errors="ignore")
        assert "fullCalcOnLoad" in workbook_xml
        assert "forceFullCalc" in workbook_xml
        assert "mc:Ignorable" in workbook_xml
        assert "xmlns:x15=" in workbook_xml

    template_file = settings.template_dre_path
    unchanged_parts = [
        "xl/worksheets/sheet1.xml",
        "xl/worksheets/sheet2.xml",
        "xl/worksheets/sheet4.xml",
        "xl/worksheets/sheet5.xml",
        "xl/worksheets/_rels/sheet1.xml.rels",
        "xl/worksheets/_rels/sheet2.xml.rels",
        "xl/worksheets/_rels/sheet4.xml.rels",
        "xl/worksheets/_rels/sheet5.xml.rels",
    ]
    for part in unchanged_parts:
        assert _zip_read(output_file, part) == _zip_read(template_file, part), part

    lookup = service.obter_processamento(resultado.id)
    assert lookup is not None
    assert lookup.id == resultado.id
    assert lookup.download_url == resultado.download_url


def test_processamento_dre_bloqueia_arquivo_nao_cumulativo(tmp_path):
    service = DREProcessamentoService(
        output_dir=tmp_path / "output",
        logs_dir=tmp_path / "logs",
        temp_dir=tmp_path / "tmp",
    )

    arquivo = settings.base_dir / "RELATORIO DRE MES 05.xls"
    resultado = service.processar(arquivo, arquivo.name, "05/2025")

    assert resultado.valido is False
    assert resultado.status == ProcessingStatus.ERROR
    assert any(e.campo == "competencia" for e in resultado.erros)


def test_processamento_dre_permite_modo_nao_cumulativo_para_teste(tmp_path):
    service = DREProcessamentoService(
        output_dir=tmp_path / "output",
        logs_dir=tmp_path / "logs",
        temp_dir=tmp_path / "tmp",
    )

    arquivo = settings.base_dir / "RELATORIO DRE MES 05.xls"
    resultado = service.processar(
        arquivo,
        arquivo.name,
        "05/2025",
        modo_cumulativo=False,
    )

    assert resultado.valido is True
    assert resultado.status == ProcessingStatus.COMPLETED
    assert resultado.metadata["dre_periodo_modo_cumulativo"] is False
