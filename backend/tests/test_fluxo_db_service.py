"""Testes de persistência e geração do Fluxo de Caixa a partir do banco."""

import tempfile
from datetime import date
from decimal import Decimal
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.contracts.persistence import FluxoMovimentoDB, FluxoUpload
from app.db.connection import DatabaseConnection
from app.db.manager import MigrationManager
from app.ingestao.fluxo_caixa_ingestao import FluxoCaixaIngestaoService
from app.processamento.fluxo_caixa_db import FluxoCaixaGeracaoService
from app.repository.fluxo_repository import FluxoCaixaRepository

HEADERS = [
    "Data",
    "Fornecedor",
    "Crédito",
    "Débito",
    "Saldo",
    "Classificação",
    "Ano",
    "C.M.",
    "Mês",
    "Banco",
    "Empresa",
]


def _novo_db() -> DatabaseConnection:
    tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
    tmp.close()
    db = DatabaseConnection(tmp.name)
    MigrationManager(db).migrate()
    return db


def _criar_arquivo_fluxo(path: Path, mes: int = 8) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.append(["Relatório de movimentos financeiros"])
    ws.append(["Conta:", "BANCO ITAU"])
    ws.append([])
    ws.append(
        [
            "Data Mov.",
            "Tipo",
            "Desc. Mov.",
            "Valor (R$)",
            "Saldo (R$)",
            "Conta Gerencial Mov",
        ]
    )
    ws.append(
        [
            f"04/{mes:02d}/2025",
            "Crédito",
            "Recebimento Cliente",
            1000.0,
            1000.0,
            "Recebimento de Clientes",
        ]
    )
    ws.append(
        [
            f"05/{mes:02d}/2025",
            "Débito",
            "Pagamento Fornecedor",
            200.0,
            800.0,
            "Fornecedores",
        ]
    )
    wb.save(path)


def _criar_template_fluxo(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado"
    ws.append(HEADERS)
    ws.append(
        [
            date(2025, 7, 1),
            "linha antiga de julho",
            1,
            None,
            101,
            "Antiga",
            "=YEAR(A2)",
            "=MONTH(A2)",
            "=INDEX(mês[],Consolidado!H2,2)",
            "ITAU",
            "A IDEAL",
        ]
    )
    table = Table(displayName="FluxoConsol", ref="A1:K2")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)

    apoio = wb.create_sheet("Apoio")
    apoio["B5"] = "Rótulos de Linha"
    apoio["B6"] = "Recebimento de Clientes"
    apoio["B7"] = "Fornecedores"

    fluxo = wb.create_sheet("Fluxo de Caixa ")
    for idx, label in enumerate(
        [
            "JANEIRO",
            "FEVEREIRO",
            "MARÇO",
            "ABRIL",
            "MAIO",
            "JUNHO",
            "JULHO",
            "AGOSTO",
            "SETEMBRO",
            "OUTUBRO",
            "NOVEMBRO",
            "DEZEMBRO",
        ],
        start=4,
    ):
        fluxo.cell(row=5, column=idx, value=label)
    fluxo["B7"] = "(+) SALDO INICIAL"
    fluxo["D7"] = "=SUM(D9:D18)"
    fluxo["Q7"] = "=D7"
    fluxo["C8"] = "Saldo Inicial Aplicações"
    fluxo["D8"] = (
        "=IFERROR(INDEX(Apoio!$B:$AB,MATCH($C8,Apoio!$B$1:$B$176,0),"
        "MATCH('Fluxo de Caixa '!D$5,Apoio!$B$4:$AB$4,0)),0)"
    )
    fluxo["Q8"] = "=D8"

    apresentacao = wb.create_sheet("Apresentação GMP")
    for idx, label in enumerate(
        ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"],
        start=3,
    ):
        apresentacao.cell(row=2, column=idx, value=label)

    wb.save(path)


def test_fluxo_salva_mes_no_banco_e_gera_somente_mes_selecionado(tmp_path):
    db = _novo_db()
    arquivo = tmp_path / "RELATORIO DE MOVIMENTO ITAU SISTEMA.xlsx"
    template = tmp_path / "template_fluxo.xlsx"
    _criar_arquivo_fluxo(arquivo, mes=8)
    _criar_template_fluxo(template)

    ingestao = FluxoCaixaIngestaoService(db)
    resultado_ingestao = ingestao.ingestar_lote(
        arquivos=[(arquivo, arquivo.name)],
        competencia="08/2025",
        replace=True,
    )

    assert resultado_ingestao["success"] is True
    assert resultado_ingestao["status"] == "completed"
    assert resultado_ingestao["inseridos"] == 2
    assert resultado_ingestao["meses_disponiveis_ano"] == [8]

    geracao = FluxoCaixaGeracaoService(
        db=db,
        template_path=template,
        output_dir=tmp_path / "output",
        logs_dir=tmp_path / "logs",
        temp_dir=tmp_path / "tmp",
    )
    resultado_geracao = geracao.gerar_arquivo(
        competencia="08/2025",
        meses_incluir=[8],
    )

    assert resultado_geracao["fonte_dados"] == "db"
    assert resultado_geracao["meses_utilizados"] == [8]
    assert resultado_geracao["meses_ocultos"] == [1, 2, 3, 4, 5, 6, 7, 9, 10, 11, 12]

    wb = load_workbook(resultado_geracao["output_path"], data_only=False)
    consolidado = wb["Consolidado"]
    meses_consolidado = {
        cell.value.month for cell in consolidado["A"][1:] if getattr(cell.value, "month", None)
    }
    assert meses_consolidado == {8}
    assert wb["Apresentação GMP"].column_dimensions["J"].hidden is False
    assert wb["Apresentação GMP"].column_dimensions["I"].hidden is True
    assert wb["Apresentação GMP"].column_dimensions["P"].hidden is True


def test_fluxo_geracao_inclui_saldo_ano_anterior_manual_no_documento(tmp_path):
    db = _novo_db()
    template = tmp_path / "template_fluxo.xlsx"
    _criar_template_fluxo(template)

    wb_template = load_workbook(template)
    fluxo_template = wb_template["Fluxo de Caixa "]
    fluxo_template["B227"] = "(=) SALDO FINAL"
    fluxo_template["D227"] = "=SUM(D7)"
    fluxo_template["E8"] = "=D228"
    wb_template.save(template)

    with db.transaction() as conn:
        conn.execute(
            """
            INSERT INTO fluxo_uploads
            (id, created_at, arquivo_nome, arquivo_sha256, competencia_ano, competencia_mes,
             banco, status, total_linhas, linhas_validas, linhas_rejeitadas, observacao)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "fc-saldo-anterior",
                "2026-04-01T11:00:00",
                "fluxo.xlsx",
                "hash-fc-saldo-anterior",
                2025,
                1,
                "itau",
                "completed",
                1,
                1,
                0,
                None,
            ),
        )
        conn.execute(
            """
            INSERT INTO fluxo_movimentos
            (upload_id, competencia_ano, competencia_mes, data_movimento, tipo, descricao,
             valor, saldo, classificacao, conta_gerencial, banco_origem, arquivo_origem,
             linha_origem, aba_origem, hash_linha, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "fc-saldo-anterior",
                2025,
                1,
                "2025-01-04",
                "credito",
                "Recebimento",
                1000,
                2234.56,
                "Recebimento de Clientes",
                "Recebimento de Clientes",
                "itau",
                "fluxo.xlsx",
                1,
                "Sheet",
                "saldo-anterior-h1",
                "2026-04-01T11:00:00",
            ),
        )
        conn.execute(
            """
            INSERT INTO fluxo_indicadores_manuais
            (competencia_ano, saldo_ano_anterior, created_at, updated_at)
            VALUES (?, ?, ?, ?)
            """,
            (2025, 1234.56, "2026-04-01T12:00:00", "2026-04-01T12:00:00"),
        )

    geracao = FluxoCaixaGeracaoService(
        db=db,
        template_path=template,
        output_dir=tmp_path / "output",
        logs_dir=tmp_path / "logs",
        temp_dir=tmp_path / "tmp",
    )

    resultado_geracao = geracao.gerar_arquivo(
        competencia="01/2025",
        meses_incluir=[1],
    )

    assert resultado_geracao["saldo_ano_anterior"] == 1234.56

    wb = load_workbook(resultado_geracao["output_path"], data_only=False)
    consolidado = wb["Consolidado"]
    linhas_saldo = [
        row for row in consolidado.iter_rows(min_row=2, values_only=True)
        if row[5] == "Saldo do Ano Anterior"
    ]
    assert len(linhas_saldo) == 1
    assert linhas_saldo[0][0].date() == date(2025, 1, 1)
    assert linhas_saldo[0][2] == 1234.56
    assert all(
        row[5] != "Saldo Inicial Itau"
        for row in consolidado.iter_rows(min_row=2, values_only=True)
    )
    linhas_saldo_final = [
        row for row in consolidado.iter_rows(min_row=2, values_only=True)
        if row[5] == "Saldo Final Itau"
    ]
    assert len(linhas_saldo_final) == 1
    assert linhas_saldo_final[0][3] == 2234.56

    apoio = wb["Apoio"]
    apoio_labels = [apoio.cell(row=row, column=2).value for row in range(6, apoio.max_row + 1)]
    assert "Saldo do Ano Anterior" in apoio_labels
    assert "Saldo Inicial Itau" not in apoio_labels
    assert "Saldo Final Itau" in apoio_labels

    fluxo = wb["Fluxo de Caixa "]
    assert fluxo["C8"].value == "Saldo Inicial Aplicações"
    assert fluxo["E8"].value == "=D228"
    assert fluxo["C19"].value == "Saldo do Ano Anterior"
    assert "MATCH($C19" in fluxo["D19"].value
    assert fluxo["D7"].value == "=SUM(D9:D19)"
    assert fluxo["E7"].value == "=D227"
    assert fluxo["K7"].value == "=J227"
    assert fluxo["E19"].value == "=0"


def test_fluxo_geracao_fevereiro_usa_fechamento_de_janeiro_sem_saldo_ano_anterior(tmp_path):
    db = _novo_db()
    template = tmp_path / "template_fluxo.xlsx"
    _criar_template_fluxo(template)

    with db.transaction() as conn:
        conn.execute(
            """
            INSERT INTO fluxo_uploads
            (id, created_at, arquivo_nome, arquivo_sha256, competencia_ano, competencia_mes,
             banco, status, total_linhas, linhas_validas, linhas_rejeitadas, observacao)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "fc-fev-usa-fechamento",
                "2026-02-01T11:00:00",
                "fluxo.xlsx",
                "hash-fc-fev-usa-fechamento",
                2026,
                1,
                "itau",
                "completed",
                1,
                1,
                0,
                None,
            ),
        )
        conn.execute(
            """
            INSERT INTO fluxo_uploads
            (id, created_at, arquivo_nome, arquivo_sha256, competencia_ano, competencia_mes,
             banco, status, total_linhas, linhas_validas, linhas_rejeitadas, observacao)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "fc-fevereiro",
                "2026-02-02T11:00:00",
                "fluxo_fevereiro.xlsx",
                "hash-fc-fevereiro",
                2026,
                2,
                "itau",
                "completed",
                1,
                1,
                0,
                None,
            ),
        )
        conn.executemany(
            """
            INSERT INTO fluxo_movimentos
            (upload_id, competencia_ano, competencia_mes, data_movimento, tipo, descricao,
             valor, saldo, classificacao, conta_gerencial, banco_origem, arquivo_origem,
             linha_origem, aba_origem, hash_linha, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    "fc-fev-usa-fechamento",
                    2026,
                    1,
                    "2026-01-31",
                    "credito",
                    "Fechamento janeiro",
                    150,
                    150,
                    "Recebimento de Clientes",
                    "Recebimento de Clientes",
                    "itau",
                    "fluxo.xlsx",
                    1,
                    "Sheet",
                    "fechamento-janeiro-h1",
                    "2026-02-01T11:00:00",
                ),
                (
                    "fc-fevereiro",
                    2026,
                    2,
                    "2026-02-05",
                    "credito",
                    "Recebimento fevereiro",
                    20,
                    170,
                    "Recebimento de Clientes",
                    "Recebimento de Clientes",
                    "itau",
                    "fluxo_fevereiro.xlsx",
                    1,
                    "Sheet",
                    "fevereiro-h1",
                    "2026-02-02T11:00:00",
                ),
            ],
        )
        conn.execute(
            """
            INSERT INTO fluxo_indicadores_manuais
            (competencia_ano, saldo_ano_anterior, created_at, updated_at)
            VALUES (?, ?, ?, ?)
            """,
            (2026, 9999, "2026-01-01T12:00:00", "2026-01-01T12:00:00"),
        )

    geracao = FluxoCaixaGeracaoService(
        db=db,
        template_path=template,
        output_dir=tmp_path / "output",
        logs_dir=tmp_path / "logs",
        temp_dir=tmp_path / "tmp",
    )

    resultado_geracao = geracao.gerar_arquivo(
        competencia="02/2026",
        meses_incluir=[2],
    )

    assert resultado_geracao["saldo_ano_anterior"] == 0

    wb = load_workbook(resultado_geracao["output_path"], data_only=False)
    consolidado = wb["Consolidado"]
    linhas = list(consolidado.iter_rows(min_row=2, values_only=True))

    assert all(row[5] != "Saldo do Ano Anterior" for row in linhas)
    linha_saldo_inicial = next(row for row in linhas if row[5] == "Saldo Inicial Itau")
    assert linha_saldo_inicial[0].date() == date(2026, 2, 5)
    assert linha_saldo_inicial[2] == 150.0
    assert linha_saldo_inicial[4] == 150.0


def test_fluxo_geracao_nao_pula_mes_para_resolver_saldo_inicial(tmp_path):
    db = _novo_db()
    template = tmp_path / "template_fluxo.xlsx"
    _criar_template_fluxo(template)

    with db.transaction() as conn:
        for upload_id, mes in (
            ("fc-janeiro-sem-fevereiro", 1),
            ("fc-marco-sem-fevereiro", 3),
        ):
            conn.execute(
                """
                INSERT INTO fluxo_uploads
                (id, created_at, arquivo_nome, arquivo_sha256, competencia_ano, competencia_mes,
                 banco, status, total_linhas, linhas_validas, linhas_rejeitadas, observacao)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    upload_id,
                    "2026-03-01T11:00:00",
                    f"fluxo_{mes}.xlsx",
                    f"hash-{upload_id}",
                    2026,
                    mes,
                    "itau",
                    "completed",
                    1,
                    1,
                    0,
                    None,
                ),
            )
        conn.executemany(
            """
            INSERT INTO fluxo_movimentos
            (upload_id, competencia_ano, competencia_mes, data_movimento, tipo, descricao,
             valor, saldo, classificacao, conta_gerencial, banco_origem, arquivo_origem,
             linha_origem, aba_origem, hash_linha, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    "fc-janeiro-sem-fevereiro",
                    2026,
                    1,
                    "2026-01-31",
                    "credito",
                    "Fechamento janeiro",
                    100,
                    100,
                    "Recebimento de Clientes",
                    "Recebimento de Clientes",
                    "itau",
                    "fluxo_1.xlsx",
                    1,
                    "Sheet",
                    "janeiro-sem-fevereiro-h1",
                    "2026-03-01T11:00:00",
                ),
                (
                    "fc-marco-sem-fevereiro",
                    2026,
                    3,
                    "2026-03-05",
                    "credito",
                    "Recebimento março",
                    10,
                    210,
                    "Recebimento de Clientes",
                    "Recebimento de Clientes",
                    "itau",
                    "fluxo_3.xlsx",
                    1,
                    "Sheet",
                    "marco-sem-fevereiro-h1",
                    "2026-03-01T11:00:00",
                ),
            ],
        )

    geracao = FluxoCaixaGeracaoService(
        db=db,
        template_path=template,
        output_dir=tmp_path / "output",
        logs_dir=tmp_path / "logs",
        temp_dir=tmp_path / "tmp",
    )

    resultado_geracao = geracao.gerar_arquivo(
        competencia="03/2026",
        meses_incluir=[3],
    )

    wb = load_workbook(resultado_geracao["output_path"], data_only=False)
    consolidado = wb["Consolidado"]
    linhas = list(consolidado.iter_rows(min_row=2, values_only=True))
    linha_saldo_inicial = next(row for row in linhas if row[5] == "Saldo Inicial Itau")

    assert linha_saldo_inicial[0].date() == date(2026, 3, 5)
    assert linha_saldo_inicial[2] == 200.0
    assert linha_saldo_inicial[4] == 200.0


def test_fluxo_geracao_janeiro_permite_saldo_ano_anterior_sem_movimentos(tmp_path):
    db = _novo_db()
    template = tmp_path / "template_fluxo.xlsx"
    _criar_template_fluxo(template)

    wb_template = load_workbook(template)
    fluxo_template = wb_template["Fluxo de Caixa "]
    fluxo_template["B227"] = "(=) SALDO FINAL"
    fluxo_template["D227"] = "=SUM(D7)"
    wb_template.save(template)

    with db.transaction() as conn:
        conn.execute(
            """
            INSERT INTO fluxo_indicadores_manuais
            (competencia_ano, saldo_ano_anterior, created_at, updated_at)
            VALUES (?, ?, ?, ?)
            """,
            (2026, 9876.54, "2026-01-01T08:00:00", "2026-01-01T08:00:00"),
        )

    geracao = FluxoCaixaGeracaoService(
        db=db,
        template_path=template,
        output_dir=tmp_path / "output",
        logs_dir=tmp_path / "logs",
        temp_dir=tmp_path / "tmp",
    )

    verificacao = geracao.verificar_dados("01/2026")
    assert verificacao["valido"] is True
    assert verificacao["meses_utilizados"] == [1]

    resultado_geracao = geracao.gerar_arquivo("01/2026")

    assert resultado_geracao["total_movimentos"] == 0
    assert resultado_geracao["saldo_ano_anterior"] == 9876.54
    assert resultado_geracao["saldo_liquido_apresentacao"] == 9876.54

    wb = load_workbook(resultado_geracao["output_path"], data_only=False)
    consolidado = wb["Consolidado"]
    linhas_saldo = [
        row for row in consolidado.iter_rows(min_row=2, values_only=True)
        if row[5] == "Saldo do Ano Anterior"
    ]
    assert len(linhas_saldo) == 1
    assert linhas_saldo[0][0].date() == date(2026, 1, 1)
    assert linhas_saldo[0][2] == 9876.54

    fluxo = wb["Fluxo de Caixa "]
    assert fluxo["C19"].value == "Saldo do Ano Anterior"
    assert "MATCH($C19" in fluxo["D19"].value
    assert fluxo["D7"].value == "=SUM(D9:D19)"
    assert fluxo.column_dimensions["D"].hidden is False
    assert fluxo.column_dimensions["E"].hidden is True


def test_fluxo_atualiza_nome_gerencial_por_codigo_em_movimentos_passados():
    db = _novo_db()
    repo = FluxoCaixaRepository(db)

    upload_antigo = FluxoUpload(
        id=str(uuid4()),
        arquivo_nome="fluxo_05.xlsx",
        arquivo_sha256="hash_fluxo_nome_antigo",
        competencia_ano=2025,
        competencia_mes=5,
        banco="itau",
        status="completed",
    )
    movimento_antigo = FluxoMovimentoDB(
        upload_id=upload_antigo.id,
        competencia_ano=2025,
        competencia_mes=5,
        data_movimento="2025-05-10",
        tipo="debito",
        descricao="Conta antiga",
        valor=Decimal("100.00"),
        conta_gerencial="11.2 - AGUA ADM (100,00%);",
        banco_origem="itau",
        hash_linha="hash_fluxo_antigo",
    )
    repo.upsert_competencia([(upload_antigo, [movimento_antigo])])

    upload_novo = FluxoUpload(
        id=str(uuid4()),
        arquivo_nome="fluxo_06.xlsx",
        arquivo_sha256="hash_fluxo_nome_novo",
        competencia_ano=2025,
        competencia_mes=6,
        banco="itau",
    )
    movimento_novo = FluxoMovimentoDB(
        upload_id=upload_novo.id,
        competencia_ano=2025,
        competencia_mes=6,
        data_movimento="2025-06-10",
        tipo="debito",
        descricao="Conta nova",
        valor=Decimal("200.00"),
        conta_gerencial="11.2 - AGUA ADMINISTRATIVA (100,00%);",
        banco_origem="itau",
        hash_linha="hash_fluxo_novo",
    )
    repo.upsert_competencia([(upload_novo, [movimento_novo])])

    movimentos = repo.movimentos.get_by_meses(2025, [5, 6])
    assert [mov.conta_gerencial for mov in movimentos] == [
        "11.2 - AGUA ADMINISTRATIVA (100,00%);",
        "11.2 - AGUA ADMINISTRATIVA (100,00%);",
    ]
