"""Testes da detecção automática de competência por upload."""

from pathlib import Path

from conftest import login_admin
from fastapi.testclient import TestClient
from openpyxl import Workbook

import app.main as main_module


def _criar_dre_cumulativo(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "RELATORIO"
    ws.append(["metadata"])
    ws.append(["Emissão", "Descri.", "Vlr.bruto (R$)", "CLASSIFICAÇÃO"])
    for mes in range(1, 6):
        ws.append(
            [
                f"10/{mes:02d}/2025",
                f"Lançamento {mes}",
                100.0 * mes,
                "1 - ENTRADA",
            ]
        )
    wb.save(path)


def _criar_dre_mes_com_outlier(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "RELATORIO"
    ws.append(["metadata"])
    ws.append(["Emissão", "Descri.", "Vlr.bruto (R$)", "CLASSIFICAÇÃO"])
    for dia in range(1, 6):
        ws.append([f"{dia:02d}/05/2025", f"Lançamento {dia}", 100.0, "1 - ENTRADA"])
    ws.append(["01/06/2025", "Outlier", 50.0, "1 - ENTRADA"])
    wb.save(path)


def _criar_fluxo_mensal(path: Path, mes: int) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.append(["Relatório de movimentos financeiros"])
    ws.append(["Conta:", "BANCO"])
    ws.append([])
    ws.append(["Data Mov.", "Tipo", "Desc. Mov.", "Valor (R$)", "Saldo (R$)"])
    ws.append([f"01/{mes:02d}/2025", "Crédito", "Entrada", 1000.0, 1000.0])
    ws.append([f"15/{mes:02d}/2025", "Débito", "Saída", 250.0, 750.0])
    wb.save(path)


def test_detectar_competencia_dre_usa_maior_data_do_arquivo(tmp_path):
    client = TestClient(main_module.app)
    login_admin(client)
    arquivo = tmp_path / "RELATORIO DRE 01 A 05 2025.xlsx"
    _criar_dre_cumulativo(arquivo)

    with open(arquivo, "rb") as fh:
        resp = client.post(
            "/api/detectar-competencia/dre",
            files={
                "arquivo": (
                    arquivo.name,
                    fh,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert resp.status_code == 200
    body = resp.json()
    assert body["detectado"] is True
    assert body["competencia"] == "05/2025"
    assert body["competencia_input"] == "2025-05"
    assert body["total_datas"] == 5


def test_detectar_competencia_dre_ignora_outlier_e_usa_mes_predominante(tmp_path):
    client = TestClient(main_module.app)
    login_admin(client)
    arquivo = tmp_path / "RELATORIO DRE MES 05.xlsx"
    _criar_dre_mes_com_outlier(arquivo)

    with open(arquivo, "rb") as fh:
        resp = client.post(
            "/api/detectar-competencia/dre",
            files={
                "arquivo": (
                    arquivo.name,
                    fh,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert resp.status_code == 200
    body = resp.json()
    assert body["competencia"] == "05/2025"
    assert body["competencia_input"] == "2025-05"


def test_detectar_competencia_fluxo_usa_maior_data_do_lote(tmp_path):
    client = TestClient(main_module.app)
    login_admin(client)
    arquivo_julho = tmp_path / "RELATORIO DE MOVIMENTO ITAU JUL 2025.xlsx"
    arquivo_agosto = tmp_path / "RELATORIO DE MOVIMENTO CEF AGO 2025.xlsx"
    _criar_fluxo_mensal(arquivo_julho, 7)
    _criar_fluxo_mensal(arquivo_agosto, 8)

    with open(arquivo_julho, "rb") as f1, open(arquivo_agosto, "rb") as f2:
        resp = client.post(
            "/api/detectar-competencia/fluxo_caixa",
            files=[
                (
                    "arquivos",
                    (
                        arquivo_julho.name,
                        f1,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    ),
                ),
                (
                    "arquivos",
                    (
                        arquivo_agosto.name,
                        f2,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    ),
                ),
            ],
        )

    assert resp.status_code == 200
    body = resp.json()
    assert body["detectado"] is True
    assert body["competencia"] == "08/2025"
    assert body["competencia_input"] == "2025-08"
    assert [item["competencia"] for item in body["meses_encontrados"]] == ["07/2025", "08/2025"]
