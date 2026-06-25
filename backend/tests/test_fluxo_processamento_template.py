"""Testes de escrita do Fluxo de Caixa no template."""

import zipfile
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.config import settings
from app.contracts.fluxo_caixa import FCLote, FCMovimento, TipoMovimento
from app.processamento.fluxo_caixa import FluxoCaixaProcessamentoService

HEADERS = [
    "Data",
    "Fornecedor",
    "Crédito",
    "Débito",
    "Saldo",
    "Classificação",
    "Ano",
    "C.M.",
    "Mês",
    "Banco",
    "Empresa",
]


def _criar_template_fluxo(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado"
    ws.append(HEADERS)
    ws.append(
        [
            date(2025, 1, 1),
            "Saldo Inicial CEF",
            100,
            None,
            100,
            "Saldo Inicial CEF",
            "=YEAR(A2)",
            "=MONTH(A2)",
            "=INDEX(mês[],Consolidado!H2,2)",
            "CEF",
            "A IDEAL",
        ]
    )
    ws.append(
        [
            date(2025, 7, 1),
            "linha antiga de julho",
            1,
            None,
            101,
            "Antiga",
            "=YEAR(A3)",
            "=MONTH(A3)",
            "=INDEX(mês[],Consolidado!H3,2)",
            "ITAU",
            "A IDEAL",
        ]
    )
    table = Table(displayName="FluxoConsol", ref="A1:K3")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    wb.save(path)


def _adicionar_abas_resumo_fluxo(path: Path) -> None:
    wb = load_workbook(path)

    apoio = wb.create_sheet("Apoio")
    apoio["B1"] = "Ano"
    apoio["C1"] = "(Tudo)"
    apoio["C4"] = "Janeiro"
    apoio["E4"] = "Fevereiro"
    apoio["O4"] = "Julho"
    apoio["Q4"] = "Total Soma de Crédito"
    apoio["R4"] = "Total Soma de Débito"
    apoio["B5"] = "Rótulos de Linha"
    for col in range(3, 27, 2):
        apoio.cell(row=5, column=col, value="Soma de Crédito")
        apoio.cell(row=5, column=col + 1, value="Soma de Débito")
    apoio["B6"] = "Recebimento de Clientes"
    apoio["B7"] = "Fornecedores"

    fluxo = wb.create_sheet("Fluxo de Caixa ")
    for idx, label in enumerate(
        [
            "JANEIRO",
            "FEVEREIRO",
            "MARÇO",
            "ABRIL",
            "MAIO",
            "JUNHO",
            "JULHO",
            "AGOSTO",
            "SETEMBRO",
            "OUTUBRO",
            "NOVEMBRO",
            "DEZEMBRO",
        ],
        start=4,
    ):
        fluxo.cell(row=5, column=idx, value=label)
    fluxo["Q5"] = "ACUMULADO"

    apresentacao = wb.create_sheet("Apresentação GMP")
    for row in (2, 9, 14, 19):
        for idx, label in enumerate(
            ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"],
            start=3,
        ):
            apresentacao.cell(row=row, column=idx, value=label)
        apresentacao.cell(row=row, column=15, value="Acum")

    wb.save(path)


def test_escrita_preserva_acumulado_e_formula_do_consolidado(tmp_path):
    template = tmp_path / "template_fluxo.xlsx"
    output = tmp_path / "saida_fluxo.xlsx"
    _criar_template_fluxo(template)

    lote = FCLote(
        periodo="07/2025",
        movimentos=[
            FCMovimento(
                data_movimento=date(2025, 7, 2),
                tipo=TipoMovimento.CREDITO,
                descricao="Recebimento",
                valor=Decimal("250.50"),
                saldo=None,
                classificacao="Recebimento de Clientes",
                banco_origem="itau",
            )
        ],
    )
    service = FluxoCaixaProcessamentoService(
        template_path=template,
        output_dir=tmp_path / "output",
        logs_dir=tmp_path / "logs",
        temp_dir=tmp_path / "tmp",
    )

    service._escrever_template(lote, output)

    wb = load_workbook(output, data_only=False)
    ws = wb["Consolidado"]

    assert ws.max_row == 3
    assert ws["B2"].value == "Saldo Inicial CEF"
    assert ws["B3"].value == "Recebimento-Recebimento de Clientes"
    assert ws["C3"].value == 250.5
    assert ws["D3"].value is None
    assert ws["G3"].value == "=YEAR(A3)"
    assert ws["H3"].value == "=MONTH(A3)"
    assert ws["I3"].value == "=INDEX(mês[],Consolidado!H3,2)"
    assert ws["J3"].value == "ITAU"
    assert ws.tables["FluxoConsol"].ref == "A1:K3"


def test_escrita_adiciona_linhas_de_saldo_para_movimentos_com_saldo(tmp_path):
    template = tmp_path / "template_fluxo.xlsx"
    output = tmp_path / "saida_fluxo.xlsx"
    _criar_template_fluxo(template)

    lote = FCLote(
        periodo="07/2025",
        movimentos=[
            FCMovimento(
                data_movimento=date(2025, 7, 2),
                tipo=TipoMovimento.DEBITO,
                descricao="Pagamento",
                valor=Decimal("10"),
                saldo=Decimal("90"),
                classificacao="Fornecedores",
                banco_origem="cef",
            )
        ],
    )
    service = FluxoCaixaProcessamentoService(
        template_path=template,
        output_dir=tmp_path / "output",
        logs_dir=tmp_path / "logs",
        temp_dir=tmp_path / "tmp",
    )

    service._escrever_template(lote, output)

    wb = load_workbook(output, data_only=False)
    ws = wb["Consolidado"]

    assert ws.max_row == 5
    assert ws["B3"].value == "saldo inicial "
    assert ws["C3"].value == 100
    assert ws["E3"].value == 100
    assert ws["F3"].value == "Saldo Inicial CEF"
    assert ws["B4"].value == "Pagamento-Fornecedores"
    assert ws["B5"].value == "saldo "
    assert ws["D5"].value == 90
    assert ws["E5"].value == 0
    assert ws["F5"].value == "Saldo Final CEF"
    assert ws.tables["FluxoConsol"].ref == "A1:K5"


def test_escrita_reaproveita_classificacao_canonica_do_template(tmp_path):
    template = tmp_path / "template_fluxo.xlsx"
    output = tmp_path / "saida_fluxo.xlsx"
    _criar_template_fluxo(template)

    wb = load_workbook(template)
    ws = wb["Consolidado"]
    ws["B3"] = "linha antiga-PARCELAMENTO"
    ws["F3"] = "Parcelamento de Impostos Exercicio Anterior "
    wb.save(template)

    lote = FCLote(
        periodo="07/2025",
        movimentos=[
            FCMovimento(
                data_movimento=date(2025, 7, 2),
                tipo=TipoMovimento.DEBITO,
                descricao="Pagamento parcelamento",
                valor=Decimal("10"),
                saldo=None,
                classificacao="PARCELAMENTO",
                conta_gerencial="PARCELAMENTO",
                banco_origem="itau",
            )
        ],
    )
    service = FluxoCaixaProcessamentoService(
        template_path=template,
        output_dir=tmp_path / "output",
        logs_dir=tmp_path / "logs",
        temp_dir=tmp_path / "tmp",
    )

    service._escrever_template(lote, output)

    wb = load_workbook(output, data_only=False)
    ws = wb["Consolidado"]

    assert ws["B3"].value == "Pagamento parcelamento-PARCELAMENTO"
    assert ws["F3"].value == "Parcelamento de Impostos Exercicio Anterior "


def test_escrita_recalcula_apoio_e_exibe_apenas_mes_selecionado(tmp_path):
    template = tmp_path / "template_fluxo.xlsx"
    output = tmp_path / "saida_fluxo.xlsx"
    _criar_template_fluxo(template)
    _adicionar_abas_resumo_fluxo(template)

    lote = FCLote(
        periodo="08/2025",
        movimentos=[
            FCMovimento(
                data_movimento=date(2025, 8, 4),
                tipo=TipoMovimento.CREDITO,
                descricao="Recebimento agosto",
                valor=Decimal("1000"),
                saldo=None,
                classificacao="Recebimento de Clientes",
                banco_origem="itau",
            ),
            FCMovimento(
                data_movimento=date(2025, 8, 5),
                tipo=TipoMovimento.DEBITO,
                descricao="Pagamento agosto",
                valor=Decimal("200"),
                saldo=None,
                classificacao="Fornecedores",
                banco_origem="cef",
            ),
        ],
    )
    service = FluxoCaixaProcessamentoService(
        template_path=template,
        output_dir=tmp_path / "output",
        logs_dir=tmp_path / "logs",
        temp_dir=tmp_path / "tmp",
    )

    service._escrever_template(
        lote,
        output,
        meses_visiveis=[8],
        preservar_historico=False,
    )

    wb = load_workbook(output, data_only=False)
    ws = wb["Consolidado"]
    meses_consolidado = {
        cell.value.month for cell in ws["A"][1:] if getattr(cell.value, "month", None)
    }
    assert meses_consolidado == {8}

    apoio = wb["Apoio"]
    assert apoio["Q4"].value == "Agosto"
    assert apoio["R4"].value is None
    assert apoio["Q6"].value == 1000.0
    assert apoio["R7"].value == 200.0

    fluxo = wb["Fluxo de Caixa "]
    apresentacao = wb["Apresentação GMP"]
    assert fluxo.column_dimensions["K"].hidden is False
    assert apresentacao.column_dimensions["J"].hidden is False
    for col_idx in [*range(4, 11), *range(12, 16)]:
        assert fluxo.column_dimensions[get_column_letter(col_idx)].hidden is True
    for col_idx in [*range(3, 10), *range(11, 15)]:
        assert apresentacao.column_dimensions[get_column_letter(col_idx)].hidden is True
    assert apresentacao.column_dimensions["P"].hidden is True


def test_escrita_protege_formulas_de_divisao_herdadas_do_template(tmp_path):
    template = tmp_path / "template_fluxo.xlsx"
    output = tmp_path / "saida_fluxo.xlsx"
    _criar_template_fluxo(template)
    _adicionar_abas_resumo_fluxo(template)

    wb = load_workbook(template)
    fluxo = wb["Fluxo de Caixa "]
    fluxo["Q169"] = "=Q168/Q20"
    fluxo["D169"] = "=IFERROR(D168/D20,0)"
    apresentacao = wb["Apresentação GMP"]
    apresentacao["P21"] = "=P20+O21/$O$20"
    apresentacao["Q21"] = "=O21/$O$20"
    apresentacao["C34"] = "=IFERROR(C33/C$10,0)"
    wb.save(template)

    lote = FCLote(
        periodo="08/2025",
        movimentos=[
            FCMovimento(
                data_movimento=date(2025, 8, 4),
                tipo=TipoMovimento.DEBITO,
                descricao="Pagamento sem recebimento",
                valor=Decimal("200"),
                saldo=None,
                classificacao="Fornecedores",
                banco_origem="cef",
            ),
        ],
    )
    service = FluxoCaixaProcessamentoService(
        template_path=template,
        output_dir=tmp_path / "output",
        logs_dir=tmp_path / "logs",
        temp_dir=tmp_path / "tmp",
    )

    service._escrever_template(
        lote,
        output,
        meses_visiveis=[8],
        preservar_historico=False,
    )

    wb = load_workbook(output, data_only=False)
    assert wb["Fluxo de Caixa "]["Q169"].value == "=IFERROR(Q168/Q20,0)"
    assert wb["Fluxo de Caixa "]["D169"].value == "=IFERROR(D168/D20,0)"
    assert wb["Apresentação GMP"]["P21"].value == "=IFERROR(P20+O21/$O$20,0)"
    assert wb["Apresentação GMP"]["Q21"].value == "=IFERROR(O21/$O$20,0)"
    assert wb["Apresentação GMP"]["C34"].value == "=IFERROR(C33/C$10,0)"


def test_escrita_resolve_rotulo_visual_do_template_por_codigo_gerencial(tmp_path):
    template = tmp_path / "template_fluxo.xlsx"
    output = tmp_path / "saida_fluxo.xlsx"
    _criar_template_fluxo(template)
    _adicionar_abas_resumo_fluxo(template)

    wb = load_workbook(template)
    apoio = wb["Apoio"]
    apoio["B6"] = "MATERIAIS DE CONSUMO EM OBRAS "
    apoio["B7"] = "REFEIÇÕES COLABORADORES"
    apoio["B8"] = "PREVISÃO FÉRIAS"
    apoio["B9"] = "MULTA RESCISORIAS FGTS"
    wb.save(template)

    lote = FCLote(
        periodo="08/2025",
        movimentos=[
            FCMovimento(
                data_movimento=date(2025, 8, 4),
                tipo=TipoMovimento.DEBITO,
                descricao="Materiais obra",
                valor=Decimal("300"),
                saldo=None,
                classificacao="8.9 - MATERIAIS DE CONSUMO EM OBRAS",
                conta_gerencial="8.9 - MATERIAIS DE CONSUMO EM OBRAS",
                banco_origem="cef",
            ),
            FCMovimento(
                data_movimento=date(2025, 8, 5),
                tipo=TipoMovimento.DEBITO,
                descricao="Refeições",
                valor=Decimal("120"),
                saldo=None,
                classificacao="12.5 - REFEIÇÕES FUNCIONARIOS",
                conta_gerencial="12.5 - REFEIÇÕES FUNCIONARIOS",
                banco_origem="cef",
            ),
            FCMovimento(
                data_movimento=date(2025, 8, 6),
                tipo=TipoMovimento.DEBITO,
                descricao="Previsão férias",
                valor=Decimal("45"),
                saldo=None,
                classificacao="12.100 - PREVISÃO FÉRIAS",
                conta_gerencial="12.100 - PREVISÃO FÉRIAS",
                banco_origem="cef",
            ),
            FCMovimento(
                data_movimento=date(2025, 8, 7),
                tipo=TipoMovimento.DEBITO,
                descricao="FGTS rescisório",
                valor=Decimal("3525.27"),
                saldo=None,
                classificacao="12.4 - FGTS",
                conta_gerencial="12.4 - FGTS",
                banco_origem="cef",
            ),
        ],
    )
    service = FluxoCaixaProcessamentoService(
        template_path=template,
        output_dir=tmp_path / "output",
        logs_dir=tmp_path / "logs",
        temp_dir=tmp_path / "tmp",
    )

    service._escrever_template(
        lote,
        output,
        meses_visiveis=[8],
        preservar_historico=False,
    )

    wb = load_workbook(output, data_only=False)
    ws = wb["Consolidado"]
    assert ws["F2"].value == "MATERIAIS DE CONSUMO EM OBRAS "
    assert ws["F3"].value == "REFEIÇÕES COLABORADORES"
    assert ws["F4"].value == "PREVISÃO FÉRIAS"
    assert ws["F5"].value == "MULTA RESCISORIAS FGTS"

    apoio = wb["Apoio"]
    assert apoio["R6"].value == 300.0
    assert apoio["R7"].value == 120.0
    assert apoio["R8"].value == 45.0
    assert apoio["R9"].value == 3525.27
    assert "8.9 - MATERIAIS DE CONSUMO EM OBRAS" not in [
        apoio.cell(row=row, column=2).value for row in range(6, apoio.max_row + 1)
    ]
    assert "12.4 - FGTS" not in [
        apoio.cell(row=row, column=2).value for row in range(6, apoio.max_row + 1)
    ]


def test_escrita_remove_marcadores_ok_da_apresentacao(tmp_path):
    template = tmp_path / "template_fluxo.xlsx"
    output = tmp_path / "saida_fluxo.xlsx"
    _criar_template_fluxo(template)
    _adicionar_abas_resumo_fluxo(template)

    wb = load_workbook(template)
    apresentacao = wb["Apresentação GMP"]
    apresentacao["A21"] = "OK"
    apresentacao["A22"] = "ok"
    apresentacao["A23"] = "Revisar"
    wb.save(template)

    lote = FCLote(
        periodo="08/2025",
        movimentos=[
            FCMovimento(
                data_movimento=date(2025, 8, 4),
                tipo=TipoMovimento.DEBITO,
                descricao="Pagamento agosto",
                valor=Decimal("200"),
                saldo=None,
                classificacao="Fornecedores",
                banco_origem="cef",
            ),
        ],
    )
    service = FluxoCaixaProcessamentoService(
        template_path=template,
        output_dir=tmp_path / "output",
        logs_dir=tmp_path / "logs",
        temp_dir=tmp_path / "tmp",
    )

    service._escrever_template(
        lote,
        output,
        meses_visiveis=[8],
        preservar_historico=False,
    )

    wb = load_workbook(output, data_only=False)
    apresentacao = wb["Apresentação GMP"]
    assert apresentacao["A21"].value is None
    assert apresentacao["A22"].value is None
    assert apresentacao["A23"].value == "Revisar"


def test_escrita_mantem_transferencia_rastreavel_sem_agregar_no_fluxo(tmp_path):
    template = tmp_path / "template_fluxo.xlsx"
    output = tmp_path / "saida_fluxo.xlsx"
    _criar_template_fluxo(template)
    _adicionar_abas_resumo_fluxo(template)

    lote = FCLote(
        periodo="08/2025",
        movimentos=[
            FCMovimento(
                data_movimento=date(2025, 8, 4),
                tipo=TipoMovimento.TRANSFERENCIA,
                descricao="Transferência para Safra",
                valor=Decimal("50"),
                saldo=Decimal("50"),
                classificacao="Transferência Emitida",
                conta_gerencial="Transferência entre Bancos",
                banco_origem="itau",
            ),
            FCMovimento(
                data_movimento=date(2025, 8, 5),
                tipo=TipoMovimento.DEBITO,
                descricao="Fornecedor",
                valor=Decimal("20"),
                saldo=Decimal("30"),
                classificacao="Fornecedores",
                banco_origem="itau",
            ),
        ],
    )
    service = FluxoCaixaProcessamentoService(
        template_path=template,
        output_dir=tmp_path / "output",
        logs_dir=tmp_path / "logs",
        temp_dir=tmp_path / "tmp",
    )

    service._escrever_template(lote, output, meses_visiveis=[8], preservar_historico=False)

    wb = load_workbook(output, data_only=False)
    consolidado = wb["Consolidado"]
    transferencia = next(
        row
        for row in consolidado.iter_rows(min_row=2, values_only=True)
        if row[5] == "Transferência Emitida"
    )
    assert transferencia[2] is None
    assert transferencia[3] == 50.0

    apoio = wb["Apoio"]
    labels_apoio = [apoio.cell(row=row, column=2).value for row in range(6, apoio.max_row + 1)]
    assert "Transferência Emitida" not in labels_apoio
    linha_fornecedores = labels_apoio.index("Fornecedores") + 6
    assert apoio.cell(row=linha_fornecedores, column=18).value == 20.0
    linha_saldo_inicial = labels_apoio.index("Saldo Inicial Itau") + 6
    assert apoio.cell(row=linha_saldo_inicial, column=17).value == 100.0


def test_linhas_de_saldo_encadeiam_fechamento_de_um_mes_na_abertura_do_proximo(tmp_path):
    service = FluxoCaixaProcessamentoService(
        template_path=tmp_path / "nao_usado.xlsx",
        output_dir=tmp_path / "output",
        logs_dir=tmp_path / "logs",
        temp_dir=tmp_path / "tmp",
    )
    movimentos = [
        FCMovimento(
            data_movimento=date(2025, 1, 30),
            tipo=TipoMovimento.CREDITO,
            descricao="Entrada janeiro",
            valor=Decimal("100"),
            saldo=Decimal("100"),
            classificacao="Receita",
            banco_origem="itau",
        ),
        FCMovimento(
            data_movimento=date(2025, 1, 31),
            tipo=TipoMovimento.DEBITO,
            descricao="Saída janeiro",
            valor=Decimal("10"),
            saldo=Decimal("90"),
            classificacao="Fornecedores",
            banco_origem="itau",
        ),
        FCMovimento(
            data_movimento=date(2025, 2, 1),
            tipo=TipoMovimento.CREDITO,
            descricao="Entrada fevereiro",
            valor=Decimal("25"),
            saldo=Decimal("115"),
            classificacao="Receita",
            banco_origem="itau",
        ),
    ]

    linhas = service._linhas_movimentos_com_saldos(movimentos)
    linhas_saldo = [linha for linha in linhas if isinstance(linha, list)]

    assert len(linhas_saldo) == 4
    assert linhas_saldo[0][4] == 0.0
    assert linhas_saldo[1][3] == 90.0
    assert linhas_saldo[2][4] == 90.0
    assert linhas_saldo[3][3] == 115.0


def test_fechamento_bancario_zero_do_extrato_prevalece_sobre_saldo_calculado(tmp_path):
    service = FluxoCaixaProcessamentoService(
        template_path=tmp_path / "nao_usado.xlsx",
        output_dir=tmp_path / "output",
        logs_dir=tmp_path / "logs",
        temp_dir=tmp_path / "tmp",
    )
    movimentos = [
        FCMovimento(
            data_movimento=date(2025, 1, 1),
            tipo=TipoMovimento.CREDITO,
            descricao="Entrada",
            valor=Decimal("100"),
            saldo=Decimal("100"),
            classificacao="Receita",
            banco_origem="itau",
        ),
        FCMovimento(
            data_movimento=date(2025, 1, 2),
            tipo=TipoMovimento.DEBITO,
            descricao="Ajuste bancário",
            valor=Decimal("10"),
            saldo=Decimal("0"),
            classificacao="Fornecedores",
            banco_origem="itau",
        ),
    ]

    assert service._saldo_final_grupo(movimentos, Decimal("90")) == Decimal("0")


def test_escrita_mapeia_4_3_para_outros_materiais_dos_fornecedores(tmp_path):
    template = tmp_path / "template_fluxo.xlsx"
    output = tmp_path / "saida_fluxo.xlsx"
    _criar_template_fluxo(template)
    _adicionar_abas_resumo_fluxo(template)

    wb_template = load_workbook(template)
    wb_template["Apoio"]["B6"] = "OUTROS MATERIAIS DE APLICAÇÃO"
    wb_template.save(template)

    lote = FCLote(
        periodo="08/2025",
        movimentos=[
            FCMovimento(
                data_movimento=date(2025, 8, 8),
                tipo=TipoMovimento.DEBITO,
                descricao="Outros materiais",
                valor=Decimal("40"),
                saldo=None,
                classificacao="4.3 - OUTROS MATERIAIS",
                conta_gerencial="4.3 - OUTROS MATERIAIS",
                banco_origem="cef",
            )
        ],
    )
    service = FluxoCaixaProcessamentoService(
        template_path=template,
        output_dir=tmp_path / "output",
        logs_dir=tmp_path / "logs",
        temp_dir=tmp_path / "tmp",
    )

    service._escrever_template(lote, output, meses_visiveis=[8], preservar_historico=False)

    wb = load_workbook(output, data_only=False)
    assert wb["Consolidado"]["F2"].value == "OUTROS MATERIAIS DE APLICAÇÃO"
    assert wb["Apoio"]["R6"].value == 40.0


def test_escrita_exibe_conta_nova_com_nome_sem_codigo_no_relatorio(tmp_path):
    template = tmp_path / "template_fluxo.xlsx"
    output = tmp_path / "saida_fluxo.xlsx"
    _criar_template_fluxo(template)
    _adicionar_abas_resumo_fluxo(template)

    lote = FCLote(
        periodo="08/2025",
        movimentos=[
            FCMovimento(
                data_movimento=date(2025, 8, 8),
                tipo=TipoMovimento.DEBITO,
                descricao="Nova despesa",
                valor=Decimal("70"),
                saldo=None,
                classificacao="18.99 - NOVA DESPESA",
                conta_gerencial="18.99 - NOVA DESPESA",
                banco_origem="cef",
            )
        ],
    )
    service = FluxoCaixaProcessamentoService(
        template_path=template,
        output_dir=tmp_path / "output",
        logs_dir=tmp_path / "logs",
        temp_dir=tmp_path / "tmp",
    )

    service._escrever_template(lote, output, meses_visiveis=[8], preservar_historico=False)

    wb = load_workbook(output, data_only=False)
    assert wb["Consolidado"]["F2"].value == "NOVA DESPESA"
    labels_apoio = [
        wb["Apoio"].cell(row=row, column=2).value
        for row in range(6, wb["Apoio"].max_row + 1)
    ]
    assert "NOVA DESPESA" in labels_apoio


def test_escrita_fluxo_remove_slicers_do_template_real(tmp_path):
    output = tmp_path / "fluxo_sem_slicers.xlsx"
    service = FluxoCaixaProcessamentoService(
        template_path=settings.template_fluxo_path,
        output_dir=tmp_path / "output",
        logs_dir=tmp_path / "logs",
        temp_dir=tmp_path / "tmp",
    )
    lote = FCLote(
        periodo="06/2025",
        movimentos=[
            FCMovimento(
                data_movimento=date(2025, 6, 10),
                tipo=TipoMovimento.CREDITO,
                descricao="Recebimento",
                valor=Decimal("100"),
                saldo=Decimal("100"),
                classificacao="Recebimento de Clientes",
                banco_origem="itau",
            )
        ],
    )

    service._escrever_template(lote, output, preservar_historico=False)

    with zipfile.ZipFile(output, "r") as workbook:
        nomes = set(workbook.namelist())
        workbook_xml = workbook.read("xl/workbook.xml").decode("utf-8")
    assert not any(nome.startswith("xl/slicers/") for nome in nomes)
    assert not any(nome.startswith("xl/slicerCaches/") for nome in nomes)
    assert "#REF!" not in workbook_xml
