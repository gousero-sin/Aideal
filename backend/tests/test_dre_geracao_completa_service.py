"""Testes unitários do serviço de geração completa DRE (modo DB)."""

import shutil
import tempfile
import zipfile
from decimal import Decimal
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook

from app.config import settings
from app.contracts.persistence import DRELancamentoDB, DREUpload
from app.db.connection import DatabaseConnection
from app.db.manager import MigrationManager
from app.processamento.dre_geracao_completa import DREGeracaoCompletaService
from app.repository.dre_repository import DRERepository
from app.templates.writer import TemplateWriter


def _build_service_com_dados(meses: list[int] | None = None) -> DREGeracaoCompletaService:
    meses = meses or [1, 2, 6]
    tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
    tmp.close()
    db = DatabaseConnection(tmp.name)
    MigrationManager(db).migrate()
    repo = DRERepository(db)

    for mes in meses:
        upload = DREUpload(
            id=str(uuid4()),
            arquivo_nome=f"{mes:02d}_2025.xls",
            arquivo_sha256=f"hash_{mes}",
            competencia_ano=2025,
            competencia_mes=mes,
            status="completed",
        )
        repo.uploads.create(upload)
        repo.lancamentos.create(
            DRELancamentoDB(
                upload_id=upload.id,
                competencia_ano=2025,
                competencia_mes=mes,
                data_lancamento=f"2025-{mes:02d}-15",
                historico=f"Lanc {mes}",
                credito=Decimal("100.00"),
                debito=Decimal("0"),
                natureza_raw="1.2.3 - Origens Financeiras",
                centro_custo="ADMINISTRATIVO",
                rubrica="Juros Recebidos",
                conta_pai="(+/-)Despesas e Receitas Financeiras",
                hash_linha=f"hash_lanc_{mes}",
            )
        )

    return DREGeracaoCompletaService(db)


def _build_service_com_receita_liquida_e_impostos() -> DREGeracaoCompletaService:
    tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
    tmp.close()
    db = DatabaseConnection(tmp.name)
    MigrationManager(db).migrate()
    repo = DRERepository(db)

    upload = DREUpload(
        id=str(uuid4()),
        arquivo_nome="01_2026.xls",
        arquivo_sha256="hash_receita_2026_01",
        competencia_ano=2026,
        competencia_mes=1,
        status="completed",
    )
    repo.uploads.create(upload)

    repo.lancamentos.create(
        DRELancamentoDB(
            upload_id=upload.id,
            competencia_ano=2026,
            competencia_mes=1,
            data_lancamento="2026-01-10",
            historico="Recebimento cliente",
            valor_bruto=Decimal("1300.00"),
            credito=Decimal("1000.00"),
            debito=Decimal("0"),
            natureza_raw="1.1.1 - Recebimento de Clientes",
            centro_custo="OBRA A",
            hash_linha="hash_receita_liquida",
        )
    )
    repo.lancamentos.create(
        DRELancamentoDB(
            upload_id=upload.id,
            competencia_ano=2026,
            competencia_mes=1,
            data_lancamento="2026-01-10",
            historico="Recebimento de empréstimo",
            valor_bruto=Decimal("100.00"),
            credito=Decimal("100.00"),
            debito=Decimal("0"),
            natureza_raw="2.2 - Recebimento de empréstimo",
            rubrica="2.2 - Recebimento de empréstimo",
            centro_custo="OBRA A",
            hash_linha="hash_emprestimo",
        )
    )
    repo.lancamentos.create(
        DRELancamentoDB(
            upload_id=upload.id,
            competencia_ano=2026,
            competencia_mes=1,
            data_lancamento="2026-01-10",
            historico="IR retido",
            valor_bruto=Decimal("200.00"),
            credito=Decimal("0"),
            debito=Decimal("200.00"),
            natureza_raw="IR Retido",
            rubrica="IR Retido",
            centro_custo="OBRA A",
            hash_linha="hash_ir_retido",
        )
    )
    repo.lancamentos.create(
        DRELancamentoDB(
            upload_id=upload.id,
            competencia_ano=2026,
            competencia_mes=1,
            data_lancamento="2026-01-10",
            historico="CSLL retido",
            valor_bruto=Decimal("100.00"),
            credito=Decimal("0"),
            debito=Decimal("100.00"),
            natureza_raw="CSLL",
            rubrica="CSLL",
            centro_custo="OBRA A",
            hash_linha="hash_csll_retido",
        )
    )
    repo.lancamentos.create(
        DRELancamentoDB(
            upload_id=upload.id,
            competencia_ano=2026,
            competencia_mes=1,
            data_lancamento="2026-01-10",
            historico="ICMS",
            valor_bruto=Decimal("20.00"),
            credito=Decimal("0"),
            debito=Decimal("20.00"),
            natureza_raw="17.9 - ICMS",
            rubrica="ICMS",
            centro_custo="OBRA A",
            hash_linha="hash_icms",
        )
    )
    repo.lancamentos.create(
        DRELancamentoDB(
            upload_id=upload.id,
            competencia_ano=2026,
            competencia_mes=1,
            data_lancamento="2026-01-10",
            historico="Investimento sede",
            valor_bruto=Decimal("50.00"),
            credito=Decimal("0"),
            debito=Decimal("50.00"),
            natureza_raw="15.4 - DESPESAS C/ CONSTRUÇÃO ESCRITORIO ADM",
            rubrica="Construção da Nova Sede",
            centro_custo="OBRA A",
            hash_linha="hash_investimento",
        )
    )

    return DREGeracaoCompletaService(db)


def _copiar_template(tmp_path: Path) -> Path:
    destino = tmp_path / "template_dre_teste.xlsx"
    shutil.copyfile(settings.template_dre_path, destino)
    return destino


def _celula_dre_por_rotulo(ws, rotulo: str, coluna: str) -> object:
    for linha in range(1, ws.max_row + 1):
        if ws.cell(linha, 1).value == rotulo:
            return ws[f"{coluna}{linha}"].value
    raise AssertionError(f"Rótulo DRE ausente: {rotulo}")


def test_verificar_dados_modo_padrao_aceita_gaps_sem_exigir_mes_alvo():
    service = _build_service_com_dados()

    ok = service.verificar_dados("06/2025")
    assert ok["valido"] is True
    assert ok["estrategia_meses"] == "competencia"
    assert ok["meses_utilizados"] == [1, 2, 6]

    com_gap = service.verificar_dados("05/2025")
    assert com_gap["valido"] is True
    assert com_gap["estrategia_meses"] == "competencia"
    assert com_gap["meses_utilizados"] == [1, 2]


def test_verificar_dados_meses_incluir_dispensa_mes_alvo():
    service = _build_service_com_dados()

    resultado = service.verificar_dados("05/2025", meses_incluir=[1, 2])
    assert resultado["valido"] is True
    assert resultado["estrategia_meses"] == "meses_incluir"
    assert resultado["meses_utilizados"] == [1, 2]


def test_verificar_dados_ano_todo_usa_todos_mes_disponiveis():
    service = _build_service_com_dados()

    resultado = service.verificar_dados("05/2025", ano_todo=True)
    assert resultado["valido"] is True
    assert resultado["estrategia_meses"] == "ano_todo"
    assert resultado["meses_utilizados"] == [1, 2, 6]


def test_verificar_dados_rejeita_meses_indisponiveis():
    service = _build_service_com_dados()

    resultado = service.verificar_dados("06/2025", meses_incluir=[1, 7])
    assert resultado["valido"] is False
    assert "Meses sem upload completed" in resultado["error"]


def test_verificar_dados_funciona_com_apenas_um_mes():
    service = _build_service_com_dados([6])

    resultado = service.verificar_dados("06/2025")
    assert resultado["valido"] is True
    assert resultado["meses_disponiveis"] == [6]
    assert resultado["meses_utilizados"] == [6]


def test_verificar_dados_rejeita_competencia_com_mes_zero():
    service = _build_service_com_dados([6])

    resultado = service.verificar_dados("00/2025")
    assert resultado["valido"] is False
    assert "Mês da competência inválido" in resultado["error"]


def test_verificar_dados_competencia_antes_do_primeiro_mes_disponivel_falha():
    service = _build_service_com_dados([6])

    resultado = service.verificar_dados("05/2025")
    assert resultado["valido"] is False
    assert "Nenhum mês disponível para geração até 05/2025" in resultado["error"]


def test_visibilidade_colunas_exibe_junho_quando_mes_6_presente(tmp_path):
    service = _build_service_com_dados()
    template_teste = _copiar_template(tmp_path)

    with TemplateWriter(template_teste) as writer:
        vis = service._controlar_visibilidade_colunas_dre(writer, [1, 2, 6])
        ws = writer._wb["DRE"]

        assert ws.column_dimensions["N"].hidden is False
        assert ws.column_dimensions["O"].hidden is False
        assert ws.column_dimensions["L"].hidden is True
        assert ws.column_dimensions["M"].hidden is True
        assert vis["meses_visiveis"] == [1, 2, 6]


def test_gerar_arquivo_persiste_visibilidade_dre_no_arquivo_final(tmp_path):
    service = _build_service_com_dados([6])
    output_path = tmp_path / "dre_mes_06.xlsx"

    resultado = service.gerar_arquivo(
        competencia="06/2025",
        output_path=output_path,
    )

    assert resultado["success"] is True
    assert output_path.exists()

    wb = load_workbook(output_path)
    ws = wb["DRE"]
    ws_bd = wb["BD_FLUXO"]
    ws_apoio = wb["APOIO"]
    ws_detalhe = wb["DETALHE_MENSAL_DB"]
    # Somente junho visível entre meses; colunas ANO sempre visíveis.
    assert ws.column_dimensions["B"].hidden is True
    assert ws.column_dimensions["C"].hidden is True
    assert ws.column_dimensions["N"].hidden is False
    assert ws.column_dimensions["O"].hidden is False
    assert ws.column_dimensions["AH"].hidden is False
    assert ws.column_dimensions["AI"].hidden is False
    assert ws_bd["H2"].value == 2025
    assert ws_bd["I2"].value == 6
    assert ws_bd["J2"].value == "Jun"
    assert ws_bd["M2"].value == 100
    assert ws_bd["N2"].value == "Juros Recebidos"
    assert ws_bd["O2"].value == "(+) Receitas Financeiras"
    assert ws_bd["P2"].value == "(+/-)Despesas e Receitas Financeiras"
    assert ws_bd["Q2"].value == 6
    assert ws_bd["R2"].value == 2025
    assert ws_apoio["B5"].value == "Conta Pai"
    assert ws_apoio["C5"].value == "Jan"
    assert ws_apoio["G5"].value == "Mai"
    assert ws_apoio["H5"].value == "Jun"
    assert ws_apoio["O5"].value == "Total Geral"
    assert ws_detalhe["A1"].value == "DRE - Detalhamento mensal (gerado do banco)"
    assert ws_detalhe["A4"].value == "Resumo mensal"
    assert ws_detalhe["A5"].value == "Mes"
    assert ws_detalhe["A6"].value == 6

    with zipfile.ZipFile(output_path, "r") as zf:
        sheet2_xml = zf.read("xl/worksheets/sheet2.xml").decode("utf-8", errors="ignore")
        sheet2_rels = zf.read("xl/worksheets/_rels/sheet2.xml.rels").decode(
            "utf-8",
            errors="ignore",
        )
        sheet6_rels = zf.read("xl/worksheets/_rels/sheet6.xml.rels").decode(
            "utf-8",
            errors="ignore",
        )
        sheet3_xml = zf.read("xl/worksheets/sheet3.xml").decode("utf-8", errors="ignore")
        workbook_xml = zf.read("xl/workbook.xml").decode("utf-8", errors="ignore")
        content_types_xml = zf.read("[Content_Types].xml").decode("utf-8", errors="ignore")
        table_bd_fluxo_xml = zf.read("xl/tables/table2.xml").decode("utf-8", errors="ignore")
        table_detalhe_xml = zf.read("xl/tables/table5.xml").decode("utf-8", errors="ignore")
        assert "slicerList" not in sheet2_xml
        assert "relationships/slicer" not in sheet2_rels
        # BD_FLUXO preserva XML estrutural do template e troca somente sheetData.
        assert "mc:Ignorable" in sheet3_xml
        assert 'tablePart r:id="rId2"' in sheet3_xml
        assert "fullCalcOnLoad" in workbook_xml
        assert "forceFullCalc" in workbook_xml
        assert "slicerCaches" not in workbook_xml
        # Excel-safe mode: pivotCache, pivotTables e charts são removidos do
        # pacote final para evitar prompt de reparo no Excel Desktop.
        assert not any(n.startswith("xl/pivotCache/") for n in zf.namelist())
        assert not any(n.startswith("xl/pivotTables/") for n in zf.namelist())
        assert not any(n.startswith("xl/charts/") for n in zf.namelist())
        assert "/xl/pivotCache/" not in content_types_xml
        assert "/xl/pivotTables/" not in content_types_xml
        assert "/xl/charts/" not in content_types_xml
        assert "<pivotCaches" not in workbook_xml
        assert not any(n.startswith("xl/slicerCaches/") for n in zf.namelist())
        assert not any(n.startswith("xl/slicers/") for n in zf.namelist())
        assert "/xl/slicerCaches/" not in content_types_xml
        assert "/xl/slicers/" not in content_types_xml
        assert 'displayName="BD_FLUXO1"' in table_bd_fluxo_xml
        assert "calculatedColumnFormula" not in table_bd_fluxo_xml
        assert 'displayName="DETALHE_MENSAL_DB"' in table_detalhe_xml
        assert "A13:R14" in table_detalhe_xml
        assert "../tables/table5.xml" in sheet6_rels
        assert "/xl/tables/table5.xml" not in sheet6_rels


def test_gerar_arquivo_materializa_receita_liquida_e_resultado_gerencial(tmp_path):
    service = _build_service_com_receita_liquida_e_impostos()
    output_path = tmp_path / "dre_receita_liquida.xlsx"

    resultado = service.gerar_arquivo(
        competencia="01/2026",
        output_path=output_path,
    )

    assert resultado["success"] is True
    assert resultado["celulas_dre_materializadas"] > 0

    wb_valores = load_workbook(output_path, data_only=True)
    ws_dre_valores = wb_valores["DRE"]
    ws_apoio_valores = wb_valores["APOIO"]

    assert ws_apoio_valores["C6"].value == 1320
    assert ws_apoio_valores["C7"].value == 1320
    assert ws_apoio_valores["C8"].value == 1000
    assert ws_dre_valores["B6"].value == 1320
    assert ws_dre_valores["B7"].value == 1000
    assert ws_dre_valores["B8"].value == -20
    assert ws_dre_valores["B12"].value == 0
    assert ws_dre_valores["B19"].value == 1300
    assert ws_dre_valores["B34"].value == 1300
    assert _celula_dre_por_rotulo(ws_dre_valores, "(=)RESULTADO OPERACIONAL", "B") == 1300
    assert _celula_dre_por_rotulo(ws_dre_valores, "(-)IRPJ/CSLL", "B") == -300
    assert _celula_dre_por_rotulo(ws_dre_valores, "CSLL", "B") == -100
    assert _celula_dre_por_rotulo(ws_dre_valores, "IRPJ", "B") == -200
    assert _celula_dre_por_rotulo(ws_dre_valores, "(=)RESULTADO LÍQUIDO", "B") == 1030
    assert _celula_dre_por_rotulo(ws_dre_valores, "(-)Investimentos", "B") == -50
    assert _celula_dre_por_rotulo(ws_dre_valores, "(=)RESULTADO GERENCIAL", "B") == 980
    assert ws_dre_valores["AH6"].value == 1320
    assert ws_dre_valores["AH7"].value == 1000
    assert ws_dre_valores["AH8"].value == -20
    assert ws_dre_valores["AH12"].value == 0
    assert ws_dre_valores["AH19"].value == 1300
    assert ws_dre_valores["AH34"].value == 1300
    assert _celula_dre_por_rotulo(ws_dre_valores, "(=)RESULTADO OPERACIONAL", "AH") == 1300
    assert _celula_dre_por_rotulo(ws_dre_valores, "(-)IRPJ/CSLL", "AH") == -300
    assert _celula_dre_por_rotulo(ws_dre_valores, "CSLL", "AH") == -100
    assert _celula_dre_por_rotulo(ws_dre_valores, "IRPJ", "AH") == -200
    assert _celula_dre_por_rotulo(ws_dre_valores, "(=)RESULTADO LÍQUIDO", "AH") == 1030
    assert _celula_dre_por_rotulo(ws_dre_valores, "(-)Investimentos", "AH") == -50
    assert _celula_dre_por_rotulo(ws_dre_valores, "(=)RESULTADO GERENCIAL", "AH") == 980

    wb_formulas = load_workbook(output_path, data_only=False)
    ws_dre_formulas = wb_formulas["DRE"]
    assert not isinstance(ws_dre_formulas["B19"].value, str)
    assert ws_dre_formulas["B19"].value == 1300
    assert ws_dre_formulas["B34"].value == 1300
    assert _celula_dre_por_rotulo(ws_dre_formulas, "(=)RESULTADO OPERACIONAL", "B") == 1300
    assert _celula_dre_por_rotulo(ws_dre_formulas, "(-)IRPJ/CSLL", "B") == -300
    assert _celula_dre_por_rotulo(ws_dre_formulas, "(=)RESULTADO LÍQUIDO", "B") == 1030
    assert _celula_dre_por_rotulo(ws_dre_formulas, "(=)RESULTADO GERENCIAL", "B") == 980


def test_gerar_arquivo_protege_formulas_auxiliares_da_apoio(tmp_path):
    """A aba APOIO não pode abrir no Excel com #N/A ou #DIV/0! herdados."""
    service = _build_service_com_receita_liquida_e_impostos()
    output_path = tmp_path / "dre_apoio_sem_erros.xlsx"

    service.gerar_arquivo(competencia="01/2026", output_path=output_path)

    wb = load_workbook(output_path, data_only=False)
    ws_apoio = wb["APOIO"]

    assert ws_apoio["R5"].value.startswith("=IFERROR(VLOOKUP(")
    assert ws_apoio["AF20"].value.startswith("=IFERROR(")


def test_converte_linha_bd_fluxo_expandida_classifica_conta_sem_grupo_como_nao_operacional():
    service = _build_service_com_dados([6])
    lanc = DRELancamentoDB(
        upload_id="upload-x",
        competencia_ano=2025,
        competencia_mes=6,
        data_lancamento="2025-06-15",
        historico="Lanc fallback",
        credito=Decimal("150.00"),
        debito=Decimal("25.00"),
        natureza_raw="Natureza nao mapeada",
        centro_custo="OBRA XPTO",
        rubrica="Rubrica Persistida",
        conta_pai="Conta Pai Persistida",
        hash_linha="hash-fallback",
    )

    linha = service._converte_linha_bd_fluxo_expandida(lanc, plano={})

    assert linha[8] == 6
    assert linha[9] == "Jun"
    assert linha[12] == 125.0
    assert linha[13] == "Recebimentos Não Operacionais"
    assert linha[14] == "(+) Recebimentos Não Operacionais"
    assert linha[15] == "(+/-)Despesas e Recebimentos Não Operacionais"
    assert linha[16] == 5


def test_agregar_apoio_receita_bruta_e_faturamento_brutos_vs_liquidos():
    """Receita Bruta soma valor_bruto; Faturamento soma o crédito líquido."""
    service = _build_service_com_dados([6])
    with TemplateWriter(settings.template_dre_path) as writer:
        plano = service._ler_plano_contas(writer)

    lancamentos = [
        # Recebimento de cliente com estorno parcial lançado como débito.
        DRELancamentoDB(
            upload_id="u1",
            competencia_ano=2025,
            competencia_mes=6,
            data_lancamento="2025-06-10",
            historico="Recebimento com impostos",
            valor_bruto=Decimal("1300.00"),
            credito=Decimal("1000.00"),
            debito=Decimal("0"),
            natureza_raw="1.1.1 - Recebimento de Clientes - prestação de serviço",
            centro_custo="OBRA A",
            hash_linha="h1",
        ),
        # Subconta irmã (herda classificação de Receita Bruta via família 1.1).
        DRELancamentoDB(
            upload_id="u1",
            competencia_ano=2025,
            competencia_mes=6,
            data_lancamento="2025-06-12",
            historico="Venda material",
            valor_bruto=Decimal("700.00"),
            credito=Decimal("500.00"),
            debito=Decimal("0"),
            natureza_raw="1.1.2 - Venda Material",
            centro_custo="OBRA A",
            hash_linha="h2",
        ),
    ]

    linhas, _ = service._agregar_para_apoio(lancamentos, plano)
    por_label = {row[1]: row for row in linhas}

    assert por_label["(=)Receita Bruta"][7] == 2000.0
    assert por_label["(=)Receita Bruta"][14] == 2000.0
    assert por_label["(+)Receita Bruta"][7] == 2000.0
    assert por_label["Faturamento"][7] == 1500.0
    assert por_label["Faturamento"][14] == 1500.0


def test_agregar_apoio_mapeia_previsoes_atuais_de_ferias_e_13_por_codigo():
    service = _build_service_com_dados([6])
    with TemplateWriter(settings.template_dre_path) as writer:
        plano = service._ler_plano_contas(writer)

    lancamentos = [
        DRELancamentoDB(
            upload_id="u1",
            competencia_ano=2026,
            competencia_mes=5,
            data_lancamento="2026-05-10",
            historico="Previsão férias",
            valor_bruto=Decimal("45.00"),
            credito=Decimal("0"),
            debito=Decimal("45.00"),
            natureza_raw="12.100 - PREVISÃO FÉRIAS",
            rubrica="12.100 - PREVISÃO FÉRIAS",
            centro_custo="OBRA A",
            hash_linha="h-ferias",
        ),
        DRELancamentoDB(
            upload_id="u1",
            competencia_ano=2026,
            competencia_mes=5,
            data_lancamento="2026-05-10",
            historico="Previsão 13",
            valor_bruto=Decimal("55.00"),
            credito=Decimal("0"),
            debito=Decimal("55.00"),
            natureza_raw="12.101 - PREVISÃO 13°",
            rubrica="12.101 - PREVISÃO 13°",
            centro_custo="OBRA A",
            hash_linha="h-decimo",
        ),
    ]

    linhas, _ = service._agregar_para_apoio(lancamentos, plano)
    por_label = {row[1]: row for row in linhas}

    assert por_label["PREVISAO FÉRIAS"][6] == -45.0
    assert por_label["13° PREVISAO"][6] == -55.0
    assert por_label["Despesas com Pessoal"][6] == -100.0
    assert por_label["(-)Gastos Fixos"][6] == -100.0
    assert "SALARIO" not in por_label


def test_saldos_painel_por_mes_trata_codigos_de_imposto_como_deducao():
    lancamentos = [
        DRELancamentoDB(
            upload_id="u1",
            competencia_ano=2026,
            competencia_mes=2,
            data_lancamento="2026-02-10",
            historico="Receita",
            valor_bruto=Decimal("1000.00"),
            credito=Decimal("1000.00"),
            debito=Decimal("0"),
            natureza_raw="1.1.1 - Recebimento de Clientes",
            rubrica="1.1.1 - Recebimento de Clientes",
            centro_custo="OBRA A",
            hash_linha="h-receita",
        ),
        DRELancamentoDB(
            upload_id="u1",
            competencia_ano=2026,
            competencia_mes=2,
            data_lancamento="2026-02-10",
            historico="ISS",
            valor_bruto=Decimal("111.00"),
            credito=Decimal("0"),
            debito=Decimal("111.00"),
            natureza_raw="17.8 - ISS",
            rubrica="17.8 - ISS",
            centro_custo="OBRA A",
            hash_linha="h-iss",
        ),
    ]

    assert DREGeracaoCompletaService._saldos_painel_por_mes(lancamentos) == {2: 1000.0}


def test_agregar_apoio_receita_bruta_recompoe_bruto_sem_abater_ir_na_receita_liquida():
    service = _build_service_com_dados([6])
    with TemplateWriter(settings.template_dre_path) as writer:
        plano = service._ler_plano_contas(writer, aplicar_overrides_dre_gerado=True)

    lancamentos = [
        DRELancamentoDB(
            upload_id="u1",
            competencia_ano=2025,
            competencia_mes=6,
            data_lancamento="2025-06-10",
            historico="Recebimento líquido persistido antes do bruto",
            valor_bruto=Decimal("1000.00"),
            credito=Decimal("1000.00"),
            debito=Decimal("0"),
            natureza_raw="1.1.1 - Recebimento de Clientes",
            centro_custo="OBRA A",
            hash_linha="h1",
        ),
        DRELancamentoDB(
            upload_id="u1",
            competencia_ano=2025,
            competencia_mes=6,
            data_lancamento="2025-06-10",
            historico="IR retido",
            valor_bruto=Decimal("300.00"),
            credito=Decimal("0"),
            debito=Decimal("300.00"),
            natureza_raw="IR Retido",
            rubrica="IR Retido",
            centro_custo="OBRA A",
            hash_linha="h2",
        ),
    ]

    linhas, _ = service._agregar_para_apoio(lancamentos, plano)
    por_label = {row[1]: row for row in linhas}

    assert por_label["(=)Receita Bruta"][7] == 1300.0
    assert por_label["(+)Receita Bruta"][7] == 1300.0
    assert por_label["Faturamento"][7] == 1000.0
    assert "IR Retido" not in por_label
    assert "(-)Deduções sobre vendas" not in por_label
    assert por_label["IRPJ"][7] == -300.0
    assert por_label["IRPJ/CSLL"][7] == -300.0
    assert por_label["(-)IRPJ/CSLL"][7] == -300.0


def test_agregar_apoio_aplica_correcoes_do_painel_por_codigo_gerencial():
    """As correções do painel devem prevalecer pelo código, não pelo texto legado."""
    service = _build_service_com_dados([6])
    with TemplateWriter(settings.template_dre_path) as writer:
        plano = service._ler_plano_contas(writer, aplicar_overrides_dre_gerado=True)

    def lancamento(codigo: str, valor: str, *, credito: bool = False) -> DRELancamentoDB:
        return DRELancamentoDB(
            upload_id="u-correcoes",
            competencia_ano=2026,
            competencia_mes=6,
            data_lancamento="2026-06-10",
            historico=f"Lançamento {codigo}",
            credito=Decimal(valor) if credito else Decimal("0"),
            debito=Decimal("0") if credito else Decimal(valor),
            natureza_raw=f"{codigo} - descrição alterável no relatório",
            rubrica=f"{codigo} - descrição alterável no relatório",
            centro_custo="OBRA A",
            hash_linha=f"hash-{codigo}-{valor}-{credito}",
        )

    linhas, _ = service._agregar_para_apoio(
        [
            lancamento("8.5", "85"),
            lancamento("16.1", "161"),
            lancamento("16.3", "163"),
            lancamento("16.4", "164"),
            lancamento("15.5", "155"),
            lancamento("15.7", "157"),
            lancamento("11.17", "117"),
            lancamento("11.18", "118"),
            lancamento("3.20", "320"),
            lancamento("3.21", "321"),
            lancamento("12.19", "1219"),
            lancamento("12.20", "1220"),
            lancamento("11.7", "117"),
            lancamento("6.11", "611"),
            lancamento("6.11", "612", credito=True),
            lancamento("2.2", "22", credito=True),
        ],
        plano,
    )
    por_label = {row[1]: row for row in linhas}

    assert por_label["Aquisição de Maquinas e Equipamentos"][7] == -85.0
    assert por_label["Despesas Não Operacionais"][7] == -606.0
    assert por_label["Manutenção da Sede"][7] == -429.0
    assert por_label["Mão de Obra Terceirizada"][7] == -320.0
    assert por_label["Marketing"][7] == -321.0
    assert por_label["AUXILIO MORADIA"][7] == -1219.0
    assert por_label["VALE TRANSPORTE"][7] == -1220.0
    assert por_label["Financiamento"][7] == -117.0
    assert por_label["(-) Despesas Financeiras"][7] == -117.0
    assert por_label["Compra de veiculos"][7] == -611.0
    assert por_label["Recebimentos Não Operacionais"][7] == 612.0
    assert por_label["Recebimentos de emprestimos"][7] == 22.0


def test_agregar_apoio_classifica_contas_sem_grupo_como_nao_operacionais():
    """Contas sem mapeamento canônico não podem desaparecer do resultado geral."""
    service = _build_service_com_dados([6])
    with TemplateWriter(settings.template_dre_path) as writer:
        plano = service._ler_plano_contas(writer, aplicar_overrides_dre_gerado=True)

    def lancamento(codigo: str, valor: str, *, credito: bool) -> DRELancamentoDB:
        return DRELancamentoDB(
            upload_id="u-nao-operacional",
            competencia_ano=2026,
            competencia_mes=6,
            data_lancamento="2026-06-10",
            historico=f"Lançamento {codigo}",
            credito=Decimal(valor) if credito else Decimal("0"),
            debito=Decimal("0") if credito else Decimal(valor),
            natureza_raw=f"{codigo} - conta sem grupo",
            rubrica=f"{codigo} - conta sem grupo",
            centro_custo="OBRA A",
            hash_linha=f"hash-nao-operacional-{codigo}-{valor}-{credito}",
        )

    linhas, _ = service._agregar_para_apoio(
        [lancamento("99.1", "88", credito=True), lancamento("99.2", "77", credito=False)],
        plano,
    )
    por_label = {row[1]: row for row in linhas}

    assert por_label["Recebimentos Não Operacionais"][7] == 88.0
    assert por_label["Despesas Não Operacionais"][7] == -77.0
    assert por_label["(+/-)Despesas e Recebimentos Não Operacionais"][7] == 11.0


def test_linha_bd_fluxo_classifica_venda_de_veiculo_como_recebimento_nao_operacional():
    service = _build_service_com_dados([6])
    with TemplateWriter(settings.template_dre_path) as writer:
        plano = service._ler_plano_contas(writer, aplicar_overrides_dre_gerado=True)

    venda = DRELancamentoDB(
        upload_id="u-veiculo",
        competencia_ano=2026,
        competencia_mes=6,
        data_lancamento="2026-06-10",
        historico="Venda de veículo",
        credito=Decimal("5000"),
        debito=Decimal("0"),
        natureza_raw="6.11 - COMPRA VENDA VEICULOS",
        rubrica="6.11 - COMPRA VENDA VEICULOS",
        centro_custo="OBRA A",
        hash_linha="hash-venda-veiculo",
    )

    linha = service._converte_linha_bd_fluxo_expandida(venda, plano)

    assert linha[13] == "Recebimentos Não Operacionais"
    assert linha[14] == "(+) Recebimentos Não Operacionais"
    assert linha[15] == "(+/-)Despesas e Recebimentos Não Operacionais"
    assert linha[16] == 5


def test_gerar_arquivo_mostra_correcoes_do_painel_em_linhas_recolhidas(tmp_path):
    service = _build_service_com_dados([6])
    upload_id = service.repository.uploads.list_all()[0].id

    def lancamento(codigo: str, valor: str, *, credito: bool = False) -> DRELancamentoDB:
        return DRELancamentoDB(
            upload_id=upload_id,
            competencia_ano=2025,
            competencia_mes=6,
            data_lancamento="2025-06-20",
            historico=f"DRE corrigida {codigo}",
            credito=Decimal(valor) if credito else Decimal("0"),
            debito=Decimal("0") if credito else Decimal(valor),
            natureza_raw=f"{codigo} - descrição variável",
            rubrica=f"{codigo} - descrição variável",
            centro_custo="OBRA A",
            hash_linha=f"saida-{codigo}-{valor}-{credito}",
        )

    service.repository.lancamentos.create_many(
        [
            lancamento("8.5", "85"),
            lancamento("16.1", "161"),
            lancamento("16.3", "163"),
            lancamento("16.4", "164"),
            lancamento("15.5", "155"),
            lancamento("15.7", "157"),
            lancamento("11.17", "117"),
            lancamento("11.18", "118"),
            lancamento("3.20", "320"),
            lancamento("3.21", "321"),
            lancamento("12.19", "1219"),
            lancamento("12.20", "1220"),
            lancamento("11.7", "117"),
            lancamento("6.11", "611"),
            lancamento("6.11", "612", credito=True),
            lancamento("2.2", "22", credito=True),
        ]
    )
    output_path = tmp_path / "dre_painel_corrigido.xlsx"

    resultado = service.gerar_arquivo("06/2025", output_path=output_path)

    assert resultado["success"] is True
    workbook = load_workbook(output_path, data_only=True)
    ws_dre = workbook["DRE"]
    assert _celula_dre_por_rotulo(ws_dre, "Aquisição de Maquinas e Equipamentos", "N") == -85.0
    assert _celula_dre_por_rotulo(ws_dre, "Despesas Não Operacionais", "N") == -606.0
    assert _celula_dre_por_rotulo(ws_dre, "Manutenção da Sede", "N") == -429.0
    assert _celula_dre_por_rotulo(ws_dre, "Mão de Obra Terceirizada", "N") == -320.0
    assert _celula_dre_por_rotulo(ws_dre, "Marketing", "N") == -321.0
    assert _celula_dre_por_rotulo(ws_dre, "AUXILIO MORADIA", "N") == -1219.0
    assert _celula_dre_por_rotulo(ws_dre, "VALE TRANSPORTE", "N") == -1220.0
    assert _celula_dre_por_rotulo(ws_dre, "Financiamento", "N") == -117.0
    assert _celula_dre_por_rotulo(ws_dre, "(-) Despesas Financeiras", "N") == -117.0
    assert _celula_dre_por_rotulo(ws_dre, "Compra de veiculos", "N") == -611.0
    assert _celula_dre_por_rotulo(ws_dre, "Recebimentos Não Operacionais", "N") == 612.0
    assert _celula_dre_por_rotulo(ws_dre, "Recebimentos de emprestimos", "N") == 22.0

    with zipfile.ZipFile(output_path, "r") as zf:
        nomes = set(zf.namelist())
        assert not any(nome.startswith("xl/slicers/") for nome in nomes)
        assert not any(nome.startswith("xl/slicerCaches/") for nome in nomes)


def test_template_dre_exibe_novas_rubricas_somente_no_nivel_expandido():
    workbook = load_workbook(settings.template_dre_path, data_only=False)
    ws_dre = workbook["DRE"]
    ws_plano = workbook["PLANO_CONTAS"]

    linhas_por_label = {
        str(ws_dre.cell(row, 1).value).strip(): row
        for row in range(1, ws_dre.max_row + 1)
        if ws_dre.cell(row, 1).value
    }
    for label in (
        "AUXILIO MORADIA",
        "VALE TRANSPORTE",
        "Mão de Obra Terceirizada",
        "Manutenção da Sede",
        "Marketing",
    ):
        linha = linhas_por_label[label]
        assert ws_dre.row_dimensions[linha].outlineLevel == 1
        assert ws_dre.row_dimensions[linha].hidden is True
        formula_esperada = (
            f"=IFERROR(VLOOKUP($A{linha},APOIO!$B:$N,"
            "MATCH(B$5,APOIO!$B$5:$N$5,0),FALSE),0)"
        )
        assert ws_dre.cell(linha, 2).value == formula_esperada
    assert next(iter(ws_dre.tables.values())).ref == f"A5:AI{ws_dre.max_row}"

    plano_por_codigo = {
        str(row[0]).split(" - ", 1)[0].strip(): row[1:5]
        for row in ws_plano.iter_rows(min_row=2, max_col=5, values_only=True)
        if row[0]
    }
    assert plano_por_codigo["8.5"] == (
        "Aquisição de Maquinas e Equipamentos",
        "Investimentos",
        "(-)Investimentos",
        8,
    )
    assert plano_por_codigo["11.7"] == (
        "Financiamento",
        "(-) Despesas Financeiras",
        "(+/-)Despesas e Receitas Financeiras",
        6,
    )
    assert plano_por_codigo["15.5"] == (
        "Manutenção da Sede",
        "Despesas Administrativas",
        "(-)Gastos Fixos",
        4,
    )
    assert plano_por_codigo["12.20"] == (
        "VALE TRANSPORTE",
        "Despesas com Pessoal",
        "(-)Gastos Fixos",
        4,
    )
    assert next(iter(ws_plano.tables.values())).ref == f"A1:E{ws_plano.max_row}"
